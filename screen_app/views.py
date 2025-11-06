import math
import time
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
import json

from .models import (Station, Product, ProductMedia, AssemblyStage, 
                    AssemblyProcess, AssemblySession, BillOfMaterial,BOMTemplate, BOMItem, BOMTemplateItem)

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.core.cache import cache


# Bom with pagination 
# Improved BOM Pagination Implementation

import math
import time
from django.core.cache import cache
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import json


class BOMPaginator:
    """Handle BOM pagination logic - FIXED for pre-split data from Station"""
    
    def __init__(self, bom_data, mode='split', items_per_screen=8):
        self.bom_data = bom_data or []
        self.mode = mode  # 'split' or 'single'
        self.items_per_screen = items_per_screen
        self.total_items = len(self.bom_data)
        
        print(f"DEBUG PAGINATOR: Initialized with {self.total_items} items (pre-split data)")
        
        # IMPORTANT: Since Station.get_current_bom_data() already handles display splitting,
        # we don't need to do additional splitting here. Just handle pagination within
        # the data that's already been assigned to this display.
        
        if mode == 'split':
            # For split mode, we're dealing with pre-split data
            # Just paginate the items we received for this specific display
            self.screens_count = 1  # We're only dealing with one display's data
            self.items_per_page = self.items_per_screen
        else:
            # Single mode
            self.screens_count = 1
            self.items_per_page = self.items_per_screen
            
        self.total_pages = math.ceil(self.total_items / self.items_per_page) if self.total_items > 0 else 1
        
        print(f"DEBUG PAGINATOR: Mode={mode}, items_per_page={self.items_per_page}, total_pages={self.total_pages}")
    
    def get_page_data(self, page_number=1, screen_number=1):
        """Get BOM data for specific page - FIXED for pre-split data"""
        print(f"DEBUG PAGINATOR: get_page_data called with page={page_number}, screen={screen_number}")
        print(f"DEBUG PAGINATOR: Working with {len(self.bom_data)} pre-split items")
        
        page_number = max(1, min(page_number, self.total_pages))
        
        # FIXED: Since data is already split by display, just paginate within this display's data
        page_start_idx = (page_number - 1) * self.items_per_page
        page_end_idx = page_start_idx + self.items_per_page
        
        page_items = self.bom_data[page_start_idx:page_end_idx]
        
        print(f"DEBUG PAGINATOR: Page {page_number} contains {len(page_items)} items")
        
        # Create adjusted items with proper serial numbers
        adjusted_items = []
        for i, item in enumerate(page_items):
            adjusted_item = {
                'serial_number': item['serial_number'],  # Keep original serial number from template
                'screen_position': i + 1,  # Position on current page/screen
                'item': item['item'],
                'base_quantity': item['base_quantity'],
                'calculated_quantity': item['calculated_quantity'],
                'formatted_quantity': item['formatted_quantity'],
                'notes': item.get('notes', '')
            }
            adjusted_items.append(adjusted_item)
            
        print(f"DEBUG PAGINATOR: Returning {len(adjusted_items)} adjusted items")
        return adjusted_items
    
    def get_pagination_info(self, current_page=1):
        """Get pagination information - UPDATED"""
        return {
            'current_page': current_page,
            'total_pages': self.total_pages,
            'total_items': self.total_items,  # Items for THIS display only
            'items_per_screen': self.items_per_page,
            'items_per_page': self.items_per_page,
            'screens_count': 1,  # We're handling one display's data
            'mode': self.mode,
            'has_next': current_page < self.total_pages,
            'has_previous': current_page > 1
        }
  



class BOMPaginationManager:
    """Enhanced pagination state manager with stage change detection - COMPLETE REPLACEMENT"""
    
    CACHE_TIMEOUT = 3600  # 1 hour
    
    @classmethod
    def _get_cache_key(cls, product_id, bom_type):
        """Generate cache key for pagination state"""
        return f"bom_pagination_{product_id}_{bom_type}"
    
    @classmethod
    def _get_stage_cache_key(cls, station_id):
        """Generate cache key for tracking current stage/process"""
        return f"station_stage_{station_id}"
    
    @classmethod
    def _get_process_cache_key(cls, station_id):
        """Generate cache key for tracking current process"""
        return f"station_process_{station_id}"
    
    @classmethod
    def get_current_page(cls, product_id, bom_type):
        """Get current page for a product/bom_type combination"""
        cache_key = cls._get_cache_key(product_id, bom_type)
        page = cache.get(cache_key, 1)
        return page
    
    @classmethod
    def set_current_page(cls, product_id, bom_type, page):
        """Set current page for a product/bom_type combination"""
        cache_key = cls._get_cache_key(product_id, bom_type)
        page = max(1, page)  # Ensure page is at least 1
        cache.set(cache_key, page, cls.CACHE_TIMEOUT)
    
    @classmethod
    def reset_pagination(cls, product_id, bom_type):
        """Reset pagination to page 1 for a product/bom_type combination"""
        cache_key = cls._get_cache_key(product_id, bom_type)
        cache.set(cache_key, 1, cls.CACHE_TIMEOUT)
    
    @classmethod
    def clear_product_pagination(cls, product_id):
        """Clear all pagination state for a product"""
        bom_types = [
            'SINGLE_UNIT', 'BATCH_50', 'SUB_ASSEMBLY_1', 'SUB_ASSEMBLY_2', 
            'SUB_ASSEMBLY_3', 'SUB_ASSEMBLY_4', 'SUB_ASSEMBLY_5', 'SUB_ASSEMBLY_6', 
            'FINAL_ASSEMBLY', 'BOM_DISPLAY'
        ]
        cleared_count = 0
        for bom_type in bom_types:
            cache_key = cls._get_cache_key(product_id, bom_type)
            if cache.delete(cache_key):
                cleared_count += 1
        print(f"DEBUG PAGINATION MANAGER: Cleared {cleared_count} pagination states for product {product_id}")
    
    @classmethod
    def check_and_reset_on_stage_change(cls, station):
        """
        Check if stage/process has changed and reset pagination if needed.
        Call this before rendering BOM to ensure fresh pagination on stage changes.
        """
        station_id = station.id
        current_stage = station.current_stage.name if station.current_stage else 'None'
        current_process = getattr(station, 'current_process', None)
        current_process_name = current_process.name if current_process else 'None'
        
        # Get cached values
        cached_stage = cache.get(cls._get_stage_cache_key(station_id))
        cached_process = cache.get(cls._get_process_cache_key(station_id))
        
        stage_changed = cached_stage is not None and cached_stage != current_stage
        process_changed = cached_process is not None and cached_process != current_process_name
        
        if stage_changed or process_changed:
            
            # Clear all pagination for this product
            if station.current_product:
                cls.clear_product_pagination(station.current_product.id)
            
            # Update cached values
            cache.set(cls._get_stage_cache_key(station_id), current_stage, cls.CACHE_TIMEOUT)
            cache.set(cls._get_process_cache_key(station_id), current_process_name, cls.CACHE_TIMEOUT)
            
            return True  # Indicates pagination was reset
        else:
            # Update cached values for first-time tracking
            if cached_stage is None:
                cache.set(cls._get_stage_cache_key(station_id), current_stage, cls.CACHE_TIMEOUT)
            if cached_process is None:
                cache.set(cls._get_process_cache_key(station_id), current_process_name, cls.CACHE_TIMEOUT)
            
            return False  # No reset needed
    
    @classmethod
    def force_reset_station_pagination(cls, station):
        """Force reset pagination for a station (useful for manual resets)"""
        
        if station.current_product:
            cls.clear_product_pagination(station.current_product.id)
        
        # Update stage/process tracking
        current_stage = station.current_stage.name if station.current_stage else 'None'
        current_process = getattr(station, 'current_process', None)
        current_process_name = current_process.name if current_process else 'None'
        
        cache.set(cls._get_stage_cache_key(station.id), current_stage, cls.CACHE_TIMEOUT)
        cache.set(cls._get_process_cache_key(station.id), current_process_name, cls.CACHE_TIMEOUT)
    
    @classmethod
    def get_pagination_status(cls, product_id, bom_type):
        """Get detailed pagination status for debugging"""
        cache_key = cls._get_cache_key(product_id, bom_type)
        current_page = cache.get(cache_key, 1)
        
        return {
            'cache_key': cache_key,
            'current_page': current_page,
            'cache_exists': cache.get(cache_key) is not None
        }
    
    @classmethod
    def next_page(cls, product_id, bom_type, total_pages):
        """Go to next page"""
        current = cls.get_current_page(product_id, bom_type)
        next_page = min(current + 1, total_pages)
        cls.set_current_page(product_id, bom_type, next_page)
        return next_page
    
    @classmethod
    def previous_page(cls, product_id, bom_type):
        """Go to previous page"""
        current = cls.get_current_page(product_id, bom_type)
        prev_page = max(current - 1, 1)
        cls.set_current_page(product_id, bom_type, prev_page)
        return prev_page
    
    @classmethod
    def set_page(cls, product_id, bom_type, target_page, total_pages):
        """Set specific page with bounds checking"""
        target_page = max(1, min(target_page, total_pages))
        cls.set_current_page(product_id, bom_type, target_page)
        return target_page



def debug_process_change(request, station_id):
    """Debug what happens when process changes"""
    station = get_object_or_404(Station, pk=station_id)
    
    # Get current state
    current_state = {
        'station_id': station_id,
        'display_number': station.display_number,
        'current_process': {
            'id': station.current_process.id if station.current_process else None,
            'name': station.current_process.name if station.current_process else None,
            'display_name': station.current_process.display_name if station.current_process else None,
        },
        'current_stage': {
            'id': station.current_stage.id if station.current_stage else None,
            'name': station.current_stage.name if station.current_stage else None,
            'display_name': station.current_stage.display_name if station.current_stage else None,
        },
        'bom_settings': {
            'show_batch_bom': station.show_batch_bom,
            'show_single_unit_bom': station.show_single_unit_bom,
            'product_quantity': station.product_quantity
        }
    }
    
    # Check what BOM data this should produce
    bom_data = station.get_current_bom_data() or []
    current_bom = {
        'item_count': len(bom_data),
        'items': [
            {
                'serial_number': item['serial_number'],
                'item_description': item['item'].item_description,
            }
            for item in bom_data
        ]
    }
    
    # Get cache state
    from django.core.cache import cache
    cache_states = {}
    if station.current_product:
        bom_types = ['SINGLE_UNIT', 'BATCH_50', 'SUB_ASSEMBLY_1', 'SUB_ASSEMBLY_2', 'FINAL_ASSEMBLY', 'BOM_DISPLAY']
        for bom_type in bom_types:
            cache_key = f"bom_pagination_{station.current_product.id}_{bom_type}"
            cache_value = cache.get(cache_key, 'NOT_SET')
            cache_states[bom_type] = cache_value
    
    # Test what different processes would show
    process_scenarios = {}
    
    # Test some other processes
    test_processes = AssemblyProcess.objects.all()[:5]  # Get first 5 processes for testing
    
    for process in test_processes:
        # Temporarily change process
        original_process = station.current_process
        original_stage = station.current_stage
        
        station.current_process = process
        station.current_stage = process.stage
        
        # Get BOM data for this process
        test_bom_data = station.get_current_bom_data() or []
        
        process_scenarios[process.name] = {
            'process_id': process.id,
            'stage_name': process.stage.name if process.stage else None,
            'bom_item_count': len(test_bom_data),
            'items': [
                {
                    'serial_number': item['serial_number'],
                    'item_description': item['item'].item_description,
                }
                for item in test_bom_data[:3]  # Just first 3 for brevity
            ] if test_bom_data else []
        }
        
        # Restore original
        station.current_process = original_process
        station.current_stage = original_stage
    
    return JsonResponse({
        'current_state': current_state,
        'current_bom': current_bom,
        'cache_states': cache_states,
        'process_scenarios': process_scenarios,
        'explanation': {
            'why_same_items': 'If you see the same items when changing processes, it could be because:',
            'reasons': [
                '1. You are staying in the same BOM stage (BOM Display)',
                '2. Both processes use the same BOM template (BATCH_50)',
                '3. Display 3 always gets items 7-9 for this BOM template',
                '4. The process change does not affect which BOM template is used'
            ],
            'when_items_change': 'Items will change when:',
            'change_conditions': [
                '1. You move to a different stage that has its own BOM template',
                '2. You toggle between single unit and batch BOM',
                '3. You change to a process that does not show BOMs'
            ]
        }
    })

def debug_bom_template(request, station_id):
    """Debug BOM template to see what's happening"""
    station = get_object_or_404(Station, pk=station_id)
    
    if not station.current_product:
        return JsonResponse({'error': 'No product selected'}, status=400)
    
    debug_info = {
        'station_id': station_id,
        'display_number': station.display_number,
        'current_product': station.current_product.code,
        'show_batch_bom': station.show_batch_bom,
        'show_single_unit_bom': station.show_single_unit_bom,
        'product_quantity': station.product_quantity
    }
    
    # Check what BOM template exists
    try:
        if station.show_batch_bom:
            bom_template = BOMTemplate.objects.get(
                product=station.current_product,
                bom_type='BATCH_50',
                is_active=True
            )
        elif station.show_single_unit_bom:
            bom_template = BOMTemplate.objects.get(
                product=station.current_product,
                bom_type='SINGLE_UNIT',
                is_active=True
            )
        else:
            return JsonResponse({'error': 'No BOM type selected'}, status=400)
        
        # Get template details
        template_info = {
            'id': bom_template.id,
            'name': bom_template.template_name,
            'bom_type': bom_template.bom_type,
            'should_split': bom_template.should_split_across_displays(),
            'total_items': bom_template.bom_items.filter(is_active=True).count()
        }
        
        # Get all items in the template
        all_template_items = []
        for item in bom_template.bom_items.filter(is_active=True).order_by('serial_number'):
            all_template_items.append({
                'serial_number': item.serial_number,
                'item_code': item.item.item_code,
                'item_description': item.item.item_description,
                'base_quantity': float(item.base_quantity)
            })
        
        # Test split logic for all 3 displays
        split_results = {}
        for display_num in [1, 2, 3]:
            items_for_display = bom_template.get_items_for_display(display_num, station.product_quantity)
            split_results[f'display_{display_num}'] = {
                'item_count': len(items_for_display),
                'items': [
                    {
                        'serial_number': item['serial_number'],
                        'item_description': item['item'].item_description,
                        'calculated_quantity': item['calculated_quantity']
                    }
                    for item in items_for_display
                ]
            }
        
        # Test what station.get_current_bom_data() returns
        station_bom_data = station.get_current_bom_data() or []
        station_result = {
            'item_count': len(station_bom_data),
            'items': [
                {
                    'serial_number': item['serial_number'],
                    'item_description': item['item'].item_description,
                    'calculated_quantity': item['calculated_quantity']
                }
                for item in station_bom_data
            ]
        }
        
        return JsonResponse({
            'debug_info': debug_info,
            'template_info': template_info,
            'all_template_items': all_template_items,
            'split_test_results': split_results,
            'station_bom_result': station_result,
            'summary': {
                'template_has_items': template_info['total_items'],
                'should_split': template_info['should_split'],
                'display_1_gets': split_results['display_1']['item_count'],
                'display_2_gets': split_results['display_2']['item_count'],
                'display_3_gets': split_results['display_3']['item_count'],
                'station_method_returns': station_result['item_count']
            }
        })
        
    except BOMTemplate.DoesNotExist:
        return JsonResponse({
            'error': 'BOM template not found',
            'debug_info': debug_info,
            'suggestion': 'Create a BOM template in admin panel'
        }, status=404)
           
def validate_pagination_params(request):
    """Centralized parameter validation"""
    # Validate page parameter
    try:
        page = int(request.GET.get('page', 1))
        page = max(1, page)
    except (ValueError, TypeError):
        page = 1
    
    # Validate mode parameter
    mode = request.GET.get('mode', 'split')
    if mode not in ['split', 'single']:
        mode = 'split'
    
    # Validate items_per_screen parameter
    try:
        items_per_screen_param = request.GET.get('items_per_screen', '8')
        if items_per_screen_param in ['undefined', 'null', '', None]:
            items_per_screen = 8
        else:
            items_per_screen = int(items_per_screen_param)
            if not (1 <= items_per_screen <= 20):
                items_per_screen = 8
    except (ValueError, TypeError):
        items_per_screen = 8
    
    return page, mode, items_per_screen

def serialize_bom_item(item_data):
    """Serialize BOM item for JSON response"""
    return {
        'serial_number': item_data['serial_number'],
        'screen_position': item_data.get('screen_position', item_data['serial_number']),
        'item_code': item_data['item'].item_code,
        'item_description': item_data['item'].item_description,
        'part_number': item_data['item'].part_number,
        'unit_of_measure': item_data['item'].unit_of_measure,
        'base_quantity': item_data['base_quantity'],
        'calculated_quantity': item_data['calculated_quantity'],
        'formatted_quantity': item_data['formatted_quantity'],
        'notes': item_data['notes'],
        'item_photo_url': item_data['item'].item_photo.url if item_data['item'].item_photo else None,
        'supplier': item_data['item'].supplier,
        'cost_per_unit': float(item_data['item'].cost_per_unit) if item_data['item'].cost_per_unit else None,
    }



def get_station_bom_data_paginated(request, station_id):
    """Get paginated BOM data for a specific station (without get_current_bom_data)."""
    station = get_object_or_404(Station, pk=station_id)
    page_number = request.GET.get('page', 1)

    if not station.current_product:
        return JsonResponse({'error': 'No product selected'}, status=400)

    # Validate parameters
    page, mode, items_per_screen = validate_pagination_params(request)

    # Set defaults
    bom_type = 'UNKNOWN'
    quantity = station.product_quantity or 1
    bom_template = None
    is_stage_specific = False
    bom_data = []

    if station.current_stage:
        bom_type = station.current_stage.name
        is_stage_specific = True
        quantity = station.product_quantity or 1
        try:
            bom_template = station.current_product.bom_templates.get(
                stage=station.current_stage,
                is_active=True
            )
        except:
            try:
                bom_template = station.current_product.bom_templates.get(
                    bom_type=bom_type,
                    is_active=True
                )
            except:
                bom_template = None
    elif station.current_stage and station.current_stage.name == 'BOM_DISPLAY':
        if station.show_single_unit_bom:
            bom_type = 'SINGLE_UNIT'
            quantity = 1
        elif station.show_batch_bom:
            bom_type = 'BATCH_50'
            quantity = station.product_quantity
        try:
            bom_template = station.current_product.bom_templates.get(
                bom_type=bom_type,
                is_active=True
            )
        except:
            bom_template = None

    # If we have a template, generate BOM data from it
    if bom_template:
        full_bom = bom_template.generate_bom_for_quantity(quantity)
        paginator = Paginator(full_bom, 8 * (1 if is_stage_specific else 3))
        try:
            page_obj = paginator.page(page)
            bom_data = page_obj.object_list
        except:
            bom_data = []

    # If no bom_data found, fallback to empty pagination
    if not bom_data:
        return JsonResponse({
            'bom_data': [],
            'pagination': {
                'current_page': page,
                'total_pages': 1,
                'total_items': 0,
                'items_per_screen': 0,
                'items_per_page': 0,
                'screens_count': 1,
                'mode': mode,
                'has_next': False,
                'has_previous': False
            },
            'station_info': {
                'name': station.name,
                'display_number': station.display_number,
                'product_code': station.current_product.code,
                'product_name': station.current_product.name,
                'quantity': quantity,
                'bom_type': bom_type,
                'mode': mode,
                'current_stage': station.current_stage.name if station.current_stage else None,
                'is_stage_specific': is_stage_specific
            },
            'summary': {
                'items_on_screen': 0,
                'total_items': 0,
                'display_context': f"Display {station.display_number} - Page {page}/1"
            },
            'debug_info': {
                'is_stage_specific': is_stage_specific,
                'bom_type': bom_type,
                'template_found': False,
                'stage': station.current_stage.name if station.current_stage else None
            }
        })

    # Calculate pagination info
    if bom_template:
        template_pagination_info = bom_template.get_pagination_info_for_split(quantity=quantity, items_per_screen=8)
        if is_stage_specific and station.display_number != 1:
            # Only Display 1 shows data in stage-specific mode
            bom_data = []
            pagination_info = {
                'current_page': 1,
                'total_pages': 1,
                'total_items': 0,
                'items_per_screen': 0,
                'items_per_page': 0,
                'screens_count': 1,
                'mode': mode,
                'has_next': False,
                'has_previous': False
            }
        else:
            pagination_info = {
                'current_page': page,
                'total_pages': template_pagination_info['total_pages'],
                'total_items': template_pagination_info['total_items'],
                'items_per_screen': 8,
                'items_per_page': template_pagination_info['items_per_page'],
                'screens_count': 1 if is_stage_specific else 3,
                'mode': mode,
                'has_next': page < template_pagination_info['total_pages'],
                'has_previous': page > 1
            }
    else:
        pagination_info = {
            'current_page': page,
            'total_pages': 1,
            'total_items': len(bom_data),
            'items_per_screen': len(bom_data),
            'items_per_page': len(bom_data),
            'screens_count': 1,
            'mode': mode,
            'has_next': False,
            'has_previous': False
        }

    # Assign screen_position to items
    for i, item in enumerate(bom_data):
        if 'screen_position' not in item:
            item['screen_position'] = i + 1

    formatted_bom = [serialize_bom_item(item) for item in bom_data]

    return JsonResponse({
        'bom_data': formatted_bom,
        'pagination': pagination_info,
        'station_info': {
            'name': station.name,
            'display_number': station.display_number,
            'product_code': station.current_product.code,
            'product_name': station.current_product.name,
            'quantity': quantity,
            'bom_type': bom_type,
            'mode': mode,
            'current_stage': station.current_stage.name if station.current_stage else None,
            'is_stage_specific': is_stage_specific
        },
        'summary': {
            'items_on_screen': len(formatted_bom),
            'total_items': pagination_info['total_items'],
            'display_context': f"Display {station.display_number} - Page {page}/{pagination_info['total_pages']}"
        },
        'debug_info': {
            'is_stage_specific': is_stage_specific,
            'bom_type': bom_type,
            'template_found': bom_template is not None,
            'stage': station.current_stage.name if station.current_stage else None
        }
    })


def get_station_media_with_bom_pagination(request, station_id):
    """Get media for a specific station - FIXED TO USE BOMPaginationManager"""
    station = get_object_or_404(Station, pk=station_id)
    
    # Validate pagination parameters
    try:
        mode = request.GET.get('mode', 'split')
        if mode not in ['split', 'single']:
            mode = 'split'
    except:
        mode = 'split'
    
    # CRITICAL FIX: Get current page from BOMPaginationManager instead of always using request parameter
    try:
        # Check if page is explicitly provided in request
        page_param = request.GET.get('page')
        if page_param:
            current_page = int(page_param)
            if current_page < 1:
                current_page = 1
        else:
            # CRITICAL FIX: No page parameter provided, get current page from BOMPaginationManager
            current_page = 1  # Default fallback
            
            if station.current_product and station.current_stage:
                # Find BOM template to get the correct BOM type key
                bom_template = None
                bom_type_key = 'UNKNOWN'
                
                try:
                    # First try stage-specific template
                    bom_template = station.current_product.bom_templates.filter(
                        stage=station.current_stage,
                        is_active=True
                    ).first()
                    
                    # If no stage-specific template, try by BOM type
                    if not bom_template:
                        if hasattr(station, 'show_single_unit_bom') and station.show_single_unit_bom:
                            bom_type_key = 'SINGLE_UNIT'
                        elif hasattr(station, 'show_batch_bom') and station.show_batch_bom:
                            bom_type_key = 'BATCH_50'
                        else:
                            bom_type_key = 'BATCH_50'
                        
                        bom_template = station.current_product.bom_templates.filter(
                            bom_type=bom_type_key,
                            is_active=True
                        ).first()
                    
                    if bom_template:
                        bom_type_key = bom_template.bom_type
                        
                        # CRITICAL FIX: Get current page from BOMPaginationManager
                        current_page = BOMPaginationManager.get_current_page(
                            station.current_product.id, 
                            bom_type_key
                        )
                    else:
                        print(f"DEBUG VIEW: No BOM template found, using default page 1")
                        
                except Exception as e:
                    print(f"DEBUG VIEW: Error getting pagination state: {e}")
            
    except (ValueError, TypeError):
        current_page = 1
    
    try:
        items_per_screen_param = request.GET.get('items_per_screen', '8')
        if items_per_screen_param in ['undefined', 'null', '', None]:
            items_per_screen = 8
        else:
            items_per_screen = int(items_per_screen_param)
            if items_per_screen < 1 or items_per_screen > 20:
                items_per_screen = 8
    except (ValueError, TypeError):
        items_per_screen = 8
    
    
    media_data = []
    
    # FIRST: Try to get BOM data with the CORRECT current page
    display_bom_data = station.get_current_bom_data(page=current_page) or []
    
    
    # DYNAMIC: Check if current BOM template should split across displays
    bom_template = None
    is_splitting_bom_stage = False
    should_show_on_this_display = True
    
    if station.current_product and station.current_stage:
        # Try to find BOM template for current stage/product
        try:
            # First try to find stage-specific BOM template
            bom_template = station.current_product.bom_templates.filter(
                stage=station.current_stage,
                is_active=True
            ).first()
            
            # If no stage-specific template, try to find by BOM type
            if not bom_template:
                # Determine BOM type based on station settings
                if hasattr(station, 'show_single_unit_bom') and station.show_single_unit_bom:
                    bom_type_key = 'SINGLE_UNIT'
                elif hasattr(station, 'show_batch_bom') and station.show_batch_bom:
                    bom_type_key = 'BATCH_50'
                else:
                    # Fallback to BATCH_50 for general BOM stages
                    bom_type_key = 'BATCH_50'
                
                bom_template = station.current_product.bom_templates.filter(
                    bom_type=bom_type_key,
                    is_active=True
                ).first()
            
            if bom_template:
                
                # Check if this BOM should split across displays
                is_splitting_bom_stage = bom_template.should_split_across_displays()
                
                # Check if this display should show this BOM based on template settings
                display_field_name = f'display_screen_{station.display_number}'
                should_show_on_this_display = getattr(bom_template, display_field_name, True)
                
            else:
                print(f"DEBUG VIEW: Display {station.display_number} - No BOM template found")
                
        except Exception as e:
            print(f"DEBUG VIEW: Display {station.display_number} - Error finding BOM template: {e}")
    
    # CRITICAL FIX: For splitting BOM stages, if this display should not show BOM or has no BOM data, return empty media
    if is_splitting_bom_stage:
        
        if not should_show_on_this_display:
            
            pagination_info = {
                'current_page': current_page,
                'total_pages': 1,
                'total_items': 0,
                'items_per_screen': 0,
                'items_per_page': 0,
                'screens_count': 1,
                'mode': mode,
                'has_next': False,
                'has_previous': False
            }
            
            return JsonResponse({
                'media': [],  # EMPTY - BOM disabled for this display
                'station_info': {
                    'name': station.name,
                    'display_number': station.display_number,
                    'current_product': {
                        'id': station.current_product.id,
                        'code': station.current_product.code,
                        'name': station.current_product.name
                    } if station.current_product else None,
                    'current_stage': {
                        'id': station.current_stage.id,
                        'name': station.current_stage.display_name,
                        'order': station.current_stage.order
                    } if station.current_stage else None,
                    'current_process': {
                        'id': station.current_process.id,
                        'name': station.current_process.name,
                        'display_name': station.current_process.display_name,
                        'order': station.current_process.order,
                        'is_looped': station.current_process.is_looped,
                        'loop_group': station.current_process.loop_group
                    } if station.current_process else None,
                    'quantity': station.product_quantity,
                    'loop_mode': station.loop_mode,
                    'clicker_enabled': station.clicker_enabled
                },
                'pagination_info': {
                    'mode': mode,
                    'items_per_screen': 0,
                    'has_bom_data': False,
                    'current_page': current_page
                },
                'debug_info': {
                    'display_number': station.display_number,
                    'current_stage': station.current_stage.name if station.current_stage else None,
                    'bom_items_from_station': 0,
                    'regular_media_count': 0,
                    'total_media_items': 0,
                    'logic': 'Splitting BOM stage - BOM disabled for this display',
                    'is_splitting_bom_stage': True,
                    'should_show_on_this_display': False,
                    'bom_template_found': bom_template is not None,
                    'bom_template_type': bom_template.bom_type if bom_template else None,
                    'reason': 'BOM template disabled for this display'
                }
            })
            
        elif not display_bom_data or len(display_bom_data) == 0:
            
            # Still need to provide pagination info for consistency
            if bom_template:
                template_pagination_info = bom_template.get_pagination_info_for_split(
                    quantity=station.product_quantity or 1, 
                    items_per_screen=8
                )
                pagination_info = {
                    'current_page': current_page,
                    'total_pages': template_pagination_info['total_pages'],
                    'total_items': template_pagination_info['total_items'],
                    'items_per_screen': 8,
                    'items_per_page': template_pagination_info['items_per_page'],
                    'screens_count': 3,
                    'mode': mode,
                    'has_next': current_page < template_pagination_info['total_pages'],
                    'has_previous': current_page > 1
                }
            else:
                pagination_info = {
                    'current_page': current_page,
                    'total_pages': 1,
                    'total_items': 0,
                    'items_per_screen': 0,
                    'items_per_page': 0,
                    'screens_count': 1,
                    'mode': mode,
                    'has_next': False,
                    'has_previous': False
                }
            
            return JsonResponse({
                'media': [],  # EMPTY - no BOM data for this display
                'station_info': {
                    'name': station.name,
                    'display_number': station.display_number,
                    'current_product': {
                        'id': station.current_product.id,
                        'code': station.current_product.code,
                        'name': station.current_product.name
                    } if station.current_product else None,
                    'current_stage': {
                        'id': station.current_stage.id,
                        'name': station.current_stage.display_name,
                        'order': station.current_stage.order
                    } if station.current_stage else None,
                    'current_process': {
                        'id': station.current_process.id,
                        'name': station.current_process.name,
                        'display_name': station.current_process.display_name,
                        'order': station.current_process.order,
                        'is_looped': station.current_process.is_looped,
                        'loop_group': station.current_process.loop_group
                    } if station.current_process else None,
                    'quantity': station.product_quantity,
                    'loop_mode': station.loop_mode,
                    'clicker_enabled': station.clicker_enabled
                },
                'pagination_info': {
                    'mode': mode,
                    'items_per_screen': pagination_info['items_per_screen'],
                    'has_bom_data': False,
                    'current_page': current_page
                },
                'debug_info': {
                    'display_number': station.display_number,
                    'current_stage': station.current_stage.name if station.current_stage else None,
                    'bom_items_from_station': 0,
                    'regular_media_count': 0,
                    'total_media_items': 0,
                    'logic': 'Splitting BOM stage - this display has no BOM items',
                    'is_splitting_bom_stage': True,
                    'should_show_on_this_display': True,
                    'bom_template_found': bom_template is not None,
                    'bom_template_type': bom_template.bom_type if bom_template else None,
                    'reason': 'No BOM data for this display in splitting stage'
                }
            })
        else:
            print(f"DEBUG VIEW: Display {station.display_number} - Splitting BOM stage with {len(display_bom_data)} BOM items - proceeding normally")
    
    # If we reach here, either it's not a splitting BOM stage OR there is BOM data to show
    # Add regular media ONLY if it's NOT a splitting BOM stage OR this display doesn't show BOM
    if not is_splitting_bom_stage or not should_show_on_this_display:
        current_media = station.get_current_media()
        for media in current_media:
            media_info = {
                'id': media.id,
                'url': media.file.url,
                'type': media.file.name.split('.')[-1].lower(),
                'duration': media.duration,
                'media_type': media.get_media_type_display(),
                'product_name': media.product.name,
                'product_code': media.product.code,
            }
            
            if media.process:
                media_info['process'] = {
                    'id': media.process.id,
                    'name': media.process.name,
                    'display_name': media.process.display_name,
                    'stage': media.process.stage.display_name,
                    'order': media.process.order,
                    'is_looped': media.process.is_looped,
                    'loop_group': media.process.loop_group
                }
            
            media_data.append(media_info)
    
    # Process BOM data if available
    if display_bom_data and len(display_bom_data) > 0:
        # Determine BOM type and get the template for duration info
        is_stage_specific = False
        bom_type_key = 'UNKNOWN'
        bom_display_name = 'BOM'
        
        if bom_template:
            bom_type_key = bom_template.bom_type
            is_stage_specific = not bom_template.should_split_across_displays()
            
            if is_stage_specific:
                bom_display_name = f'{station.current_stage.display_name} BOM' if station.current_stage else 'Stage BOM'
            else:
                if bom_template.bom_type == 'SINGLE_UNIT':
                    bom_display_name = 'Single Unit BOM'
                elif bom_template.bom_type == 'BATCH_50':
                    bom_display_name = f'{station.product_quantity} Units BOM'
                    if station.display_number in [2, 3]:
                        bom_display_name += ' (Split)'
                else:
                    bom_display_name = f'{bom_template.get_bom_type_display()} BOM'
        
        # Determine duration from BOM template or use default
        bom_duration = 400  # Default fallback
        is_duration_active = False  # Default fallback
        
        if bom_template:
            bom_duration = bom_template.duration
            is_duration_active = bom_template.is_duration_active
        else:
            print(f"DEBUG VIEW: Display {station.display_number} - No BOM template found, using default duration: {bom_duration}s")
        
        # Get pagination info from template
        if bom_template and is_stage_specific:
            # Stage-specific BOMs: Pagination on Display 1 only (8 items per page)
            if station.display_number == 1:
                template_pagination_info = bom_template.get_pagination_info_for_split(
                    quantity=station.product_quantity or 1, 
                    items_per_screen=8
                )
                pagination_info = {
                    'current_page': current_page,
                    'total_pages': template_pagination_info['total_pages'],
                    'total_items': template_pagination_info['total_items'],
                    'items_per_screen': 8,
                    'items_per_page': 8,
                    'screens_count': 1,
                    'mode': mode,
                    'has_next': current_page < template_pagination_info['total_pages'],
                    'has_previous': current_page > 1
                }
            else:
                # This shouldn't happen for stage-specific BOMs, but handle gracefully
                pagination_info = {
                    'current_page': 1,
                    'total_pages': 1,
                    'total_items': 0,
                    'items_per_screen': 0,
                    'items_per_page': 0,
                    'screens_count': 1,
                    'mode': mode,
                    'has_next': False,
                    'has_previous': False
                }
        elif bom_template:
            # Splitting BOMs: 8 items per display, multiple displays per page
            template_pagination_info = bom_template.get_pagination_info_for_split(
                quantity=station.product_quantity or 1, 
                items_per_screen=8
            )
            pagination_info = {
                'current_page': current_page,
                'total_pages': template_pagination_info['total_pages'],
                'total_items': template_pagination_info['total_items'],
                'items_per_screen': 8,
                'items_per_page': template_pagination_info['items_per_page'],
                'screens_count': 3,
                'mode': mode,
                'has_next': current_page < template_pagination_info['total_pages'],
                'has_previous': current_page > 1
            }
        else:
            # Fallback
            pagination_info = {
                'current_page': current_page,
                'total_pages': 1,
                'total_items': len(display_bom_data),
                'items_per_screen': len(display_bom_data),
                'items_per_page': len(display_bom_data),
                'screens_count': 1,
                'mode': mode,
                'has_next': False,
                'has_previous': False
            }
        
        # Format the data we received from the station
        serialized_bom_data = []
        for i, item_data in enumerate(display_bom_data):
            serialized_item = {
                'serial_number': item_data['serial_number'],
                'screen_position': i + 1,
                'item': {
                    'item_code': item_data['item'].item_code,
                    'item_description': item_data['item'].item_description,
                    'part_number': item_data['item'].part_number,
                    'unit_of_measure': item_data['item'].unit_of_measure,
                    'supplier': item_data['item'].supplier,
                    'cost_per_unit': float(item_data['item'].cost_per_unit) if item_data['item'].cost_per_unit else None,
                    'item_photo_url': item_data['item'].item_photo.url if item_data['item'].item_photo else None,
                },
                'base_quantity': item_data['base_quantity'],
                'calculated_quantity': item_data['calculated_quantity'],
                'formatted_quantity': item_data['formatted_quantity'],
                'notes': item_data['notes']
            }
            serialized_bom_data.append(serialized_item)
        
        # Add BOM as media item with dynamic duration
        bom_media = {
            'id': f'bom_paginated_{station.id}',
            'url': f'/station/{station.id}/bom-render-paginated/',
            'type': 'bom',
            'duration': bom_duration,
            'is_duration_active': is_duration_active,
            'media_type': 'Paginated BOM',
            'is_bom_data': True,
            'bom_items': len(display_bom_data),
            'bom_type': bom_display_name,
            'is_split': not is_stage_specific,
            'display_info': f"Display {station.display_number} - Page {current_page} - {len(display_bom_data)} items",
            'bom_data': serialized_bom_data,
            'pagination': pagination_info,
            'bom_hash': f"{station.current_product.id}_{bom_type_key}_{current_page}_{mode}_{items_per_screen}_{station.display_number}"
        }
        
        # Add BOM at beginning (priority)
        media_data.insert(0, bom_media)
    else:
        # Set default pagination_info if not set
        if 'pagination_info' not in locals():
            pagination_info = {
                'current_page': current_page,
                'total_pages': 1,
                'total_items': 0,
                'items_per_screen': items_per_screen,
                'items_per_page': items_per_screen,
                'screens_count': 1,
                'mode': mode,
                'has_next': False,
                'has_previous': False
            }
    
    return JsonResponse({
        'media': media_data,
        'station_info': {
            'name': station.name,
            'display_number': station.display_number,
            'current_product': {
                'id': station.current_product.id,
                'code': station.current_product.code,
                'name': station.current_product.name
            } if station.current_product else None,
            'current_stage': {
                'id': station.current_stage.id,
                'name': station.current_stage.display_name,
                'order': station.current_stage.order
            } if station.current_stage else None,
            'current_process': {
                'id': station.current_process.id,
                'name': station.current_process.name,
                'display_name': station.current_process.display_name,
                'order': station.current_process.order,
                'is_looped': station.current_process.is_looped,
                'loop_group': station.current_process.loop_group
            } if station.current_process else None,
            'quantity': station.product_quantity,
            'loop_mode': station.loop_mode,
            'clicker_enabled': station.clicker_enabled
        },
        'pagination_info': {
            'mode': mode,
            'items_per_screen': pagination_info['items_per_screen'] if 'pagination_info' in locals() else items_per_screen,
            'has_bom_data': len([m for m in media_data if m.get('is_bom_data')]) > 0,
            'current_page': current_page
        },
        'debug_info': {
            'display_number': station.display_number,
            'current_stage': station.current_stage.name if station.current_stage else None,
            'bom_items_from_station': len(display_bom_data) if 'display_bom_data' in locals() else 0,
            'regular_media_count': len([m for m in media_data if not m.get('is_bom_data')]),
            'total_media_items': len(media_data),
            'logic': 'Dynamic BOM detection - no hardcoded stages',
            'is_stage_specific': 'is_stage_specific' in locals() and is_stage_specific,
            'is_splitting_bom_stage': is_splitting_bom_stage,
            'should_show_on_this_display': should_show_on_this_display,
            'bom_template_found': bom_template is not None,
            'bom_template_type': bom_template.bom_type if bom_template else None
        }
    })
   
   
# ===== BACKEND FIX =====

@csrf_exempt
@require_http_methods(["POST"])
def bom_pagination_control(request, station_id):
    """Handle BOM pagination controls - COMPREHENSIVE FIX"""
    station = get_object_or_404(Station, pk=station_id)
    
    if not station.current_product:
        return JsonResponse({'error': 'No product selected'}, status=400)
    
    try:
        data = json.loads(request.body)
        
        action = data.get('action')
        
        
        # Validate parameters
        mode = data.get('mode', 'split')
        if mode not in ['split', 'single']:
            mode = 'split'
        
        # Find BOM template using same logic as media view
        bom_template = None
        is_splitting_bom_stage = False
        should_show_on_this_display = True
        is_stage_specific = False
        bom_type_key = 'UNKNOWN'
        quantity = station.product_quantity or 1
        
        if station.current_product and station.current_stage:
            try:
                # First try to find stage-specific BOM template
                bom_template = station.current_product.bom_templates.filter(
                    stage=station.current_stage,
                    is_active=True
                ).first()
                
                # If no stage-specific template, try to find by BOM type
                if not bom_template:
                    # Determine BOM type based on station settings
                    if hasattr(station, 'show_single_unit_bom') and station.show_single_unit_bom:
                        bom_type_key = 'SINGLE_UNIT'
                    elif hasattr(station, 'show_batch_bom') and station.show_batch_bom:
                        bom_type_key = 'BATCH_50'
                    else:
                        # Fallback to BATCH_50 for general BOM stages
                        bom_type_key = 'BATCH_50'
                    
                    bom_template = station.current_product.bom_templates.filter(
                        bom_type=bom_type_key,
                        is_active=True
                    ).first()
                
                if bom_template:
                    
                    # Update variables based on template
                    bom_type_key = bom_template.bom_type
                    
                    # Check if this BOM should split across displays
                    is_splitting_bom_stage = bom_template.should_split_across_displays()
                    is_stage_specific = not is_splitting_bom_stage
                    
                    # Check if this display should show this BOM based on template settings
                    display_field_name = f'display_screen_{station.display_number}'
                    should_show_on_this_display = getattr(bom_template, display_field_name, True)
                    
                    # Set quantity based on BOM type
                    if is_stage_specific:
                        quantity = 1  # Stage-specific BOMs always use quantity 1
                    else:
                        quantity = station.product_quantity or 1
                    

                else:
                    print(f"DEBUG PAGINATION CONTROL: No BOM template found")
                    
            except Exception as e:
                print(f"DEBUG PAGINATION CONTROL: Error finding BOM template: {e}")
        
        if not bom_template:
            return JsonResponse({'error': 'No BOM template available'}, status=404)
        
        # Check if pagination is allowed for this display
        if is_splitting_bom_stage and not should_show_on_this_display:
            return JsonResponse({
                'error': 'BOM disabled for this display in splitting BOM stage',
                'current_page': 1,
                'total_pages': 1
            }, status=400)
        
        if is_stage_specific and station.display_number != 1:
            return JsonResponse({
                'error': 'Pagination only available on Display 1 for stage-specific BOMs',
                'current_page': 1,
                'total_pages': 1
            }, status=400)
        
        # Get pagination info from template
        template_pagination_info = bom_template.get_pagination_info_for_split(quantity=quantity, items_per_screen=8)
        total_pages = template_pagination_info['total_pages']
        total_items = template_pagination_info['total_items']
        
        
        # Get current page from cache
        current_page = BOMPaginationManager.get_current_page(station.current_product.id, bom_type_key)
        
        # Handle pagination actions
        if action == 'next_page':
            new_page = min(current_page + 1, total_pages)
        elif action == 'previous_page':
            new_page = max(current_page - 1, 1)
        elif action == 'set_page':
            try:
                target_page = int(data.get('page', 1))
                new_page = max(1, min(target_page, total_pages))
            except (ValueError, TypeError):
                new_page = 1
        elif action == 'first_page':
            new_page = 1
        elif action == 'last_page':
            new_page = total_pages
        else:
            return JsonResponse({'error': 'Invalid action'}, status=400)
        
        
        # Update pagination state in cache
        BOMPaginationManager.set_current_page(station.current_product.id, bom_type_key, new_page)
        
        # Verify cache update
        verify_page = BOMPaginationManager.get_current_page(station.current_product.id, bom_type_key)
        
        # Get the updated BOM data directly from template
        try:
            updated_bom_data = bom_template.get_items_for_display(
                display_number=station.display_number,
                quantity=quantity,
                page=new_page,
                items_per_screen=8
            )
            
            # Debug: Show which items we got
            if updated_bom_data:
                first_item = updated_bom_data[0]
                last_item = updated_bom_data[-1] if len(updated_bom_data) > 1 else first_item
        except Exception as e:
            print(f"DEBUG PAGINATION CONTROL: Error getting data from template: {e}")
            updated_bom_data = []
        
        # Format the updated BOM data
        formatted_bom = []
        for i, item_data in enumerate(updated_bom_data):
            try:
                # Handle both template format and station format
                item_obj = item_data.get('item') if isinstance(item_data, dict) else None
                
                if item_obj:
                    serialized_item = {
                        'serial_number': item_data.get('serial_number', i + 1),
                        'screen_position': i + 1,
                        'item': {
                            'item_code': getattr(item_obj, 'item_code', ''),
                            'item_description': getattr(item_obj, 'item_description', ''),
                            'part_number': getattr(item_obj, 'part_number', ''),
                            'unit_of_measure': getattr(item_obj, 'unit_of_measure', ''),
                            'supplier': getattr(item_obj, 'supplier', ''),
                            'cost_per_unit': float(item_obj.cost_per_unit) if hasattr(item_obj, 'cost_per_unit') and item_obj.cost_per_unit else None,
                            'item_photo_url': item_obj.item_photo.url if hasattr(item_obj, 'item_photo') and item_obj.item_photo else None,
                        },
                        'base_quantity': item_data.get('base_quantity', 1),
                        'calculated_quantity': item_data.get('calculated_quantity', 1),
                        'formatted_quantity': item_data.get('formatted_quantity', '1'),
                        'notes': item_data.get('notes', '')
                    }
                    formatted_bom.append(serialized_item)
                else:
                    print(f"DEBUG PAGINATION CONTROL: Invalid item data at index {i}: {item_data}")
            except Exception as e:
                print(f"DEBUG PAGINATION CONTROL: Error formatting item {i}: {e}")
        
        
        # Create pagination info
        if is_stage_specific:
            pagination_info = {
                'current_page': new_page,
                'total_pages': total_pages,
                'total_items': total_items,
                'items_per_screen': 8,
                'items_per_page': 8,
                'screens_count': 1,
                'mode': mode,
                'has_next': new_page < total_pages,
                'has_previous': new_page > 1
            }
        else:
            pagination_info = {
                'current_page': new_page,
                'total_pages': total_pages,
                'total_items': total_items,
                'items_per_screen': 8,
                'items_per_page': template_pagination_info['items_per_page'],
                'screens_count': 3,
                'mode': mode,
                'has_next': new_page < total_pages,
                'has_previous': new_page > 1
            }
        
        # Create response
        response = {
            'success': True,
            'action': action,
            'current_page': new_page,
            'pagination': pagination_info,
            'bom_data': formatted_bom,
            'summary': {
                'items_on_screen': len(formatted_bom),
                'total_items': total_items,
                'display_context': f"Display {station.display_number} - Page {new_page}/{total_pages}",
                'page_change': f"{current_page} → {new_page}"
            },
            'message': f'Successfully moved to page {new_page} of {total_pages}',
            'debug_info': {
                'bom_type': bom_type_key,
                'is_stage_specific': is_stage_specific,
                'display_number': station.display_number,
                'old_page': current_page,
                'new_page': new_page,
                'cache_verified': verify_page,
                'items_returned': len(formatted_bom),
                'total_pages': total_pages,
                'total_items': total_items,
                'template_found': True
            }
        }
        
        
        return JsonResponse(response)
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        print(f"DEBUG PAGINATION CONTROL ERROR: {str(e)}")
        import traceback
        print(f"DEBUG PAGINATION CONTROL TRACEBACK: {traceback.format_exc()}")
        return JsonResponse({'error': str(e)}, status=500)
    
    
    
def render_bom_fragment_paginated(request, station_id):
    """Render BOM fragment with pagination support - USING MEDIA VIEW LOGIC"""
    station = get_object_or_404(Station, pk=station_id)
    
    # Validate parameters
    try:
        mode = request.GET.get('mode', 'split')
        if mode not in ['split', 'single']:
            mode = 'split'
    except:
        mode = 'split'
    
    try:
        items_per_screen_param = request.GET.get('items_per_screen', '8')
        if items_per_screen_param in ['undefined', 'null', '', None]:
            items_per_screen = 8
        else:
            items_per_screen = int(items_per_screen_param)
            if items_per_screen < 1 or items_per_screen > 20:
                items_per_screen = 8
    except (ValueError, TypeError):
        items_per_screen = 8
    
    try:
        current_page = int(request.GET.get('page', '1'))
        if current_page < 1:
            current_page = 1
    except (ValueError, TypeError):
        current_page = 1
    
    
    if not station.current_product:
        context = {
            'station': station,
            'bom_data': [],
            'bom_type': 'No Product Selected',
            'quantity': 0,
            'duration': 400,
            'is_duration_active': False,
            'pagination': {
                'current_page': 1,
                'total_pages': 1,
                'total_items': 0,
                'items_per_screen': 0,
                'items_per_page': 0,
                'screens_count': 1,
                'mode': mode,
                'has_next': False,
                'has_previous': False
            },
            'mode': mode,
            'is_split': False,
            'bom_info': {},
            'debug_info': {
                'message': 'No product selected'
            }
        }
        return render(request, 'bom_slider_fragment_paginated.html', context)
    
    # ENHANCED: Check for stage/process changes and reset pagination if needed
    pagination_was_reset = BOMPaginationManager.check_and_reset_on_stage_change(station)
    if pagination_was_reset:
        current_page = 1  # Reset to page 1 if pagination was cleared
    
    # FIXED: Check if we need to reset pagination due to manual reset flag
    should_reset_pagination = request.GET.get('reset_pagination', 'false').lower() == 'true'
    if should_reset_pagination:
        BOMPaginationManager.force_reset_station_pagination(station)
        current_page = 1
    
    # USING MEDIA VIEW LOGIC: Get BOM data from station method with current page
    display_bom_data = station.get_current_bom_data(page=current_page) or []
    
    
    # USING MEDIA VIEW LOGIC: Find BOM template using same logic as media view
    bom_template = None
    is_splitting_bom_stage = False
    should_show_on_this_display = True
    is_stage_specific = False
    bom_type_key = 'UNKNOWN'
    bom_type_display = 'Unknown BOM'
    quantity = station.product_quantity or 1
    
    if station.current_product and station.current_stage:
        try:
            # First try to find stage-specific BOM template
            bom_template = station.current_product.bom_templates.filter(
                stage=station.current_stage,
                is_active=True
            ).first()
            
            # If no stage-specific template, try to find by BOM type
            if not bom_template:
                # Determine BOM type based on station settings
                if hasattr(station, 'show_single_unit_bom') and station.show_single_unit_bom:
                    bom_type_key = 'SINGLE_UNIT'
                elif hasattr(station, 'show_batch_bom') and station.show_batch_bom:
                    bom_type_key = 'BATCH_50'
                else:
                    # Fallback to BATCH_50 for general BOM stages
                    bom_type_key = 'BATCH_50'
                
                bom_template = station.current_product.bom_templates.filter(
                    bom_type=bom_type_key,
                    is_active=True
                ).first()
            
            if bom_template:
                
                # Update variables based on template
                bom_type_key = bom_template.bom_type
                bom_type_display = bom_template.get_bom_type_display()
                
                # Check if this BOM should split across displays
                is_splitting_bom_stage = bom_template.should_split_across_displays()
                is_stage_specific = not is_splitting_bom_stage
                
                # Check if this display should show this BOM based on template settings
                display_field_name = f'display_screen_{station.display_number}'
                should_show_on_this_display = getattr(bom_template, display_field_name, True)
                
                # Set quantity based on BOM type
                if is_stage_specific:
                    quantity = 1  # Stage-specific BOMs always use quantity 1
                else:
                    quantity = station.product_quantity or 1
                
            else:
                print(f"DEBUG RENDER FRAGMENT: No BOM template found")
                
        except Exception as e:
            print(f"DEBUG RENDER FRAGMENT: Error finding BOM template: {e}")
    
    # Update pagination cache with current page for this BOM type
    BOMPaginationManager.set_current_page(station.current_product.id, bom_type_key, current_page)
    
    # USING MEDIA VIEW LOGIC: Handle splitting BOM stages where this display shouldn't show BOM
    if is_splitting_bom_stage and not should_show_on_this_display:
        context = {
            'station': station,
            'bom_data': [],
            'bom_type': 'BOM disabled for this display',
            'quantity': 0,
            'duration': 400,
            'is_duration_active': False,
            'pagination': {
                'current_page': 1,
                'total_pages': 1,
                'total_items': 0,
                'items_per_screen': 0,
                'items_per_page': 0,
                'screens_count': 1,
                'mode': mode,
                'has_next': False,
                'has_previous': False
            },
            'mode': mode,
            'is_split': False,
            'bom_info': {},
            'debug_info': {
                'display_number': station.display_number,
                'stage': station.current_stage.name if station.current_stage else 'None',
                'message': 'Splitting BOM stage - BOM disabled for this display',
                'is_splitting_bom_stage': True,
                'should_show_on_this_display': False,
                'bom_template_found': bom_template is not None,
                'bom_type_key': bom_type_key
            }
        }
        return render(request, 'bom_slider_fragment_paginated.html', context)
    
    # USING MEDIA VIEW LOGIC: Handle case where splitting BOM stage has no data for this display
    if is_splitting_bom_stage and (not display_bom_data or len(display_bom_data) == 0):
        
        # Get pagination info for consistency
        if bom_template:
            template_pagination_info = bom_template.get_pagination_info_for_split(
                quantity=quantity, 
                items_per_screen=8
            )
            pagination_info = {
                'current_page': current_page,
                'total_pages': template_pagination_info['total_pages'],
                'total_items': template_pagination_info['total_items'],
                'items_per_screen': 8,
                'items_per_page': template_pagination_info['items_per_page'],
                'screens_count': 3,
                'mode': mode,
                'has_next': current_page < template_pagination_info['total_pages'],
                'has_previous': current_page > 1
            }
        else:
            pagination_info = {
                'current_page': current_page,
                'total_pages': 1,
                'total_items': 0,
                'items_per_screen': 0,
                'items_per_page': 0,
                'screens_count': 1,
                'mode': mode,
                'has_next': False,
                'has_previous': False
            }
        
        context = {
            'station': station,
            'bom_data': [],
            'bom_type': 'No BOM data for this display',
            'quantity': 0,
            'duration': 400,
            'is_duration_active': False,
            'pagination': pagination_info,
            'mode': mode,
            'is_split': is_splitting_bom_stage,
            'bom_info': {},
            'debug_info': {
                'display_number': station.display_number,
                'stage': station.current_stage.name if station.current_stage else 'None',
                'message': 'Splitting BOM stage - no BOM data for this display',
                'is_splitting_bom_stage': True,
                'should_show_on_this_display': True,
                'bom_template_found': bom_template is not None,
                'bom_type_key': bom_type_key,
                'current_page': current_page
            }
        }
        return render(request, 'bom_slider_fragment_paginated.html', context)
    
    print(f"DEBUG RENDER FRAGMENT: Processing {len(display_bom_data)} BOM items")
    
    # Get duration info from template
    bom_duration = 400  # Default fallback
    is_duration_active = False  # Default fallback
    
    if bom_template:
        bom_duration = bom_template.duration
        is_duration_active = bom_template.is_duration_active
        print(f"DEBUG RENDER FRAGMENT: Using BOM template duration: {bom_duration}s, is_duration_active: {is_duration_active}")
    else:
        print(f"DEBUG RENDER FRAGMENT: No BOM template found, using default duration: {bom_duration}s")
    
    # USING MEDIA VIEW LOGIC: Get pagination info from template
    if bom_template and is_stage_specific:
        # Stage-specific BOMs: Pagination on Display 1 only (8 items per page)
        if station.display_number == 1:
            template_pagination_info = bom_template.get_pagination_info_for_split(
                quantity=quantity, 
                items_per_screen=8
            )
            pagination_info = {
                'current_page': current_page,
                'total_pages': template_pagination_info['total_pages'],
                'total_items': template_pagination_info['total_items'],
                'items_per_screen': 8,
                'items_per_page': 8,
                'screens_count': 1,
                'mode': mode,
                'has_next': current_page < template_pagination_info['total_pages'],
                'has_previous': current_page > 1
            }
        else:
            # This shouldn't happen for stage-specific BOMs, but handle gracefully
            pagination_info = {
                'current_page': 1,
                'total_pages': 1,
                'total_items': 0,
                'items_per_screen': 0,
                'items_per_page': 0,
                'screens_count': 1,
                'mode': mode,
                'has_next': False,
                'has_previous': False
            }
    elif bom_template:
        # Splitting BOMs: 8 items per display, multiple displays per page
        template_pagination_info = bom_template.get_pagination_info_for_split(
            quantity=quantity, 
            items_per_screen=8
        )
        pagination_info = {
            'current_page': current_page,
            'total_pages': template_pagination_info['total_pages'],
            'total_items': template_pagination_info['total_items'],
            'items_per_screen': 8,
            'items_per_page': template_pagination_info['items_per_page'],
            'screens_count': 3,
            'mode': mode,
            'has_next': current_page < template_pagination_info['total_pages'],
            'has_previous': current_page > 1
        }
    else:
        # Fallback
        pagination_info = {
            'current_page': current_page,
            'total_pages': 1,
            'total_items': len(display_bom_data),
            'items_per_screen': len(display_bom_data),
            'items_per_page': len(display_bom_data),
            'screens_count': 1,
            'mode': mode,
            'has_next': False,
            'has_previous': False
        }
    
    # USING MEDIA VIEW LOGIC: Format BOM data (convert from station format to template format)
    formatted_bom_data = []
    for i, item_data in enumerate(display_bom_data):
        # Convert from station format to the format expected by template
        formatted_item = {
            'serial_number': item_data['serial_number'],
            'screen_position': i + 1,
            'item': item_data['item'],  # Keep the actual item object
            'base_quantity': item_data['base_quantity'],
            'calculated_quantity': item_data['calculated_quantity'],
            'formatted_quantity': item_data['formatted_quantity'],
            'notes': item_data['notes']
        }
        formatted_bom_data.append(formatted_item)
    
    
    # Create bom_info structure
    bom_info = {
        'bom_type': bom_type_key,
        'display_name': bom_type_display,
        'quantity': quantity,
        'total_items': pagination_info['total_items'],
        'current_page': current_page,
        'total_pages': pagination_info['total_pages']
    }
    
    context = {
        'station': station,
        'bom_data': formatted_bom_data,  # Use formatted data
        'bom_type': bom_type_display,
        'quantity': quantity,
        'duration': bom_duration,
        'is_duration_active': is_duration_active,
        'pagination': pagination_info,
        'mode': mode,
        'is_split': is_splitting_bom_stage,  # Use the splitting logic from media view
        'bom_info': bom_info,
        'debug_info': {
            'received_params': {
                'mode': request.GET.get('mode'),
                'items_per_screen': request.GET.get('items_per_screen'),
                'reset_pagination': request.GET.get('reset_pagination'),
                'page': request.GET.get('page')
            },
            'processed_params': {
                'mode': mode,
                'items_per_screen': items_per_screen,
                'current_page': current_page
            },
            'current_page': current_page,
            'total_pages': pagination_info['total_pages'],
            'screen_items': len(formatted_bom_data),
            'bom_type_key': bom_type_key,
            'stage': station.current_stage.name if station.current_stage else 'None',
            'is_stage_specific': is_stage_specific,
            'is_splitting_bom_stage': is_splitting_bom_stage,
            'should_show_on_this_display': should_show_on_this_display,
            'template_found': bom_template is not None,
            'display_number': station.display_number,
            'quantity': quantity,
            'template_id': bom_template.id if bom_template else None,
            'pagination_was_reset': pagination_was_reset,
            'logic': 'Using media view logic for BOM template detection and data retrieval'
        }
    }
    
    return render(request, 'bom_slider_fragment_paginated.html', context)
    

def get_station_bom_data(request, station_id):
    """Get database BOM data for a specific station with split logic - FIXED"""
    station = get_object_or_404(Station, pk=station_id)
    
    if not station.current_product:
        return JsonResponse({'error': 'No product selected'}, status=400)
    
    # Get BOM data based on station settings and display number
    bom_data = station.get_current_bom_data()
    
    if not bom_data:
        return JsonResponse({'error': 'No BOM data available'}, status=404)
    
    # Get BOM info for display context
    bom_info = station.get_current_bom_info()
    
    # Format BOM data for frontend
    formatted_bom = []
    for item_data in bom_data:
        item_info = {
            'serial_number': item_data['serial_number'],
            'item_code': item_data['item'].item_code,
            'item_description': item_data['item'].item_description,
            'part_number': item_data['item'].part_number,
            'unit_of_measure': item_data['item'].unit_of_measure,
            'base_quantity': item_data['base_quantity'],
            'calculated_quantity': item_data['calculated_quantity'],
            'formatted_quantity': item_data['formatted_quantity'],
            'notes': item_data['notes'],
            'item_photo_url': item_data['item'].item_photo.url if item_data['item'].item_photo else None,
            'supplier': item_data['item'].supplier,
            'cost_per_unit': float(item_data['item'].cost_per_unit) if item_data['item'].cost_per_unit else None,
        }
        formatted_bom.append(item_info)
    
    # Determine BOM type and quantity
    if station.show_single_unit_bom:
        bom_type = 'SINGLE_UNIT'
        quantity = 1
    elif station.show_batch_bom:
        bom_type = 'BATCH_50'
        quantity = station.product_quantity
    else:
        bom_type = station.current_stage.name if station.current_stage else None
        quantity = station.product_quantity
    
    # FIXED: Get BOM template information without model objects
    bom_template_info = None
    try:
        bom_template = BOMTemplate.objects.get(
            product=station.current_product,
            bom_type=bom_type,
            is_active=True
        )
        bom_template_info = {
            'id': bom_template.id,
            'name': bom_template.template_name,
            'bom_type': bom_template.get_bom_type_display(),
            'stage': bom_template.stage.display_name if bom_template.stage else None,
            'description': bom_template.description,
            'is_split': bom_info['is_split'] if bom_info else False,
            'display_info': bom_info['display_info'] if bom_info else None,
            'split_info': bom_info.get('split_info') if bom_info else None,
        }
    except BOMTemplate.DoesNotExist:
        pass
    
    # FIXED: Serialize bom_info properly
    serialized_bom_info = {}
    if bom_info:
        for key, value in bom_info.items():
            if hasattr(value, 'id'):  # It's a model object
                if hasattr(value, 'template_name'):  # BOMTemplate
                    serialized_bom_info[key] = {
                        'id': value.id,
                        'name': value.template_name,
                        'bom_type': value.bom_type,
                    }
                elif hasattr(value, 'display_name'):  # AssemblyStage
                    serialized_bom_info[key] = {
                        'id': value.id,
                        'name': value.name,
                        'display_name': value.display_name,
                    }
                else:
                    serialized_bom_info[key] = str(value)
            else:
                serialized_bom_info[key] = value
    
    return JsonResponse({
        'bom_data': formatted_bom,
        'bom_template': bom_template_info,
        'station_info': {
            'name': station.name,
            'display_number': station.display_number,
            'product_code': station.current_product.code,
            'product_name': station.current_product.name,
            'quantity': quantity,
            'bom_type': bom_type,
            'is_split_bom': bom_info['is_split'] if bom_info else False,
            'display_info': bom_info['display_info'] if bom_info else None,
            'bom_info': serialized_bom_info,  # FIXED: Use serialized version
        },
        'summary': {
            'total_items': len(formatted_bom),
            'total_cost': sum(
                item['calculated_quantity'] * (item['cost_per_unit'] or 0) 
                for item in formatted_bom
            ),
            'display_context': f"Display {station.display_number}" + (
                f" - {bom_info['display_info']}" if bom_info and bom_info.get('display_info') else ""
            )
        }
    })
    
    

def render_bom_display(request, station_id):
    """Render BOM display page for a station with split logic"""
    station = get_object_or_404(Station, pk=station_id)
    
    # Get BOM data for this specific display
    bom_data = station.get_current_bom_data() if station.current_product else []
    bom_info = station.get_current_bom_info()
    
    # Determine BOM type and quantity
    if station.show_single_unit_bom:
        bom_type = 'Single Unit'
        if bom_info and bom_info['is_split']:
            bom_type += f" (Display {station.display_number})"
        quantity = 1
    elif station.show_batch_bom:
        bom_type = f'{station.product_quantity} Units Batch'
        if bom_info and bom_info['is_split']:
            bom_type += f" (Display {station.display_number})"
        quantity = station.product_quantity
    else:
        bom_type = station.current_stage.display_name if station.current_stage else 'Unknown'
        quantity = station.product_quantity
    
    # Add split information to context
    split_context = {}
    if bom_info and bom_info.get('is_split'):
        split_context = {
            'is_split': True,
            'display_info': bom_info.get('display_info', ''),
            'split_info': bom_info.get('split_info', {}),
            'current_display': station.display_number
        }
    
    context = {
        'station': station,
        'bom_data': bom_data,
        'bom_type': bom_type,
        'quantity': quantity,
        'product': station.current_product,
        'stage': station.current_stage,
        'process': station.current_process,
        'split_context': split_context,
    }
    
    return render(request, 'bom_display.html', context)

@csrf_exempt
@require_http_methods(["POST"])
def update_bom_settings(request, station_id):
    """Update BOM display settings for a station"""
    station = get_object_or_404(Station, pk=station_id)
    
    try:
        data = json.loads(request.body)
        
        # Update BOM display settings
        if 'show_single_unit_bom' in data:
            station.show_single_unit_bom = data['show_single_unit_bom']
        if 'show_batch_bom' in data:
            station.show_batch_bom = data['show_batch_bom']
        if 'product_quantity' in data:
            quantity = int(data['product_quantity'])
            if quantity > 0:
                station.product_quantity = quantity
        
        station.save()
        
        # Get updated BOM data
        bom_data = station.get_current_bom_data()
        
        return JsonResponse({
            'success': True,
            'message': 'BOM settings updated successfully',
            'current_settings': {
                'show_single_unit_bom': station.show_single_unit_bom,
                'show_batch_bom': station.show_batch_bom,
                'product_quantity': station.product_quantity,
            },
            'bom_item_count': len(bom_data) if bom_data else 0,
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except ValueError:
        return JsonResponse({'error': 'Invalid quantity value'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_bom_templates(request):
    """Get all available BOM templates"""
    templates = BOMTemplate.objects.filter(is_active=True).select_related('product', 'stage')
    
    templates_data = []
    for template in templates:
        template_info = {
            'id': template.id,
            'name': template.template_name,
            'product_code': template.product.code,
            'product_name': template.product.name,
            'bom_type': template.bom_type,
            'bom_type_display': template.get_bom_type_display(),
            'stage': template.stage.name if template.stage else None,
            'stage_display': template.stage.display_name if template.stage else None,
            'description': template.description,
            'item_count': template.bom_items.filter(is_active=True).count(),
            'displays': {
                'screen_1': template.display_screen_1,
                'screen_2': template.display_screen_2,
                'screen_3': template.display_screen_3,
            },
            'created_date': template.created_date.isoformat(),
        }
        templates_data.append(template_info)
    
    return JsonResponse({
        'templates': templates_data,
        'total_count': len(templates_data),
    })

def preview_bom_template(request, template_id):
    """Preview BOM template for different quantities"""
    template = get_object_or_404(BOMTemplate, pk=template_id)
    quantity = int(request.GET.get('quantity', 1))
    
    # Generate BOM data for the specified quantity
    bom_data = template.generate_bom_for_quantity(quantity)
    
    # Format for JSON response
    formatted_bom = []
    total_cost = 0
    
    for item_data in bom_data:
        item_cost = (item_data['calculated_quantity'] * 
                    (item_data['item'].cost_per_unit or 0))
        total_cost += item_cost
        
        item_info = {
            'serial_number': item_data['serial_number'],
            'item_code': item_data['item'].item_code,
            'item_description': item_data['item'].item_description,
            'part_number': item_data['item'].part_number,
            'unit_of_measure': item_data['item'].unit_of_measure,
            'base_quantity': item_data['base_quantity'],
            'calculated_quantity': item_data['calculated_quantity'],
            'formatted_quantity': item_data['formatted_quantity'],
            'notes': item_data['notes'],
            'cost_per_unit': float(item_data['item'].cost_per_unit) if item_data['item'].cost_per_unit else None,
            'line_cost': float(item_cost),
            'supplier': item_data['item'].supplier,
            'item_photo_url': item_data['item'].item_photo.url if item_data['item'].item_photo else None,
        }
        formatted_bom.append(item_info)
    
    return JsonResponse({
        'template': {
            'id': template.id,
            'name': template.template_name,
            'bom_type': template.get_bom_type_display(),
            'product': {
                'code': template.product.code,
                'name': template.product.name,
            },
            'stage': template.stage.display_name if template.stage else None,
            'description': template.description,
        },
        'quantity': quantity,
        'bom_data': formatted_bom,
        'summary': {
            'total_items': len(formatted_bom),
            'total_cost': float(total_cost),
            'unique_suppliers': len(set(
                item['supplier'] for item in formatted_bom 
                if item['supplier']
            )),
        }
    })

def get_bom_items(request):
    """Get all BOM items with search and filtering"""
    items = BOMItem.objects.filter(is_active=True)
    
    # Search functionality
    search = request.GET.get('search', '').strip()
    if search:
        items = items.filter(
            item_description__icontains=search
        ) | items.filter(
            item_code__icontains=search
        ) | items.filter(
            part_number__icontains=search
        )
    
    # Filter by unit
    unit_filter = request.GET.get('unit')
    if unit_filter:
        items = items.filter(unit_of_measure=unit_filter)
    
    # Filter by supplier
    supplier_filter = request.GET.get('supplier')
    if supplier_filter:
        items = items.filter(supplier__icontains=supplier_filter)
    
    # Pagination
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 50))
    start = (page - 1) * per_page
    end = start + per_page
    
    total_count = items.count()
    items_page = items[start:end]
    
    items_data = []
    for item in items_page:
        item_info = {
            'id': item.id,
            'item_code': item.item_code,
            'item_description': item.item_description,
            'part_number': item.part_number,
            'unit_of_measure': item.unit_of_measure,
            'supplier': item.supplier,
            'cost_per_unit': float(item.cost_per_unit) if item.cost_per_unit else None,
            'weight_per_unit': float(item.weight_per_unit) if item.weight_per_unit else None,
            'item_photo_url': item.item_photo.url if item.item_photo else None,
            'created_date': item.created_date.isoformat(),
            'is_active': item.is_active,
        }
        items_data.append(item_info)
    
    return JsonResponse({
        'items': items_data,
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total_count': total_count,
            'total_pages': (total_count + per_page - 1) // per_page,
            'has_next': end < total_count,
            'has_prev': page > 1,
        },
        'filters': {
            'search': search,
            'unit': unit_filter,
            'supplier': supplier_filter,
        }
    })




#  For station 
def get_station_media(request, station_id):
    """Get media for a specific station based on current assembly state"""
    station = get_object_or_404(Station, pk=station_id)
    
    # Get media based on current assembly state and display number
    current_media = station.get_current_media()
    
    media_data = []
    for media in current_media:
        media_info = {
            'id': media.id,
            'url': media.file.url,
            'type': media.file.name.split('.')[-1].lower(),
            'duration': media.duration,
            'media_type': media.get_media_type_display(),
            'product_name': media.product.name,
            'product_code': media.product.code,
        }
        
        # Add process info if available
        if media.process:
            media_info['process'] = {
                'id': media.process.id,
                'name': media.process.name,
                'display_name': media.process.display_name,
                'stage': media.process.stage.display_name,
                'order': media.process.order,
                'is_looped': media.process.is_looped,
                'loop_group': media.process.loop_group
            }
        
        # Add BOM info if available
        if media.bom:
            media_info['bom'] = {
                'id': media.bom.id,
                'type': media.bom.get_bom_type_display(),
                'stage': media.bom.stage.display_name if media.bom.stage else None
            }
        
        media_data.append(media_info)
    
    return JsonResponse({
        'media': media_data,
        'station_info': {
            'name': station.name,
            'display_number': station.display_number,
            'current_product': {
                'id': station.current_product.id,
                'code': station.current_product.code,
                'name': station.current_product.name
            } if station.current_product else None,
            'current_stage': {
                'id': station.current_stage.id,
                'name': station.current_stage.display_name,
                'order': station.current_stage.order
            } if station.current_stage else None,
            'current_process': {
                'id': station.current_process.id,
                'name': station.current_process.name,
                'display_name': station.current_process.display_name,
                'order': station.current_process.order,
                'is_looped': station.current_process.is_looped,
                'loop_group': station.current_process.loop_group
            } if station.current_process else None,
            'quantity': station.product_quantity,
            'loop_mode': station.loop_mode,
            'clicker_enabled': station.clicker_enabled
        }
    })

def station_media_slider(request, station_id):
    """Render the media slider for a station"""
    station = get_object_or_404(Station, pk=station_id)
    request.session['last_station_id'] = station_id

    # Get current media for this display
    current_media = station.get_current_media()
    
    context = {
        'station': station,
        'selected_media': current_media,
        'current_product': station.current_product,
        'current_stage': station.current_stage,
        'current_process': station.current_process,
        'product_quantity': station.product_quantity,
        'clicker_enabled': station.clicker_enabled,
        'loop_mode': station.loop_mode,
        'next_process': station.get_next_process(),
        'previous_process': station.get_previous_process(),
        'display_number': station.display_number,
    }
    
    return render(request, 'brg_station_slider.html', context)



def station_media_stream(request, station_id):
    """Real-time streaming for BRG assembly workflow"""
    def event_stream():
        last_update = None
        while True:
            try:
                station = get_object_or_404(Station, pk=station_id)
                current_media = station.get_current_media()
                
                # Create comprehensive comparison data
                media_comparison = []
                for media in current_media:
                    comparison_item = {
                        'id': media.id,
                        'file_url': media.file.url,
                        'media_type': media.media_type,
                        'process_id': media.process.id if media.process else None,
                        'bom_id': media.bom.id if media.bom else None
                    }
                    media_comparison.append(comparison_item)
                
                # Include assembly state in comparison
                assembly_state = {
                    'current_product_id': station.current_product.id if station.current_product else None,
                    'current_stage_id': station.current_stage.id if station.current_stage else None,
                    'current_process_id': station.current_process.id if station.current_process else None,
                    'product_quantity': station.product_quantity,
                    'loop_mode': station.loop_mode,
                    'display_number': station.display_number,
                }
                
                current_state = {
                    'media': media_comparison,
                    'assembly': assembly_state
                }
                
                # Only send update if something changed
                if current_state != last_update:
                    # Prepare detailed media data for frontend
                    media_data = []
                    for media in current_media:
                        media_info = {
                            'id': media.id,
                            'url': media.file.url,
                            'type': media.file.name.split('.')[-1].lower(),
                            'duration': media.duration,
                            'media_type': media.get_media_type_display(),
                            'product_name': media.product.name,
                            'product_code': media.product.code,
                        }
                        
                        if media.process:
                            media_info['process'] = {
                                'id': media.process.id,
                                'name': media.process.name,
                                'display_name': media.process.display_name,
                                'stage': media.process.stage.display_name,
                                'order': media.process.order,
                                'is_looped': media.process.is_looped,
                                'loop_group': media.process.loop_group
                            }
                        
                        if media.bom:
                            media_info['bom'] = {
                                'id': media.bom.id,
                                'type': media.bom.get_bom_type_display(),
                                'stage': media.bom.stage.display_name if media.bom.stage else None
                            }
                        
                        media_data.append(media_info)
                    
                    # Prepare assembly info
                    assembly_info = {
                        'current_product': {
                            'id': station.current_product.id,
                            'code': station.current_product.code,
                            'name': station.current_product.name
                        } if station.current_product else None,
                        'current_stage': {
                            'id': station.current_stage.id,
                            'name': station.current_stage.display_name,
                            'order': station.current_stage.order
                        } if station.current_stage else None,
                        'current_process': {
                            'id': station.current_process.id,
                            'name': station.current_process.name,
                            'display_name': station.current_process.display_name,
                            'order': station.current_process.order,
                            'is_looped': station.current_process.is_looped,
                            'loop_group': station.current_process.loop_group
                        } if station.current_process else None,
                        'quantity': station.product_quantity,
                        'loop_mode': station.loop_mode,
                        'clicker_enabled': station.clicker_enabled,
                        'display_number': station.display_number,
                        'next_process': {
                            'id': station.get_next_process().id,
                            'name': station.get_next_process().name,
                            'display_name': station.get_next_process().display_name
                        } if station.get_next_process() else None,
                        'previous_process': {
                            'id': station.get_previous_process().id,  
                            'name': station.get_previous_process().name,
                            'display_name': station.get_previous_process().display_name
                        } if station.get_previous_process() else None,
                    }
                    
                    response_data = {
                        'media': media_data,
                        'station_name': station.name,
                        'assembly': assembly_info,
                        'timestamp': time.time()
                    }
                    
                    last_update = current_state
                    yield f"data: {json.dumps(response_data)}\n\n"
                
                time.sleep(1)  # Check every 3 seconds for BRG workflow
                
            except Station.DoesNotExist:
                yield f"data: {json.dumps({'error': 'Station not found'})}\n\n"
                break
            except Exception as e:
                yield f"data: {json.dumps({'error': str(e)})}\n\n"
                time.sleep(10)

    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['Connection'] = 'keep-alive'
    response['X-Accel-Buffering'] = 'no'
    return response

# UPDATED: Add pagination reset to your clicker_action function

@csrf_exempt
@require_http_methods(["POST"])
def clicker_action(request, station_id):
    station = get_object_or_404(Station, pk=station_id)

    if not station.clicker_enabled:
        return JsonResponse({'error': 'Clicker not enabled for this station'}, status=400)

    try:
        data = json.loads(request.body)
        action = data.get('action')  # 'forward', 'backward', 'toggle_loop'

        if action == 'forward':
            next_process = station.get_next_process()
            exit_loop_mode = False
            enter_loop_mode = False
            current_proc = station.current_process

            if (station.loop_mode and 
                current_proc and 
                current_proc.loop_group == 'final_assembly_1abc'):

                # Exit loop and jump directly to Process 2
                process_2 = AssemblyProcess.objects.filter(name='PROCESS 2 OF 6').first()
                if process_2:
                    next_process = process_2
                    exit_loop_mode = True
                    print(f"CLICKER: Exiting loop mode via manual forward → Jumping to {next_process.name}")

            elif (station.loop_mode and 
                  current_proc and 
                  current_proc.loop_group == 'final_assembly_1abc' and
                  next_process.loop_group != 'final_assembly_1abc'):

                # Exit loop due to process transition
                exit_loop_mode = True
                print(f"MANUAL FORWARD: Exiting loop - from {current_proc.name} to {next_process.name}")

            # ✅ Auto-enable loop when arriving at 1A from non-loop process
            elif (
                not station.loop_mode and 
                next_process and 
                next_process.name == 'PROCESS 1A OF 6'
            ):
                enter_loop_mode = True
                print(f"AUTO-ENTER LOOP MODE: Entering loop at {next_process.name}")

            if next_process:
                all_stations = Station.objects.all()
                updated_stations = []

                for st in all_stations:
                    if st.current_product == station.current_product:
                        old_process_name = st.current_process.name if st.current_process else None
                        st.current_process = next_process

                        if next_process.stage != st.current_stage:
                            st.current_stage = next_process.stage

                        if exit_loop_mode:
                            st.loop_mode = False
                        if enter_loop_mode:
                            st.loop_mode = True

                        st.save()
                        updated_stations.append({
                            'id': st.id,
                            'name': st.name,
                            'display_number': st.display_number,
                            'old_process': old_process_name,
                            'new_process': next_process.name,
                            'loop_mode_changed': exit_loop_mode or enter_loop_mode
                        })

                # NEW: Reset BOM pagination when process changes
                if station.current_product:
                    print(f"PAGINATION RESET: Process changed from {current_proc.name if current_proc else 'None'} to {next_process.name}")
                    BOMPaginationManager.clear_product_pagination(station.current_product.id)

                response_data = {
                    'success': True,
                    'message': f'All stations moved to {next_process.display_name}',
                    'current_process': {
                        'id': next_process.id,
                        'name': next_process.name,
                        'display_name': next_process.display_name,
                        'stage': next_process.stage.display_name,
                        'is_looped': next_process.is_looped,
                        'loop_group': next_process.loop_group
                    },
                    'loop_mode': True if enter_loop_mode else (station.loop_mode if not exit_loop_mode else False),
                    'updated_stations': updated_stations,
                    'exit_loop': exit_loop_mode,
                    'manual_action': True,
                    'pagination_reset': True,  # NEW: Flag indicating pagination was reset
                    'next_process': {
                        'id': station.get_next_process().id,
                        'name': station.get_next_process().name,
                        'display_name': station.get_next_process().display_name
                    } if station.get_next_process() else None
                }

                if exit_loop_mode:
                    response_data['exit_message'] = f'Exited loop mode - moved from {current_proc.name} to {next_process.name}'

                return JsonResponse(response_data)
            else:
                return JsonResponse({'error': 'No next process available'}, status=400)


        elif action == 'backward':
            exit_loop_mode = False
            enter_loop_mode = False

            current_proc = station.current_process
            previous_process = station.get_previous_process()

            # 🔧 Manual override: If current is 1A, jump back to Sub Assembly 2
            if (
                station.loop_mode and 
                current_proc and 
                current_proc.name == 'PROCESS 1A OF 6'
            ):
                previous_process = AssemblyProcess.objects.filter(
                    name='PROCESS 1 OF 1'
                ).first()

                if previous_process:
                    exit_loop_mode = True
                    print(f"OVERRIDE BACKWARD: From 1A → {previous_process.name}")
                else:
                    return JsonResponse({'error': 'Could not locate Sub Assembly 2 process'}, status=404)

            elif (
                station.loop_mode and 
                current_proc and 
                current_proc.loop_group == 'final_assembly_1abc' and 
                (previous_process.loop_group != 'final_assembly_1abc' or previous_process.loop_group is None)
            ):
                exit_loop_mode = True
                print(f"BACKWARD: Exiting loop mode – {current_proc.name} → {previous_process.name}")

            elif (
                not station.loop_mode and 
                previous_process and 
                previous_process.name == 'PROCESS 1A OF 6'
            ):
                enter_loop_mode = True
                print(f"BACKWARD: Auto-entering loop mode at {previous_process.name}")

            if previous_process is None:
                return JsonResponse({'error': 'No previous process available'}, status=400)

            # 🔁 Update all stations with same product
            all_stations = Station.objects.all()
            updated_stations = []

            for st in all_stations:
                if st.current_product == station.current_product:
                    old_process_name = st.current_process.name if st.current_process else None
                    st.current_process = previous_process

                    if previous_process.stage != st.current_stage:
                        st.current_stage = previous_process.stage

                    if exit_loop_mode:
                        st.loop_mode = False
                    if enter_loop_mode:
                        st.loop_mode = True

                    st.save()
                    updated_stations.append({
                        'id': st.id,
                        'name': st.name,
                        'display_number': st.display_number,
                        'old_process': old_process_name,
                        'new_process': previous_process.name,
                        'loop_mode_changed': exit_loop_mode or enter_loop_mode
                    })

            # NEW: Reset BOM pagination when process changes
            if station.current_product:
                print(f"PAGINATION RESET: Process changed from {current_proc.name if current_proc else 'None'} to {previous_process.name}")
                BOMPaginationManager.clear_product_pagination(station.current_product.id)

            return JsonResponse({
                'success': True,
                'message': f'All stations moved to {previous_process.display_name}',
                'current_process': {
                    'id': previous_process.id,
                    'name': previous_process.name,
                    'display_name': previous_process.display_name,
                    'stage': previous_process.stage.display_name,
                    'is_looped': previous_process.is_looped,
                    'loop_group': previous_process.loop_group
                },
                'loop_mode': True if enter_loop_mode else (station.loop_mode if not exit_loop_mode else False),
                'updated_stations': updated_stations,
                'manual_action': True,
                'pagination_reset': True,  # NEW: Flag indicating pagination was reset
                'previous_process': {
                    'id': previous_process.id,
                    'name': previous_process.name,
                    'display_name': previous_process.display_name
                }
            })

        elif action == 'toggle_loop':
            # No pagination reset needed for loop toggle
            if (station.current_process and 
                station.current_process.loop_group == 'final_assembly_1abc'):

                new_loop_mode = not station.loop_mode
                all_stations = Station.objects.all()
                updated_stations = []

                for st in all_stations:
                    if (st.current_product == station.current_product and
                        st.current_process and 
                        st.current_process.loop_group == 'final_assembly_1abc'):
                        st.loop_mode = new_loop_mode
                        st.save()
                        updated_stations.append({
                            'id': st.id,
                            'name': st.name,
                            'display_number': st.display_number,
                            'loop_mode': new_loop_mode
                        })

                return JsonResponse({
                    'success': True,
                    'loop_mode': new_loop_mode,
                    'message': f"Loop mode {'enabled' if new_loop_mode else 'disabled'} on all stations",
                    'updated_stations': updated_stations,
                    'manual_action': True
                })
            else:
                return JsonResponse({'error': 'Loop mode only available for processes 1A, 1B, 1C'}, status=400)

        else:
            return JsonResponse({'error': 'Invalid action'}, status=400)

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except ValueError:
        return JsonResponse({'error': 'Invalid quantity value'}, status=400)
    except Exception as e:
        print(f"Clicker action error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)    
      
      
    
def get_assembly_options(request, station_id):
    """Get available options for BRG assembly configuration"""
    station = get_object_or_404(Station, pk=station_id)
    
    # Get available products for this station
    products = [
        {
            'id': p.id,
            'code': p.code,
            'name': p.name
        }
        for p in station.products.all()
    ]
    
    # Get all assembly stages
    stages = [
        {
            'id': s.id,
            'name': s.name,
            'display_name': s.display_name,
            'order': s.order,
            'processes': [
                {
                    'id': p.id,
                    'name': p.name,
                    'display_name': p.display_name,
                    'order': p.order,
                    'location': p.get_location_display(),
                    'is_looped': p.is_looped,
                    'loop_group': p.loop_group
                }
                for p in s.processes.all()
            ]
        }
        for s in AssemblyStage.objects.all().prefetch_related('processes')
    ]
    
    # Get current configuration
    current_config = {
        'product_id': station.current_product.id if station.current_product else None,
        'stage_id': station.current_stage.id if station.current_stage else None,
        'process_id': station.current_process.id if station.current_process else None,
        'quantity': station.product_quantity,
        'clicker_enabled': station.clicker_enabled,
        'loop_mode': station.loop_mode,
        'display_number': station.display_number,
        'bom_settings': {
            'show_single_unit': station.show_single_unit_bom,
            'show_batch': station.show_batch_bom
        }
    }
    
    return JsonResponse({
        'products': products,
        'stages': stages,
        'current_config': current_config,
        'workflow_info': {
            'total_stages': len(stages),
            'loop_processes': ['PROCESS 1A OF 6', 'PROCESS 1B OF 6', 'PROCESS 1C OF 6'],
            'display_assignments': {
                1: 'BOMs and Reference Materials',
                2: 'Process Documents and Instructions', 
                3: 'Instructional Videos'
            }
        }
    })

def get_workflow_status(request):
    """Get overall workflow status across all displays"""
    stations = Station.objects.all().select_related(
        'current_product', 'current_stage', 'current_process'
    )
    
    workflow_status = {
        'stations': [],
        'active_sessions': AssemblySession.objects.filter(completed=False).count(),
        'current_time': time.time()
    }
    
    for station in stations:
        station_info = {
            'id': station.id,
            'name': station.name,
            'display_number': station.display_number,
            'current_product': {
                'code': station.current_product.code,
                'name': station.current_product.name
            } if station.current_product else None,
            'current_stage': station.current_stage.display_name if station.current_stage else None,
            'current_process': {
                'name': station.current_process.name,
                'display_name': station.current_process.display_name,
                'is_looped': station.current_process.is_looped
            } if station.current_process else None,
            'quantity': station.product_quantity,
            'loop_mode': station.loop_mode,
            'clicker_enabled': station.clicker_enabled,
            'media_count': station.get_current_media().count()
        }
        workflow_status['stations'].append(station_info)
    
    return JsonResponse(workflow_status)

@csrf_exempt
@require_http_methods(["POST"])
def sync_all_displays(request):
    """Synchronize all displays to the same process step"""
    try:
        data = json.loads(request.body)
        target_stage_id = data.get('stage_id')
        target_process_id = data.get('process_id')
        
        if not target_stage_id or not target_process_id:
            return JsonResponse({'error': 'Stage ID and Process ID required'}, status=400)
        
        target_stage = get_object_or_404(AssemblyStage, id=target_stage_id)
        target_process = get_object_or_404(AssemblyProcess, id=target_process_id)
        
        # Update all stations
        stations = Station.objects.all()
        updated_stations = []
        
        for station in stations:
            station.current_stage = target_stage
            station.current_process = target_process
            
            # Handle loop mode for processes 1A, 1B, 1C
            if target_process.loop_group == 'final_assembly_1abc':
                station.loop_mode = True
            else:
                station.loop_mode = False
                
            station.save()
            updated_stations.append({
                'id': station.id,
                'name': station.name,
                'display_number': station.display_number
            })
        
        return JsonResponse({
            'success': True,
            'message': f'All displays synchronized to {target_process.display_name}',
            'updated_stations': updated_stations,
            'target_process': {
                'id': target_process.id,
                'name': target_process.name,
                'display_name': target_process.display_name,
                'stage': target_stage.display_name
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_bom_files(request, station_id):
    """Get BOM files for a specific station"""
    station = get_object_or_404(Station, pk=station_id)
    
    if not station.current_product:
        return JsonResponse({'error': 'No product selected'}, status=400)
    
    boms = BillOfMaterial.objects.filter(product=station.current_product)
    
    # Filter BOMs based on station settings
    if station.show_single_unit_bom:
        boms = boms.filter(bom_type='SINGLE_UNIT')
    elif station.show_batch_bom:
        boms = boms.filter(bom_type='BATCH_50')
    
    # Include stage-specific BOMs if in that stage
    if station.current_stage:
        stage_boms = BillOfMaterial.objects.filter(
            product=station.current_product,
            stage=station.current_stage
        )
        boms = boms.union(stage_boms)
    
    bom_data = [
        {
            'id': bom.id,
            'type': bom.get_bom_type_display(),
            'file_url': bom.file.url,
            'stage': bom.stage.display_name if bom.stage else None
        }
        for bom in boms
    ]
    
    return JsonResponse({
        'boms': bom_data,
        'station_info': {
            'display_number': station.display_number,
            'show_single_unit': station.show_single_unit_bom,
            'show_batch': station.show_batch_bom
        }
    })

# File streaming views
def stream_video(request, video_path):
    """Stream video files with range support"""
    clean_path = video_path.lstrip('/').replace('media/', '', 1)
    path = os.path.join(settings.MEDIA_ROOT, clean_path)
    
    if not os.path.exists(path):
        return JsonResponse({'error': 'File not found'}, status=404)
    
    range_header = request.META.get('HTTP_RANGE', '').strip()
    size = os.path.getsize(path)
    
    def file_iterator(path, chunk_size=8192, offset=0, length=None):
        with open(path, 'rb') as f:
            if offset:
                f.seek(offset)
            remaining = length if length is not None else None
            while True:
                if remaining is not None:
                    chunk_size = min(chunk_size, remaining)
                data = f.read(chunk_size)
                if not data:
                    break
                if remaining is not None:
                    remaining -= len(data)
                yield data
                if remaining == 0:
                    break
    
    if range_header:
        range_s, range_e = range_header.split('=')[-1].split('-')
        range_start = int(range_s)
        range_end = int(range_e) if range_e else size - 1
        length = range_end - range_start + 1

        response = StreamingHttpResponse(
            file_iterator(path, offset=range_start, length=length),
            status=206,
            content_type='video/mp4'
        )
        response['Content-Range'] = f'bytes {range_start}-{range_end}/{size}'
    else:
        response = StreamingHttpResponse(
            file_iterator(path),
            content_type='video/mp4'
        )
    
    response['Accept-Ranges'] = 'bytes'
    response['Content-Length'] = str(size)
    return response

def stream_pdf(request, pdf_path):
    """Stream PDF files"""
    clean_path = pdf_path.lstrip('/').replace('media/', '', 1)
    path = os.path.join(settings.MEDIA_ROOT, clean_path)
    
    if not os.path.exists(path):
        return JsonResponse({'error': 'File not found'}, status=404)
    
    def file_iterator(path, chunk_size=8192):
        with open(path, 'rb') as f:
            while True:
                data = f.read(chunk_size)
                if not data:
                    break
                yield data
    
    response = StreamingHttpResponse(
        file_iterator(path),
        content_type='application/pdf'
    )
    response['Content-Length'] = os.path.getsize(path)
    return response
        
@csrf_exempt
@require_http_methods(["POST"])
def update_assembly_config(request, station_id):
    """Enhanced assembly configuration with AUTO LOOP MODE DETECTION"""
    station = get_object_or_404(Station, pk=station_id)
    
    try:
        data = json.loads(request.body)
        
        # Update product
        if 'product_id' in data:
            if data['product_id']:
                product = get_object_or_404(Product, id=data['product_id'])
                if product in station.products.all():
                    station.current_product = product
                    # Reset to first stage and process when product changes
                    first_stage = AssemblyStage.objects.first()
                    station.current_stage = first_stage
                    station.current_process = first_stage.processes.first() if first_stage else None
                else:
                    return JsonResponse({'error': 'Product not available for this station'}, status=400)
        
        # Update stage
        if 'stage_id' in data:
            if data['stage_id']:
                stage = get_object_or_404(AssemblyStage, id=data['stage_id'])
                station.current_stage = stage
                # Reset to first process of new stage
                station.current_process = stage.processes.first()
                
                # Auto-configure BOM settings based on stage
                if stage.name == 'BOM_DISPLAY':
                    # Set default BOM display based on quantity
                    if station.product_quantity == 1:
                        station.show_single_unit_bom = True
                        station.show_batch_bom = False
                    else:
                        station.show_single_unit_bom = False
                        station.show_batch_bom = True
        
        # Update process
        if 'process_id' in data:
            if data['process_id']:
                process = get_object_or_404(AssemblyProcess, id=data['process_id'])
                station.current_process = process
                station.current_stage = process.stage
                
                # Handle BOM-specific processes
                if process.stage.name == 'BOM_DISPLAY':
                    if process.name == 'SINGLE_UNIT_BOM_DISPLAY':
                        station.show_single_unit_bom = True
                        station.show_batch_bom = False
                        station.product_quantity = 1
                    elif process.name == 'BATCH_50_BOM_DISPLAY':
                        station.show_single_unit_bom = False
                        station.show_batch_bom = True
                        if station.product_quantity < 2:
                            station.product_quantity = 50
                    elif process.name == 'STAGE_BOM_DISPLAY':
                        station.show_single_unit_bom = False
                        station.show_batch_bom = False
                
                # ENHANCED: Auto-enable loop mode when entering Process 1A
                if (process.loop_group == 'final_assembly_1abc' and 
                    process.name == 'PROCESS 1A OF 6'):
                    station.loop_mode = True
                    print(f"Auto-enabling loop mode for Process 1A on station {station.id}")
                elif process.loop_group != 'final_assembly_1abc':
                    # Disable loop mode when leaving the loop group
                    station.loop_mode = False
        
        # Update quantity
        if 'quantity' in data:
            quantity = int(data['quantity'])
            if quantity > 0:
                station.product_quantity = quantity
                # Auto-adjust BOM display based on quantity
                if station.current_stage and station.current_stage.name == 'BOM_DISPLAY':
                    if quantity == 1:
                        station.show_single_unit_bom = True
                        station.show_batch_bom = False
                    else:
                        station.show_single_unit_bom = False
                        station.show_batch_bom = True
        
        # Update BOM display settings
        if 'show_single_unit_bom' in data:
            station.show_single_unit_bom = data['show_single_unit_bom']
            if data['show_single_unit_bom']:
                station.show_batch_bom = False
                station.product_quantity = 1
                
        if 'show_batch_bom' in data:
            station.show_batch_bom = data['show_batch_bom']
            if data['show_batch_bom']:
                station.show_single_unit_bom = False
                if station.product_quantity < 2:
                    station.product_quantity = 50
            
        station.save()
        
        # Get updated BOM data for response
        bom_data = station.get_current_bom_data()
        bom_info = station.get_current_bom_info()
        
        return JsonResponse({
            'success': True,
            'message': 'Configuration updated successfully',
            'auto_loop_enabled': (station.current_process and 
                                 station.current_process.loop_group == 'final_assembly_1abc' and
                                 station.current_process.name == 'PROCESS 1A OF 6'),  # NEW: Auto loop flag
            'current_state': {
                'product': {
                    'id': station.current_product.id,
                    'code': station.current_product.code,
                    'name': station.current_product.name
                } if station.current_product else None,
                'stage': {
                    'id': station.current_stage.id,
                    'name': station.current_stage.display_name,
                    'order': station.current_stage.order,
                    'is_bom_stage': station.current_stage.name == 'BOM_DISPLAY'
                } if station.current_stage else None,
                'process': {
                    'id': station.current_process.id,
                    'name': station.current_process.name,
                    'display_name': station.current_process.display_name,
                    'order': station.current_process.order,
                    'loop_group': station.current_process.loop_group,
                    'is_auto_loop': (station.current_process.loop_group == 'final_assembly_1abc' and
                                   station.current_process.name == 'PROCESS 1A OF 6')  # NEW: Auto loop indicator
                } if station.current_process else None,
                'quantity': station.product_quantity,
                'loop_mode': station.loop_mode,
                'bom_settings': {
                    'show_single_unit': station.show_single_unit_bom,
                    'show_batch': station.show_batch_bom
                },
                'bom_data_available': bool(bom_data),
                'bom_item_count': len(bom_data) if bom_data else 0,
                'bom_is_split': bom_info.get('is_split', False) if bom_info else False,
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except ValueError as e:
        return JsonResponse({'error': f'Invalid value: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# NEW: Auto loop 
@csrf_exempt
@require_http_methods(["POST"])
def auto_loop_progress(request, station_id):
    """Handle automatic progression in loop mode - ENHANCED VERSION"""
    station = get_object_or_404(Station, pk=station_id)
    
    try:
        data = json.loads(request.body)
        client_timestamp = data.get('timestamp')
        expected_process = data.get('expectedProcess')
        
        # FIXED: Validate that enough time has passed since last auto progression
        current_time = time.time() * 1000  # Convert to milliseconds
        if hasattr(station, 'last_auto_progress_time'):
            time_since_last = current_time - station.last_auto_progress_time
            if time_since_last < 5000:  # Prevent progressions within 5 seconds
                return JsonResponse({
                    'error': f'Auto progression too frequent. Last: {time_since_last}ms ago',
                    'time_since_last': time_since_last,
                    'minimum_interval': 5000
                }, status=429)
        
        # FIXED: Validate that we're still in the expected process
        if expected_process and station.current_process:
            if station.current_process.name != expected_process:
                return JsonResponse({
                    'error': f'Process changed. Expected: {expected_process}, Current: {station.current_process.name}',
                    'expected': expected_process,
                    'current': station.current_process.name if station.current_process else None
                }, status=409)
        
        # Existing validation logic
        if not (station.loop_mode and 
                station.current_process and 
                station.current_process.loop_group == 'final_assembly_1abc'):
            return JsonResponse({
                'error': 'Auto progression not allowed',
                'details': {
                    'loop_mode': station.loop_mode,
                    'current_process': station.current_process.name if station.current_process else None,
                    'loop_group': station.current_process.loop_group if station.current_process else None
                }
            }, status=400)
        
        # Determine next process in loop
        current_name = station.current_process.name
        next_process = None
        loop_sequence = ['PROCESS 1A OF 6', 'PROCESS 1B OF 6', 'PROCESS 1C OF 6']
        
        try:
            current_index = loop_sequence.index(current_name)
            next_index = (current_index + 1) % len(loop_sequence)
            next_process_name = loop_sequence[next_index]
            
            next_process = AssemblyProcess.objects.filter(
                name=next_process_name,
                loop_group='final_assembly_1abc'
            ).first()
            
        except ValueError:
            return JsonResponse({
                'error': f'Current process {current_name} not in loop sequence',
                'loop_sequence': loop_sequence
            }, status=400)
        
        if not next_process:
            return JsonResponse({
                'error': f'Next process {next_process_name} not found in database',
                'searched_for': next_process_name
            }, status=404)
        
        # Log the auto progression
        print(f"AUTO LOOP PROGRESSION: {current_name} → {next_process.name} (Client: {client_timestamp})")
        
        # FIXED: Update timestamp tracking
        station.last_auto_progress_time = current_time
        
        # Update ALL stations with the same product
        all_stations = Station.objects.all()
        updated_stations = []
        skipped_stations = []
        
        for st in all_stations:
            if (st.current_product == station.current_product and
                st.loop_mode and
                st.current_process and 
                st.current_process.loop_group == 'final_assembly_1abc'):
                
                old_process = st.current_process.name
                st.current_process = next_process
                
                if next_process.stage != st.current_stage:
                    st.current_stage = next_process.stage
                
                # FIXED: Update timestamp on all affected stations
                st.last_auto_progress_time = current_time
                st.save()
                
                updated_stations.append({
                    'id': st.id,
                    'name': st.name,
                    'display_number': st.display_number,
                    'old_process': old_process,
                    'new_process': next_process.name
                })
                
            else:
                skip_reason = []
                if st.current_product != station.current_product:
                    skip_reason.append('different_product')
                if not st.loop_mode:
                    skip_reason.append('not_in_loop_mode')
                if not st.current_process:
                    skip_reason.append('no_current_process')
                elif st.current_process.loop_group != 'final_assembly_1abc':
                    skip_reason.append('not_in_loop_group')
                
                skipped_stations.append({
                    'id': st.id,
                    'name': st.name,
                    'display_number': st.display_number,
                    'skip_reasons': skip_reason
                })
        
        # Enhanced response data
        response_data = {
            'success': True,
            'message': f'Auto-progressed from {current_name} to {next_process.display_name}',
            'progression': {
                'from': {
                    'name': current_name,
                    'display_name': station.current_process.display_name if station.current_process else current_name
                },
                'to': {
                    'name': next_process.name,
                    'display_name': next_process.display_name
                }
            },
            'current_process': {
                'id': next_process.id,
                'name': next_process.name,
                'display_name': next_process.display_name,
                'loop_group': next_process.loop_group,
                'stage': {
                    'id': next_process.stage.id,
                    'name': next_process.stage.name,
                    'display_name': next_process.stage.display_name
                }
            },
            'updated_stations': updated_stations,
            'skipped_stations': skipped_stations,
            'loop_continues': True,
            'loop_sequence': loop_sequence,
            'current_position': loop_sequence.index(next_process.name),
            'timestamp': current_time,
            'server_time': time.time() * 1000,
            'client_timestamp': client_timestamp,
            'progression_id': f"{station_id}_{current_time}"  # Unique ID for this progression
        }
        
        # Add next progression info
        try:
            current_index = loop_sequence.index(next_process.name)
            next_index = (current_index + 1) % len(loop_sequence)
            response_data['next_progression'] = {
                'will_go_to': loop_sequence[next_index],
                'is_cycle_complete': next_index == 0
            }
        except ValueError:
            pass
        
        print(f"AUTO LOOP SUCCESS: Updated {len(updated_stations)} stations, skipped {len(skipped_stations)}")
        
        return JsonResponse(response_data)
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        error_message = f'Auto loop progression error: {str(e)}'
        print(f"AUTO LOOP ERROR: {error_message}")
        
        return JsonResponse({
            'error': error_message,
            'success': False,
            'current_process': {
                'name': station.current_process.name if station.current_process else None,
                'loop_group': station.current_process.loop_group if station.current_process else None
            },
            'station_info': {
                'id': station.id,
                'name': station.name,
                'loop_mode': station.loop_mode
            },
            'timestamp': time.time() * 1000
        }, status=500)    
      
        
@csrf_exempt
@require_http_methods(["GET", "POST"])
def auto_loop_config(request, station_id):
    """Configure auto loop settings for a station"""
    station = get_object_or_404(Station, pk=station_id)
    
    if request.method == 'GET':
        # Return current auto loop configuration
        return JsonResponse({
            'station': {
                'id': station.id,
                'name': station.name,
                'display_number': station.display_number
            },
            'current_state': {
                'loop_mode': station.loop_mode,
                'current_process': {
                    'name': station.current_process.name if station.current_process else None,
                    'display_name': station.current_process.display_name if station.current_process else None,
                    'loop_group': station.current_process.loop_group if station.current_process else None,
                } if station.current_process else None,
                'is_in_loop_group': (station.current_process and 
                                   station.current_process.loop_group == 'final_assembly_1abc'),
            },
            'loop_configuration': {
                'loop_processes': ['PROCESS 1A OF 6', 'PROCESS 1B OF 6', 'PROCESS 1C OF 6'],
                'default_durations': {
                    'PROCESS 1A OF 6': 30000,  # 30 seconds
                    'PROCESS 1B OF 6': 45000,  # 45 seconds
                    'PROCESS 1C OF 6': 60000   # 60 seconds
                },
                'loop_group': 'final_assembly_1abc'
            }
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')
            
            if action == 'enable_auto_loop':
                # Enable auto loop mode for this station
                if (station.current_process and 
                    station.current_process.loop_group == 'final_assembly_1abc'):
                    
                    station.loop_mode = True
                    station.save()
                    
                    return JsonResponse({
                        'success': True,
                        'message': 'Auto loop mode enabled',
                        'current_process': station.current_process.name,
                        'loop_mode': station.loop_mode
                    })
                else:
                    return JsonResponse({
                        'error': 'Can only enable auto loop in processes 1A, 1B, or 1C',
                        'current_process': station.current_process.name if station.current_process else None
                    }, status=400)
            
            elif action == 'disable_auto_loop':
                # Disable auto loop mode
                station.loop_mode = False
                station.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Auto loop mode disabled',
                    'loop_mode': station.loop_mode
                })
            
            elif action == 'force_progress':
                # Force immediate progression (for testing)
                if (station.loop_mode and 
                    station.current_process and 
                    station.current_process.loop_group == 'final_assembly_1abc'):
                    
                    # Call the auto progression logic
                    progress_request = type('MockRequest', (), {
                        'method': 'POST',
                        'body': json.dumps({}).encode()
                    })()
                    
                    # This would call the auto_loop_progress view
                    return auto_loop_progress(progress_request, station_id)
                else:
                    return JsonResponse({
                        'error': 'Station not in auto loop mode or not in loop group'
                    }, status=400)
            
            else:
                return JsonResponse({'error': 'Invalid action'}, status=400)
        
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


def auto_loop_status_all(request):
    """Get auto loop status for all stations"""
    stations = Station.objects.all().select_related('current_process', 'current_stage', 'current_product')
    
    stations_status = []
    for station in stations:
        is_in_loop_group = (station.current_process and 
                           station.current_process.loop_group == 'final_assembly_1abc')
        
        station_info = {
            'id': station.id,
            'name': station.name,
            'display_number': station.display_number,
            'loop_mode': station.loop_mode,
            'current_process': {
                'name': station.current_process.name if station.current_process else None,
                'display_name': station.current_process.display_name if station.current_process else None,
                'loop_group': station.current_process.loop_group if station.current_process else None,
            } if station.current_process else None,
            'is_in_loop_group': is_in_loop_group,
            'auto_loop_eligible': is_in_loop_group and station.loop_mode,
            'current_product': {
                'code': station.current_product.code,
                'name': station.current_product.name
            } if station.current_product else None
        }
        stations_status.append(station_info)
    
    # Summary statistics
    total_stations = len(stations_status)
    auto_loop_active = len([s for s in stations_status if s['auto_loop_eligible']])
    in_loop_group = len([s for s in stations_status if s['is_in_loop_group']])
    
    return JsonResponse({
        'stations': stations_status,
        'summary': {
            'total_stations': total_stations,
            'auto_loop_active': auto_loop_active,
            'in_loop_group': in_loop_group,
            'loop_processes': ['PROCESS 1A OF 6', 'PROCESS 1B OF 6', 'PROCESS 1C OF 6']
        },
        'timestamp': time.time()
    })


def get_next_process(self):
    """Get the next process in sequence - ENHANCED for loop exit logic"""
    if not self.current_process:
        return None
    
    # Get all processes for the current product, ordered by stage and process order
    all_processes = AssemblyProcess.objects.filter(
        stage__in=self.current_product.stages.all()
    ).select_related('stage').order_by('stage__order', 'order')
    
    current_found = False
    for process in all_processes:
        if current_found:
            return process
        if process.id == self.current_process.id:
            current_found = True
    
    # If we're at the last process, return None (no next process)
    return None


#  debug the issue
def debug_process_sequence(request, station_id):
    """Debug view to check process sequence"""
    try:
        station = Station.objects.get(pk=station_id)
        
        if not station.current_product:
            return JsonResponse({'error': 'No product selected'})
        
        # Get all processes for current product
        all_processes = AssemblyProcess.objects.filter(
            stage__in=station.current_product.stages.all()
        ).select_related('stage').order_by('stage__order', 'order')
        
        process_sequence = []
        for i, process in enumerate(all_processes):
            process_info = {
                'index': i,
                'id': process.id,
                'name': process.name,
                'display_name': process.display_name,
                'stage': process.stage.display_name,
                'loop_group': process.loop_group,
                'is_current': process.id == station.current_process.id if station.current_process else False
            }
            process_sequence.append(process_info)
        
        # Find current process index
        current_index = -1
        if station.current_process:
            for i, process in enumerate(all_processes):
                if process.id == station.current_process.id:
                    current_index = i
                    break
        
        # Determine next process
        next_process_info = None
        if current_index >= 0 and current_index < len(all_processes) - 1:
            next_process = all_processes[current_index + 1]
            next_process_info = {
                'id': next_process.id,
                'name': next_process.name,
                'display_name': next_process.display_name,
                'stage': next_process.stage.display_name,
                'loop_group': next_process.loop_group
            }
        
        return JsonResponse({
            'station': {
                'id': station.id,
                'name': station.name,
                'loop_mode': station.loop_mode
            },
            'current_process': {
                'id': station.current_process.id if station.current_process else None,
                'name': station.current_process.name if station.current_process else None,
                'loop_group': station.current_process.loop_group if station.current_process else None,
                'index': current_index
            },
            'next_process': next_process_info,
            'all_processes': process_sequence,
            'total_processes': len(process_sequence)
        })
        
    except Station.DoesNotExist:
        return JsonResponse({'error': 'Station not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

    
def station_media_stream_enhanced(request, station_id):
    """Enhanced streaming with BOM data support"""
    def event_stream():
        last_update = None
        while True:
            try:
                station = get_object_or_404(Station, pk=station_id)
                current_media = station.get_current_media()
                bom_data = station.get_current_bom_data()
                
                # Create comprehensive comparison data including BOM
                media_comparison = []
                for media in current_media:
                    comparison_item = {
                        'id': media.id,
                        'file_url': media.file.url if media.file else None,
                        'media_type': media.media_type,
                        'process_id': media.process.id if media.process else None,
                        'bom_id': media.bom.id if media.bom else None
                    }
                    media_comparison.append(comparison_item)
                
                # Include BOM data in comparison
                bom_comparison = []
                if bom_data:
                    for item_data in bom_data:
                        bom_comparison.append({
                            'item_code': item_data['item'].item_code,
                            'calculated_quantity': item_data['calculated_quantity'],
                            'serial_number': item_data['serial_number']
                        })
                
                # Include assembly state in comparison
                assembly_state = {
                    'current_product_id': station.current_product.id if station.current_product else None,
                    'current_stage_id': station.current_stage.id if station.current_stage else None,
                    'current_process_id': station.current_process.id if station.current_process else None,
                    'product_quantity': station.product_quantity,
                    'loop_mode': station.loop_mode,
                    'display_number': station.display_number,
                    'show_single_unit_bom': station.show_single_unit_bom,
                    'show_batch_bom': station.show_batch_bom,
                }
                
                current_state = {
                    'media': media_comparison,
                    'bom': bom_comparison,
                    'assembly': assembly_state
                }
                
                # Only send update if something changed
                if current_state != last_update:
                    # Prepare detailed media data for frontend
                    media_data = []
                    for media in current_media:
                        media_info = {
                            'id': media.id,
                            'url': media.file.url if media.file else None,
                            'type': media.file.name.split('.')[-1].lower() if media.file else 'bom',
                            'duration': media.duration,
                            'media_type': media.get_media_type_display(),
                            'product_name': media.product.name,
                            'product_code': media.product.code,
                        }
                        
                        if media.process:
                            media_info['process'] = {
                                'id': media.process.id,
                                'name': media.process.name,
                                'display_name': media.process.display_name,
                                'stage': media.process.stage.display_name,
                                'order': media.process.order,
                                'is_looped': media.process.is_looped,
                                'loop_group': media.process.loop_group
                            }
                        
                        if media.bom:
                            media_info['bom'] = {
                                'id': media.bom.id,
                                'type': media.bom.get_bom_type_display(),
                                'stage': media.bom.stage.display_name if media.bom.stage else None
                            }
                        
                        media_data.append(media_info)
                    
                    # Prepare BOM data for frontend
                    formatted_bom = []
                    if bom_data:
                        for item_data in bom_data:
                            formatted_bom.append({
                                'serial_number': item_data['serial_number'],
                                'item_code': item_data['item'].item_code,
                                'item_description': item_data['item'].item_description,
                                'part_number': item_data['item'].part_number,
                                'formatted_quantity': item_data['formatted_quantity'],
                                'notes': item_data['notes'],
                                'item_photo_url': item_data['item'].item_photo.url if item_data['item'].item_photo else None,
                                'supplier': item_data['item'].supplier,
                            })
                    
                    # Prepare assembly info
                    assembly_info = {
                        'current_product': {
                            'id': station.current_product.id,
                            'code': station.current_product.code,
                            'name': station.current_product.name
                        } if station.current_product else None,
                        'current_stage': {
                            'id': station.current_stage.id,
                            'name': station.current_stage.display_name,
                            'order': station.current_stage.order
                        } if station.current_stage else None,
                        'current_process': {
                            'id': station.current_process.id,
                            'name': station.current_process.name,
                            'display_name': station.current_process.display_name,
                            'order': station.current_process.order,
                            'is_looped': station.current_process.is_looped,
                            'loop_group': station.current_process.loop_group
                        } if station.current_process else None,
                        'quantity': station.product_quantity,
                        'loop_mode': station.loop_mode,
                        'clicker_enabled': station.clicker_enabled,
                        'display_number': station.display_number,
                        'bom_settings': {
                            'show_single_unit': station.show_single_unit_bom,
                            'show_batch': station.show_batch_bom,
                        },
                        'next_process': {
                            'id': station.get_next_process().id,
                            'name': station.get_next_process().name,
                            'display_name': station.get_next_process().display_name
                        } if station.get_next_process() else None,
                        'previous_process': {
                            'id': station.get_previous_process().id,  
                            'name': station.get_previous_process().name,
                            'display_name': station.get_previous_process().display_name
                        } if station.get_previous_process() else None,
                    }
                    
                    response_data = {
                        'media': media_data,
                        'bom_data': formatted_bom,
                        'station_name': station.name,
                        'assembly': assembly_info,
                        'timestamp': time.time()
                    }
                    
                    last_update = current_state
                    yield f"data: {json.dumps(response_data)}\n\n"
                
                time.sleep(1)  # Check every 3 seconds
                
            except Station.DoesNotExist:
                yield f"data: {json.dumps({'error': 'Station not found'})}\n\n"
                break
            except Exception as e:
                yield f"data: {json.dumps({'error': str(e)})}\n\n"
                time.sleep(10)

    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['Connection'] = 'keep-alive'
    response['X-Accel-Buffering'] = 'no'
    return response



def get_station_media_with_bom(request, station_id):
    """Enhanced station media API that handles split BOM properly for all displays - FIXED"""
    station = get_object_or_404(Station, pk=station_id)
    
    # Get regular media (excluding PDF BOMs when we have database BOM)
    current_media = station.get_current_media()
    
    media_data = []
    
    # Check if we have database BOM data for this specific display
    bom_data = station.get_current_bom_data() if station.current_product else []
    bom_info = station.get_current_bom_info()
    has_database_bom = bool(bom_data)
    
    print(f"Station {station_id} Display {station.display_number}:")
    print(f"  - Has BOM data: {has_database_bom}")
    print(f"  - BOM items count: {len(bom_data) if bom_data else 0}")
    print(f"  - Current stage: {station.current_stage.name if station.current_stage else None}")
    
    # ALWAYS add database BOM as first item if we have data and are in BOM stage
    if (has_database_bom and station.current_stage and 
        station.current_stage.name == 'BOM_DISPLAY'):
        
        # Determine BOM display type
        bom_display_type = "Database BOM"
        if station.show_single_unit_bom:
            bom_display_type = "Single Unit BOM"
        elif station.show_batch_bom:
            bom_display_type = f"{station.product_quantity} Units BOM"
        elif station.current_stage:
            bom_display_type = f"{station.current_stage.display_name} BOM"
        
        # Add split information if applicable
        display_suffix = ""
        items_for_display = len(bom_data)
        
        if bom_info and bom_info.get('is_split'):
            split_info = bom_info.get('split_info', {})
            current_display_info = split_info.get(f'display_{station.display_number}', {})
            items_for_display = current_display_info.get('item_count', 0)
            
            if items_for_display > 0:
                display_suffix = f" (Part {station.display_number}/3 - {items_for_display} items)"
            else:
                display_suffix = f" (Display {station.display_number}/3 - No items)"
        
        bom_media_info = {
            'id': f'database_bom_{station.id}',
            'url': f'/station/{station.id}/bom-render/',
            'type': 'bom',
            'duration': 40,
            'media_type': f'Database BOM',
            'product_name': station.current_product.name,
            'product_code': station.current_product.code,
            'is_bom_data': True,
            'bom_items': items_for_display,  # Items for THIS display
            'bom_type': bom_display_type + display_suffix,
            'is_split': bom_info.get('is_split', False) if bom_info else False,
            'display_info': bom_info.get('display_info', '') if bom_info else '',
        }
        
        media_data.append(bom_media_info)
        print(f"  - Added BOM media: {bom_media_info['bom_type']}")
        
        # Filter out PDF BOMs from regular media when we have database BOM
        filtered_media = current_media.exclude(media_type='Bill of Material')
        
        # Add remaining non-BOM media
        for media in filtered_media:
            media_info = {
                'id': media.id,
                'url': media.file.url if media.file else None,
                'type': media.file.name.split('.')[-1].lower() if media.file else 'unknown',
                'duration': media.duration,
                'media_type': media.get_media_type_display(),
                'product_name': media.product.name,
                'product_code': media.product.code,
                'is_bom_data': False,
            }
            
            if media.process:
                media_info['process'] = {
                    'id': media.process.id,
                    'name': media.process.name,
                    'display_name': media.process.display_name,
                    'stage': media.process.stage.display_name,
                    'order': media.process.order,
                    'is_looped': media.process.is_looped,
                    'loop_group': media.process.loop_group
                }
            
            media_data.append(media_info)
    
    else:
        print(f"  - Not adding BOM: stage={station.current_stage.name if station.current_stage else None}, has_bom={has_database_bom}")
        
        # Regular media handling for non-BOM stages
        for media in current_media:
            media_info = {
                'id': media.id,
                'url': media.file.url if media.file else None,
                'type': media.file.name.split('.')[-1].lower() if media.file else 'unknown',
                'duration': media.duration,
                'media_type': media.get_media_type_display(),
                'product_name': media.product.name,
                'product_code': media.product.code,
                'is_bom_data': False,
            }
            
            if media.process:
                media_info['process'] = {
                    'id': media.process.id,
                    'name': media.process.name,
                    'display_name': media.process.display_name,
                    'stage': media.process.stage.display_name,
                    'order': media.process.order,
                    'is_looped': media.process.is_looped,
                    'loop_group': media.process.loop_group
                }
            
            media_data.append(media_info)
    
    # FIXED: Properly serialize BOM info without model objects
    serialized_bom_info = None
    if bom_info:
        # Convert any model objects to serializable data
        serialized_bom_info = {}
        for key, value in bom_info.items():
            if key == 'template' and value:
                # Convert BOMTemplate to serializable dict
                serialized_bom_info['template'] = {
                    'id': value.id,
                    'name': value.template_name,
                    'bom_type': value.bom_type,
                    'description': value.description,
                    'product_id': value.product.id,
                    'product_code': value.product.code,
                    'stage_id': value.stage.id if value.stage else None,
                    'stage_name': value.stage.name if value.stage else None,
                }
            elif key == 'stage' and value:
                # Convert AssemblyStage to serializable dict
                serialized_bom_info['stage'] = {
                    'id': value.id,
                    'name': value.name,
                    'display_name': value.display_name,
                    'order': value.order,
                }
            else:
                # Copy primitive values as-is
                serialized_bom_info[key] = value
    
    response_data = {
        'media': media_data,
        'station_info': {
            'name': station.name,
            'display_number': station.display_number,
            'current_product': {
                'id': station.current_product.id,
                'code': station.current_product.code,
                'name': station.current_product.name
            } if station.current_product else None,
            'current_stage': {
                'id': station.current_stage.id,
                'name': station.current_stage.display_name,
                'order': station.current_stage.order
            } if station.current_stage else None,
            'current_process': {
                'id': station.current_process.id,
                'name': station.current_process.name,
                'display_name': station.current_process.display_name,
                'order': station.current_process.order,
                'is_looped': station.current_process.is_looped,
                'loop_group': station.current_process.loop_group
            } if station.current_process else None,
            'quantity': station.product_quantity,
            'loop_mode': station.loop_mode,
            'clicker_enabled': station.clicker_enabled,
            'bom_settings': {
                'show_single_unit': station.show_single_unit_bom,
                'show_batch': station.show_batch_bom,
            },
            'bom_split_info': serialized_bom_info,  # FIXED: Use serialized version
            'has_database_bom': has_database_bom,
            'is_bom_stage': (station.current_stage and 
                           station.current_stage.name == 'BOM_DISPLAY')
        }
    }
    
    print(f"  - Returning {len(media_data)} media items")
    return JsonResponse(response_data)
  
def render_bom_for_slider(request, station_id):
    """Render BOM as HTML fragment for slider integration with enhanced split logic"""
    station = get_object_or_404(Station, pk=station_id)
    
    if not station.current_product:
        return HttpResponse('''
            <div class="bom-container-slider">
                <div class="no-bom-data-slider">
                    <div class="no-bom-icon-slider">📋</div>
                    <h3>No Product Selected</h3>
                    <p>Please select a product to view BOM data</p>
                </div>
            </div>
        ''')
    
    bom_data = station.get_current_bom_data()
    
    if not bom_data:
        return HttpResponse('''
            <div class="bom-container-slider">
                <div class="no-bom-data-slider">
                    <div class="no-bom-icon-slider">📋</div>
                    <h3>No BOM Data Available</h3>
                    <p>No bill of materials configured for current settings</p>
                    <p>Check BOM templates in admin panel</p>
                </div>
            </div>
        ''')
    
    bom_info = station.get_current_bom_info()
    
    # Determine BOM type and quantity with split information
    if station.show_single_unit_bom:
        bom_type = 'Single Unit'
        quantity = 1
    elif station.show_batch_bom:
        bom_type = f'{station.product_quantity} Units'
        quantity = station.product_quantity
    else:
        bom_type = station.current_stage.display_name if station.current_stage else 'Stage BOM'
        quantity = station.product_quantity
    
    # Add split information for display
    split_context = {'is_split': False}
    if bom_info and bom_info.get('is_split'):
        split_info = bom_info.get('split_info', {})
        current_display_info = split_info.get(f'display_{station.display_number}', {})
        
        split_context = {
            'is_split': True,
            'display_info': bom_info.get('display_info', ''),
            'current_display': station.display_number,
            'item_count': current_display_info.get('item_count', 0),
            'start_serial': current_display_info.get('start_serial', 0),
            'end_serial': current_display_info.get('end_serial', 0),
        }
        
        if current_display_info.get('item_count', 0) > 0:
            bom_type += f" (Items {current_display_info['start_serial']}-{current_display_info['end_serial']})"
        else:
            bom_type += " (No items for this display)"
    
    context = {
        'station': station,
        'bom_data': bom_data,
        'bom_type': bom_type,
        'quantity': quantity,
        'product': station.current_product,
        'stage': station.current_stage,
        'bom_info': bom_info,
        'is_split': split_context['is_split'],
        'split_context': split_context,
        'has_bom_data': bool(bom_data),
    }
    
    return render(request, 'bom_slider_fragment.html', context)

 
# BOM Management Dashboard
def bom_management_dashboard(request):
    """BOM management dashboard"""
    context = {
        'total_items': BOMItem.objects.filter(is_active=True).count(),
        'total_templates': BOMTemplate.objects.filter(is_active=True).count(),
        'total_products': Product.objects.count(),
        'recent_items': BOMItem.objects.filter(is_active=True).order_by('-created_date')[:10],
        'recent_templates': BOMTemplate.objects.filter(is_active=True).order_by('-created_date')[:5],
        'templates_by_type': {},
    }
    
    # Group templates by type
    templates = BOMTemplate.objects.filter(is_active=True).select_related('product', 'stage')
    for template in templates:
        if template.bom_type not in context['templates_by_type']:
            context['templates_by_type'][template.bom_type] = []
        context['templates_by_type'][template.bom_type].append(template)
    
    return render(request, 'bom_management_dashboard.html', context)

def bom_item_management(request):
    """BOM item management page"""
    search = request.GET.get('search', '').strip()
    unit_filter = request.GET.get('unit', '')
    supplier_filter = request.GET.get('supplier', '')
    
    items = BOMItem.objects.filter(is_active=True)
    
    if search:
        items = items.filter(
            item_description__icontains=search
        ) | items.filter(
            item_code__icontains=search
        ) | items.filter(
            part_number__icontains=search
        )
    
    if unit_filter:
        items = items.filter(unit_of_measure=unit_filter)
    
    if supplier_filter:
        items = items.filter(supplier__icontains=supplier_filter)
    
    items = items.order_by('item_description')
    
    # Get filter options
    units = BOMItem.objects.filter(is_active=True).values_list('unit_of_measure', flat=True).distinct()
    suppliers = BOMItem.objects.filter(is_active=True, supplier__isnull=False).exclude(supplier='').values_list('supplier', flat=True).distinct()
    
    context = {
        'items': items,
        'search': search,
        'unit_filter': unit_filter,
        'supplier_filter': supplier_filter,
        'units': sorted(units),
        'suppliers': sorted(suppliers),
        'total_items': items.count(),
    }
    
    return render(request, 'bom_item_management.html', context)


def bom_template_management(request):
    """BOM template management page"""
    templates = BOMTemplate.objects.filter(is_active=True).select_related('product', 'stage').prefetch_related('bom_items__item')
    
    context = {
        'templates': templates,
        'products': Product.objects.all(),
        'stages': AssemblyStage.objects.all(),
    }
    
    return render(request, 'bom_template_management.html', context)

@csrf_exempt
@require_http_methods(["POST"])
def bulk_update_bom_items(request):
    """Bulk update BOM items via AJAX"""
    try:
        data = json.loads(request.body)
        updates = data.get('updates', [])
        
        updated_count = 0
        for update in updates:
            item_id = update.get('id')
            if not item_id:
                continue
                
            try:
                item = BOMItem.objects.get(id=item_id)
                
                # Update fields if provided
                if 'cost_per_unit' in update:
                    item.cost_per_unit = update['cost_per_unit']
                if 'supplier' in update:
                    item.supplier = update['supplier']
                if 'is_active' in update:
                    item.is_active = update['is_active']
                    
                item.save()
                updated_count += 1
                
            except BOMItem.DoesNotExist:
                continue
        
        return JsonResponse({
            'success': True,
            'message': f'Updated {updated_count} items successfully',
            'updated_count': updated_count
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def export_all_bom_items_csv(request):
    """Export all BOM items to CSV"""
    import csv
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="all_bom_items.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Item Code', 'Description', 'Part Number', 'Unit', 
        'Supplier', 'Cost per Unit', 'Weight per Unit', 'Active', 'Created Date'
    ])
    
    items = BOMItem.objects.all().order_by('item_code')
    for item in items:
        writer.writerow([
            item.item_code,
            item.item_description,
            item.part_number,
            item.unit_of_measure,
            item.supplier or '',
            item.cost_per_unit or '',
            item.weight_per_unit or '',
            'Yes' if item.is_active else 'No',
            item.created_date.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    return response


from django.template import loader
from django.http import StreamingHttpResponse

def station_media_slider_enhanced(request, station_id):
    """Enhanced media slider with BOM integration — now streaming"""
    station = get_object_or_404(Station, pk=station_id)
    request.session['last_station_id'] = station_id

    current_media = station.get_current_media()
    bom_data = station.get_current_bom_data() or []
    has_bom_data = bool(bom_data)

    context = {
        'station': station,
        'selected_media': current_media,
        'current_product': station.current_product,
        'current_stage': station.current_stage,
        'current_process': station.current_process,
        'product_quantity': station.product_quantity,
        'clicker_enabled': station.clicker_enabled,
        'loop_mode': station.loop_mode,
        'next_process': station.get_next_process(),
        'previous_process': station.get_previous_process(),
        'display_number': station.display_number,
        'has_bom_data': has_bom_data,
        'bom_item_count': len(bom_data),
        'bom_data': bom_data,
    }

    template = loader.get_template('brg_station_slider_enhanced.html')

    def stream_template():
        # Render in chunks so browser can start processing early
        yield template.render(context, request)

    return StreamingHttpResponse(stream_template(), content_type='text/html')

   
   
 

@csrf_exempt
@require_http_methods(["POST"])
def quick_add_bom_item(request):
    """Quick add BOM item via AJAX"""
    try:
        data = json.loads(request.body)
        
        required_fields = ['item_code', 'item_description', 'part_number']
        if not all(field in data for field in required_fields):
            return JsonResponse({'error': 'Missing required fields'}, status=400)
        
        # Check if item already exists
        if BOMItem.objects.filter(item_code=data['item_code']).exists():
            return JsonResponse({'error': 'Item code already exists'}, status=400)
        
        item = BOMItem.objects.create(
            item_code=data['item_code'],
            item_description=data['item_description'],
            part_number=data['part_number'],
            unit_of_measure=data.get('unit_of_measure', 'NO.'),
            supplier=data.get('supplier', ''),
            cost_per_unit=data.get('cost_per_unit'),
            is_active=True
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Item created successfully',
            'item': {
                'id': item.id,
                'item_code': item.item_code,
                'item_description': item.item_description,
                'part_number': item.part_number,
                'unit_of_measure': item.unit_of_measure,
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def export_bom_csv(request, template_id):
    """Export BOM template to CSV"""
    import csv
    from django.http import HttpResponse
    
    template = get_object_or_404(BOMTemplate, pk=template_id)
    quantity = int(request.GET.get('quantity', 1))
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{template.template_name}_qty_{quantity}.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow([
        'S.NO', 'Item Code', 'Item Description', 'Part Number', 
        'Quantity', 'Unit', 'Supplier', 'Cost per Unit', 'Line Cost', 'Notes'
    ])
    
    # Write BOM data
    bom_data = template.generate_bom_for_quantity(quantity)
    total_cost = 0
    
    for item_data in bom_data:
        line_cost = (item_data['calculated_quantity'] * 
                    (item_data['item'].cost_per_unit or 0))
        total_cost += line_cost
        
        writer.writerow([
            item_data['serial_number'],
            item_data['item'].item_code,
            item_data['item'].item_description,
            item_data['item'].part_number,
            item_data['calculated_quantity'],
            item_data['item'].unit_of_measure,
            item_data['item'].supplier or '',
            item_data['item'].cost_per_unit or '',
            f"{line_cost:.2f}" if line_cost > 0 else '',
            item_data['notes'] or '',
        ])
    
    # Write summary
    writer.writerow([])
    writer.writerow(['Summary'])
    writer.writerow(['Total Items:', len(bom_data)])
    writer.writerow(['Total Cost:', f"{total_cost:.2f}"])
    writer.writerow(['Quantity:', quantity])
    writer.writerow(['Template:', template.template_name])
    writer.writerow(['Product:', f"{template.product.code} - {template.product.name}"])
    writer.writerow(['Generated:', time.strftime('%Y-%m-%d %H:%M:%S')])
    
    return response





def debug_station(request, station_id):
    """Debug view to test station setup"""
    try:
        station = Station.objects.get(pk=station_id)
        
        debug_info = {
            'station_id': station.id,
            'station_name': station.name,
            'display_number': station.display_number,
            'current_product': {
                'id': station.current_product.id if station.current_product else None,
                'code': station.current_product.code if station.current_product else None,
                'name': station.current_product.name if station.current_product else None,
            },
            'current_stage': {
                'id': station.current_stage.id if station.current_stage else None,
                'name': station.current_stage.display_name if station.current_stage else None,
            },
            'current_process': {
                'id': station.current_process.id if station.current_process else None,
                'name': station.current_process.name if station.current_process else None,
                'display_name': station.current_process.display_name if station.current_process else None,
            },
            'media_count': 0,
            'media_files': [],
            'error': None
        }
        
        # Test get_current_media
        try:
            media = station.get_current_media()
            debug_info['media_count'] = media.count()
            debug_info['media_files'] = [
                {
                    'id': m.id,
                    'file': str(m.file),
                    'type': m.media_type,
                    'display_1': m.display_screen_1,
                    'display_2': m.display_screen_2,
                    'display_3': m.display_screen_3,
                }
                for m in media
            ]
        except Exception as e:
            debug_info['error'] = str(e)
        
        return JsonResponse(debug_info)
        
    except Station.DoesNotExist:
        return JsonResponse({'error': 'Station not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
    

def station_media_stream_debug(request, station_id):
    """Simplified streaming for debugging"""
    def event_stream():
        try:
            # Test 1: Basic connection
            yield f"data: {json.dumps({'test': 'connection_ok', 'timestamp': time.time()})}\n\n"
            
            # Test 2: Get station
            station = Station.objects.get(pk=station_id)
            yield f"data: {json.dumps({'test': 'station_found', 'station_name': station.name})}\n\n"
            
            # Test 3: Get media
            current_media = station.get_current_media()
            yield f"data: {json.dumps({'test': 'media_found', 'media_count': current_media.count()})}\n\n"
            
            # Test 4: Try to serialize one media item
            if current_media.exists():
                media = current_media.first()
                media_data = {
                    'id': media.id,
                    'url': media.file.url,
                    'type': media.file.name.split('.')[-1].lower(),
                    'duration': media.duration,
                    'media_type': media.get_media_type_display(),
                }
                yield f"data: {json.dumps({'test': 'media_serialized', 'media': media_data})}\n\n"
            
            # Test 5: Full data structure
            counter = 0
            while counter < 3:  # Only run 3 times for testing
                try:
                    station = Station.objects.get(pk=station_id)
                    current_media = station.get_current_media()
                    
                    media_data = []
                    for media in current_media:
                        media_info = {
                            'id': media.id,
                            'url': media.file.url,
                            'type': media.file.name.split('.')[-1].lower(),
                            'duration': media.duration,
                            'media_type': media.get_media_type_display(),
                        }
                        
                        # Safely add process info
                        if media.process:
                            try:
                                media_info['process'] = {
                                    'id': media.process.id,
                                    'name': media.process.name,
                                    'display_name': media.process.display_name,
                                }
                            except Exception as pe:
                                media_info['process_error'] = str(pe)
                        
                        # Safely add BOM info
                        if media.bom:
                            try:
                                media_info['bom'] = {
                                    'id': media.bom.id,
                                    'type': media.bom.get_bom_type_display(),
                                }
                            except Exception as be:
                                media_info['bom_error'] = str(be)
                        
                        media_data.append(media_info)
                    
                    # Assembly info
                    assembly_info = {}
                    try:
                        if station.current_product:
                            assembly_info['current_product'] = {
                                'id': station.current_product.id,
                                'code': station.current_product.code,
                                'name': station.current_product.name
                            }
                    except Exception as pe:
                        assembly_info['product_error'] = str(pe)
                    
                    try:
                        if station.current_stage:
                            assembly_info['current_stage'] = {
                                'id': station.current_stage.id,
                                'name': station.current_stage.display_name,
                                'order': station.current_stage.order
                            }
                    except Exception as se:
                        assembly_info['stage_error'] = str(se)
                    
                    try:
                        if station.current_process:
                            assembly_info['current_process'] = {
                                'id': station.current_process.id,
                                'name': station.current_process.name,
                                'display_name': station.current_process.display_name,
                                'order': station.current_process.order,
                                'is_looped': station.current_process.is_looped,
                                'loop_group': station.current_process.loop_group
                            }
                    except Exception as pe:
                        assembly_info['process_error'] = str(pe)
                    
                    response_data = {
                        'media': media_data,
                        'station_name': station.name,
                        'assembly': assembly_info,
                        'timestamp': time.time(),
                        'counter': counter
                    }
                    
                    yield f"data: {json.dumps(response_data)}\n\n"
                    counter += 1
                    time.sleep(2)
                    
                except Exception as inner_e:
                    yield f"data: {json.dumps({'error': f'Inner loop error: {str(inner_e)}'})}\n\n"
                    break
                    
        except Exception as e:
            yield f"data: {json.dumps({'error': f'Stream error: {str(e)}'})}\n\n"

    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response





# views.py
from django.contrib.auth.decorators import login_required

from django.shortcuts import render

from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def supervisor_dashboard(request):
    return render(request, 'supervisor-dashboard.html')

@login_required
def workflow_guide(request):
    return render(request, 'workflow-guide.html')
def workflow_guide2(request):
    return render(request, 'hindi.html')




def debug_bom_stage(request, station_id):
    """Debug endpoint to check BOM stage configuration"""
    try:
        station = Station.objects.get(pk=station_id)
        
        # Check BOM stage and processes
        bom_stage = None
        bom_processes = []
        try:
            bom_stage = AssemblyStage.objects.get(name='BOM_DISPLAY')
            bom_processes = list(bom_stage.processes.all().values(
                'id', 'name', 'display_name', 'order'
            ))
        except AssemblyStage.DoesNotExist:
            pass
        
        # Check current BOM data
        bom_data = station.get_current_bom_data() if station.current_product else []
        bom_info = station.get_current_bom_info()
        
        # Check BOM templates
        bom_templates = []
        if station.current_product:
            from .models import BOMTemplate
            templates = BOMTemplate.objects.filter(
                product=station.current_product,
                is_active=True
            ).values('id', 'template_name', 'bom_type', 'stage__name')
            bom_templates = list(templates)
        
        debug_info = {
            'station': {
                'id': station.id,
                'name': station.name,
                'display_number': station.display_number,
                'current_product': {
                    'id': station.current_product.id if station.current_product else None,
                    'code': station.current_product.code if station.current_product else None,
                    'name': station.current_product.name if station.current_product else None,
                },
                'current_stage': {
                    'id': station.current_stage.id if station.current_stage else None,
                    'name': station.current_stage.name if station.current_stage else None,
                    'display_name': station.current_stage.display_name if station.current_stage else None,
                    'is_bom_stage': (station.current_stage and 
                                   station.current_stage.name == 'BOM_DISPLAY')
                },
                'current_process': {
                    'id': station.current_process.id if station.current_process else None,
                    'name': station.current_process.name if station.current_process else None,
                    'display_name': station.current_process.display_name if station.current_process else None,
                },
                'bom_settings': {
                    'show_single_unit': station.show_single_unit_bom,
                    'show_batch': station.show_batch_bom,
                    'product_quantity': station.product_quantity,
                }
            },
            'bom_stage': {
                'exists': bom_stage is not None,
                'stage_data': {
                    'id': bom_stage.id if bom_stage else None,
                    'name': bom_stage.name if bom_stage else None,
                    'display_name': bom_stage.display_name if bom_stage else None,
                } if bom_stage else None,
                'processes': bom_processes,
                'process_count': len(bom_processes)
            },
            'bom_data': {
                'available': bool(bom_data),
                'item_count': len(bom_data) if bom_data else 0,
                'is_split': bom_info.get('is_split', False) if bom_info else False,
                'display_info': bom_info.get('display_info', '') if bom_info else '',
                'bom_info': bom_info,
            },
            'bom_templates': {
                'available': len(bom_templates) > 0,
                'count': len(bom_templates),
                'templates': bom_templates,
            },
            'recommendations': []
        }
        
        # Add recommendations
        if not bom_stage:
            debug_info['recommendations'].append(
                "Create BOM_DISPLAY stage with processes for single unit, batch, and stage BOMs"
            )
        
        if not bom_data and station.current_product:
            debug_info['recommendations'].append(
                f"Create BOM templates for product {station.current_product.code}"
            )
        
        if station.current_stage and station.current_stage.name == 'BOM_DISPLAY' and not bom_data:
            debug_info['recommendations'].append(
                "Configure BOM templates or check station BOM settings"
            )
        
        return JsonResponse(debug_info, json_dumps_params={'indent': 2})
        
    except Station.DoesNotExist:
        return JsonResponse({'error': 'Station not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)











# views.py - Cleaned with Assembly Stage CRUD and Product-Filtered Data
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.db.models import Prefetch, Count
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.core.paginator import Paginator
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
import json
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from PIL import Image
import io
import os
import zipfile
import tempfile
from pathlib import Path
from .models import (
    Product, Station, ProductMedia, BOMTemplate, BOMTemplateItem, 
    AssemblyProcess, AssemblyStage, BillOfMaterial, BOMItem
)

def product_information_view(request):
    """Main view to display product information page with CRUD operations"""
    products = Product.objects.all().order_by('code')
    selected_product = None
    product_data = {}
    
    # Get selected product from request
    product_id = request.GET.get('product_id')
    if product_id:
        try:
            selected_product = get_object_or_404(Product, id=product_id)
            product_data = get_product_data(selected_product)
        except Exception as e:
            messages.error(request, f"Error loading product data: {e}")
            product_data = {}
    
    # Get additional data for dropdowns - FILTER BY SELECTED PRODUCT
    if selected_product:
        # Filter stages and processes by selected product
        assembly_stages = AssemblyStage.objects.filter(
            product=selected_product
        ).order_by('order')
        
        assembly_processes = AssemblyProcess.objects.filter(
            stage__product=selected_product
        ).select_related('stage').order_by('stage__order', 'order')
        
    else:
        # If no product selected, show all (for initial load)
        assembly_stages = AssemblyStage.objects.select_related('product').order_by('product__code', 'order')
        assembly_processes = AssemblyProcess.objects.select_related('stage', 'stage__product').order_by('stage__product__code', 'stage__order', 'order')
    
    # BOM items remain global as they can be reused across products
    bom_items = BOMItem.objects.filter(is_active=True).order_by('item_description')
    
    context = {
        'products': products,
        'selected_product': selected_product,
        'product_data': product_data,
        'assembly_stages': assembly_stages,
        'assembly_processes': assembly_processes,
        'bom_items': bom_items,
        'bom_type_choices': BOMTemplate.BOM_TYPE_CHOICES,
        'media_type_choices': ProductMedia.MEDIA_TYPE_CHOICES,
    }
    
    return render(request, 'assembly/product_information.html', context)

def get_product_data(product):
    """Get all data related to a specific product - ENHANCED VERSION"""
    
    # 1. Connected Stations
    active_stations = Station.objects.filter(
        current_product=product
    ).select_related(
        'current_stage', 'current_process', 'manager'
    ).order_by('display_number')
    
    available_stations = Station.objects.filter(
        products=product
    ).select_related('manager').order_by('display_number')
    
    # 2. Product Media
    media_files = ProductMedia.objects.filter(
        product=product
    ).select_related('process', 'bom').order_by('media_type', 'id')
    
    # 3. BOM Templates with Items
    bom_templates = BOMTemplate.objects.filter(
        product=product,
        is_active=True
    ).prefetch_related(
        Prefetch(
            'bom_items', 
            queryset=BOMTemplateItem.objects.filter(is_active=True)
                .select_related('item')
                .order_by('serial_number')
        )
    ).select_related('stage').order_by('bom_type')
    
    # 4. Assembly Stages and Processes - FILTER BY PRODUCT
    assembly_stages = AssemblyStage.objects.filter(
        product=product
    ).prefetch_related(
        Prefetch(
            'processes', 
            queryset=AssemblyProcess.objects.order_by('order')
        )
    ).order_by('order')
    
    # 5. Bill of Materials (PDF/Legacy)
    legacy_boms = BillOfMaterial.objects.filter(
        product=product
    ).select_related('stage', 'bom_template').order_by('bom_type')
    
    # Group media by type for easier display
    media_by_type = {}
    for media_item in media_files:
        media_type = media_item.get_media_type_display()
        if media_type not in media_by_type:
            media_by_type[media_type] = []
        media_by_type[media_type].append(media_item)
    
    # Group BOM templates by type
    templates_by_type = {}
    for template in bom_templates:
        template_type = template.get_bom_type_display()
        if template_type not in templates_by_type:
            templates_by_type[template_type] = []
        templates_by_type[template_type].append(template)
    
    # Group stages with their processes for better organization
    stages_with_processes = []
    for stage in assembly_stages:
        stage_info = {
            'stage': stage,
            'processes': stage.processes.all().order_by('order'),
            'process_count': stage.processes.count(),
            'bom_templates': bom_templates.filter(stage=stage)
        }
        stages_with_processes.append(stage_info)
    
    # Prepare context data
    context_data = {
        'product': product,
        'active_stations': active_stations,
        'available_stations': available_stations,
        'media_files': media_files,
        'bom_templates': bom_templates,
        'assembly_stages': assembly_stages,
        'legacy_boms': legacy_boms,
        'media_by_type': media_by_type,
        'templates_by_type': templates_by_type,
        'stages_with_processes': stages_with_processes,
        'station_summary': {
            'total_active': active_stations.count(),
            'total_available': available_stations.count(),
            'displays_in_use': list(active_stations.values_list('display_number', flat=True)),
        },
        'summary_stats': {
            'total_stages': assembly_stages.count(),
            'total_processes': sum(stage.processes.count() for stage in assembly_stages),
            'total_media_files': media_files.count(),
            'total_bom_templates': bom_templates.count(),
            'active_bom_templates': bom_templates.filter(is_active=True).count(),
        }
    }
    
    return context_data

# ==============================================
# PRODUCT CRUD OPERATIONS (NEW)
# ==============================================
from django.db import transaction

@require_http_methods(["POST"])
def create_product(request):
    """Create a new product"""
    try:
        code = request.POST.get('code', '').strip().upper()
        name = request.POST.get('name', '').strip()
        
        # Validation
        if not code:
            return JsonResponse({'success': False, 'error': 'Product code is required'})
        if not name:
            return JsonResponse({'success': False, 'error': 'Product name is required'})
        
        # Check if code already exists
        if Product.objects.filter(code=code).exists():
            return JsonResponse({'success': False, 'error': f'Product code "{code}" already exists'})
        
        with transaction.atomic():
            product = Product.objects.create(
                code=code,
                name=name
            )
        
        return JsonResponse({
            'success': True, 
            'message': f'Product "{product.code}" created successfully!',
            'product_id': product.id,
            'product_code': product.code,
            'product_name': product.name
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error creating product: {str(e)}'})

@require_http_methods(["POST"])
def update_product(request, product_id):
    """Update an existing product"""
    try:
        product = get_object_or_404(Product, id=product_id)
        
        code = request.POST.get('code', '').strip().upper()
        name = request.POST.get('name', '').strip()
        
        # Validation
        if not code:
            return JsonResponse({'success': False, 'error': 'Product code is required'})
        if not name:
            return JsonResponse({'success': False, 'error': 'Product name is required'})
        
        # Check if code already exists (excluding current product)
        if Product.objects.filter(code=code).exclude(id=product_id).exists():
            return JsonResponse({'success': False, 'error': f'Product code "{code}" already exists'})
        
        product.code = code
        product.name = name
        product.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Product updated successfully!',
            'product_code': product.code,
            'product_name': product.name
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_product(request, product_id):
    """Delete a product"""
    try:
        product = get_object_or_404(Product, id=product_id)
        product_code = product.code
        
        # Check if product has any associated data
        stage_count = product.assembly_product.count()
        bom_count = product.bom_templates.count()
        media_count = product.media.count()
        station_count = product.stations.count()
        
        total_associations = stage_count + bom_count + media_count + station_count
        
        if total_associations > 0:
            details = []
            if stage_count > 0:
                details.append(f"{stage_count} assembly stage(s)")
            if bom_count > 0:
                details.append(f"{bom_count} BOM template(s)")
            if media_count > 0:
                details.append(f"{media_count} media file(s)")
            if station_count > 0:
                details.append(f"{station_count} station(s)")
            
            return JsonResponse({
                'success': False, 
                'error': f'Cannot delete product "{product_code}" because it has associated data: {", ".join(details)}. Please remove all associated data first.'
            })
        
        product.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Product "{product_code}" deleted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_product_data_api(request, product_id):
    """Get product data for editing"""
    try:
        product = get_object_or_404(Product, id=product_id)
        product_data = {
            'id': product.id,
            'code': product.code,
            'name': product.name,
            'total_stages': product.assembly_product.count(),
            'total_bom_templates': product.bom_templates.count(),
            'total_media_files': product.media.count(),
            'total_stations': product.stations.count()
        }
        return JsonResponse({'success': True, 'product': product_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ==============================================
# ASSEMBLY STAGE CRUD OPERATIONS (EXISTING)
# ==============================================

@require_http_methods(["POST"])
def create_assembly_stage(request):
    """Create a new assembly stage"""
    try:
        product_id = request.POST.get('product_id')
        product = get_object_or_404(Product, id=product_id)
        
        stage = AssemblyStage.objects.create(
            product=product,
            name=request.POST.get('name'),
            display_name=request.POST.get('display_name', ''),
            order=int(request.POST.get('order', 1)),
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Assembly stage created successfully!',
            'stage_id': stage.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def update_assembly_stage(request, stage_id):
    """Update an existing assembly stage"""
    try:
        stage = get_object_or_404(AssemblyStage, id=stage_id)
        
        # Optional: Validate product context if needed
        product_id = request.POST.get('product_id')
        if product_id:
            product = get_object_or_404(Product, id=product_id)
            if stage.product != product:
                raise ValueError("Stage does not belong to the specified product")
        
        stage.name = request.POST.get('name')
        stage.display_name = request.POST.get('display_name', '')
        stage.order = int(request.POST.get('order', 1))
        stage.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Assembly stage updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_assembly_stage(request, stage_id):
    """Delete an assembly stage"""
    try:
        stage = get_object_or_404(AssemblyStage, id=stage_id)
        stage_name = stage.name
        
        # Check if stage has processes
        process_count = stage.processes.count()
        if process_count > 0:
            return JsonResponse({
                'success': False, 
                'error': f'Cannot delete stage "{stage_name}" because it has {process_count} associated process(es). Delete the processes first.'
            })
        
        stage.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Assembly stage "{stage_name}" deleted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_assembly_stage_data(request, stage_id):
    """Get assembly stage data for editing"""
    try:
        stage = get_object_or_404(AssemblyStage, id=stage_id)
        stage_data = {
            'id': stage.id,
            'product_id': stage.product.id if stage.product else None,
            'name': stage.name,
            'display_name': stage.display_name,
            'order': stage.order,
            'process_count': stage.processes.count()
        }
        return JsonResponse({'success': True, 'stage': stage_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    



# ==============================================
# ASSEMBLY PROCESS CRUD OPERATIONS
# ==============================================

@require_http_methods(["POST"])
def create_assembly_process(request):
    """Create a new assembly process"""
    try:
        stage_id = request.POST.get('stage_id')
        stage = get_object_or_404(AssemblyStage, id=stage_id)
        
        # Optional: Validate product context if needed
        product_id = request.POST.get('product_id')
        if product_id:
            product = get_object_or_404(Product, id=product_id)
            if stage.product != product:
                raise ValueError("Stage does not belong to the specified product")
        
        process = AssemblyProcess.objects.create(
            stage=stage,
            name=request.POST.get('name'),
            display_name=request.POST.get('display_name', ''),
            location=request.POST.get('location', ''),
            order=int(request.POST.get('order', 1)),
            is_looped=request.POST.get('is_looped') == 'on',
            loop_group=request.POST.get('loop_group', ''),
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Assembly process created successfully!',
            'process_id': process.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def update_assembly_process(request, process_id):
    """Update an existing assembly process"""
    try:
        process = get_object_or_404(AssemblyProcess, id=process_id)
        
        stage_id = request.POST.get('stage_id')
        stage = get_object_or_404(AssemblyStage, id=stage_id)
        
        process.stage = stage
        process.name = request.POST.get('name')
        process.display_name = request.POST.get('display_name', '')
        process.location = request.POST.get('location', '')
        process.order = int(request.POST.get('order', 1))
        process.is_looped = request.POST.get('is_looped') == 'on'
        process.loop_group = request.POST.get('loop_group', '')
        process.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Assembly process updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_assembly_process(request, process_id):
    """Delete an assembly process"""
    try:
        process = get_object_or_404(AssemblyProcess, id=process_id)
        process_name = process.name
        process.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Assembly process "{process_name}" deleted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_assembly_process_data(request, process_id):
    """Get assembly process data for editing"""
    try:
        process = get_object_or_404(AssemblyProcess, id=process_id)
        process_data = {
            'id': process.id,
            'stage_id': process.stage.id if process.stage else None,
            'name': process.name,
            'display_name': process.display_name,
            'location': process.location,
            'order': process.order,
            'is_looped': process.is_looped,
            'loop_group': process.loop_group
        }
        return JsonResponse({'success': True, 'process': process_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ==============================================
# BOM TEMPLATE CRUD OPERATIONS
# ==============================================

@require_http_methods(["POST"])
def create_bom_template(request):
    """Create a new BOM template"""
    try:
        product_id = request.POST.get('product_id')
        product = get_object_or_404(Product, id=product_id)
        
        # Get stage if provided - VALIDATE IT BELONGS TO THE PRODUCT
        stage_id = request.POST.get('stage_id')
        stage = None
        if stage_id:
            stage = get_object_or_404(AssemblyStage, id=stage_id, product=product)
        
        bom_template = BOMTemplate.objects.create(
            product=product,
            bom_type=request.POST.get('bom_type'),
            stage=stage,
            template_name=request.POST.get('template_name'),
            description=request.POST.get('description', ''),
            duration=int(request.POST.get('duration', 20)),
            is_duration_active=request.POST.get('is_duration_active') == 'on',
            display_screen_1=request.POST.get('display_screen_1') == 'on',
            display_screen_2=request.POST.get('display_screen_2') == 'on',
            display_screen_3=request.POST.get('display_screen_3') == 'on',
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'BOM Template created successfully!',
            'template_id': bom_template.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
   

 
@require_http_methods(["POST"])
def update_bom_template(request, template_id):
    """Update an existing BOM template"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        
        # Get stage if provided - VALIDATE IT BELONGS TO THE PRODUCT
        stage_id = request.POST.get('stage_id')
        stage = None
        if stage_id:
            stage = get_object_or_404(AssemblyStage, id=stage_id, product=template.product)
        
        template.bom_type = request.POST.get('bom_type')
        template.stage = stage
        template.template_name = request.POST.get('template_name')
        template.description = request.POST.get('description', '')
        template.duration = int(request.POST.get('duration', 20))
        template.is_duration_active = request.POST.get('is_duration_active') == 'on'
        template.display_screen_1 = request.POST.get('display_screen_1') == 'on'
        template.display_screen_2 = request.POST.get('display_screen_2') == 'on'
        template.display_screen_3 = request.POST.get('display_screen_3') == 'on'
        template.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'BOM Template updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_bom_template(request, template_id):
    """Delete a BOM template"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        template_name = template.template_name
        template.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'BOM Template "{template_name}" deleted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_bom_template_data(request, template_id):
    """Get BOM template data for editing"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        template_data = {
            'id': template.id,
            'template_name': template.template_name,
            'bom_type': template.bom_type,
            'stage_id': template.stage.id if template.stage else None,
            'description': template.description,
            'duration': template.duration,
            'is_duration_active': template.is_duration_active,
            'display_screen_1': template.display_screen_1,
            'display_screen_2': template.display_screen_2,
            'display_screen_3': template.display_screen_3,
        }
        return JsonResponse({'success': True, 'template': template_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ==============================================
# BOM TEMPLATE ITEM CRUD OPERATIONS
# ==============================================

@require_http_methods(["POST"])
def create_bom_template_item(request):
    """Create a new BOM template item"""
    try:
        template_id = request.POST.get('template_id')
        template = get_object_or_404(BOMTemplate, id=template_id)
        
        item_id = request.POST.get('item_id')
        item = get_object_or_404(BOMItem, id=item_id)
        
        bom_item = BOMTemplateItem.objects.create(
            bom_template=template,
            item=item,
            base_quantity=float(request.POST.get('base_quantity', 1)),
            serial_number=int(request.POST.get('serial_number')),
            notes=request.POST.get('notes', ''),
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'BOM item added successfully!',
            'item_id': bom_item.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def update_bom_template_item(request, item_id):
    """Update an existing BOM template item"""
    try:
        bom_item = get_object_or_404(BOMTemplateItem, id=item_id)
        
        item_id = request.POST.get('item_id')
        item = get_object_or_404(BOMItem, id=item_id)
        
        bom_item.item = item
        bom_item.base_quantity = float(request.POST.get('base_quantity', 1))
        bom_item.serial_number = int(request.POST.get('serial_number'))
        bom_item.notes = request.POST.get('notes', '')
        bom_item.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'BOM item updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_bom_template_item(request, item_id):
    """Delete a BOM template item"""
    try:
        bom_item = get_object_or_404(BOMTemplateItem, id=item_id)
        item_description = bom_item.item.item_description
        bom_item.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'BOM item "{item_description}" removed successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_bom_template_item_data(request, item_id):
    """Get BOM template item data for editing"""
    try:
        item = get_object_or_404(BOMTemplateItem, id=item_id)
        item_data = {
            'id': item.id,
            'item_id': item.item.id,
            'serial_number': item.serial_number,
            'base_quantity': float(item.base_quantity),
            'notes': item.notes
        }
        return JsonResponse({'success': True, 'item': item_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_bom_template_items(request, template_id):
    """Get BOM template items with images"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        items = template.bom_items.filter(is_active=True).order_by('serial_number')
        
        items_data = [{
            'id': item.id,
            'serial_number': item.serial_number,
            'item_code': item.item.item_code,
            'item_description': item.item.item_description,
            'part_number': item.item.part_number,
            'base_quantity': float(item.base_quantity),
            'unit_of_measure': item.item.unit_of_measure,
            'notes': item.notes,
            'item_photo_url': item.item.item_photo.url if item.item.item_photo else None,
            'has_photo': bool(item.item.item_photo)
        } for item in items]
        
        return JsonResponse({'success': True, 'items': items_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ==============================================
# PRODUCT MEDIA CRUD OPERATIONS
# ==============================================

@require_http_methods(["POST"])
def create_product_media(request):
    """Create a new product media"""
    try:
        product_id = request.POST.get('product_id')
        product = get_object_or_404(Product, id=product_id)
        
        # Get process if provided - VALIDATE IT BELONGS TO THE PRODUCT
        process_id = request.POST.get('process_id')
        process = None
        if process_id:
            process = get_object_or_404(AssemblyProcess, id=process_id, stage__product=product)
        
        # Get BOM if provided - VALIDATE IT BELONGS TO THE PRODUCT
        bom_id = request.POST.get('bom_id')
        bom = None
        if bom_id:
            bom = get_object_or_404(BillOfMaterial, id=bom_id, product=product)
        
        media = ProductMedia.objects.create(
            product=product,
            process=process,
            bom=bom,
            media_type=request.POST.get('media_type'),
            file=request.FILES.get('file'),
            duration=int(request.POST.get('duration', 15)),
            display_screen_1=request.POST.get('display_screen_1') == 'on',
            display_screen_2=request.POST.get('display_screen_2') == 'on',
            display_screen_3=request.POST.get('display_screen_3') == 'on',
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Product media created successfully!',
            'media_id': media.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
   
@require_http_methods(["POST"])
def update_product_media(request, media_id):
    """Update an existing product media"""
    try:
        media = get_object_or_404(ProductMedia, id=media_id)
        
        # Get process if provided - VALIDATE IT BELONGS TO THE PRODUCT
        process_id = request.POST.get('process_id')
        process = None
        if process_id:
            process = get_object_or_404(AssemblyProcess, id=process_id, stage__product=media.product)
        
        # Get BOM if provided - VALIDATE IT BELONGS TO THE PRODUCT
        bom_id = request.POST.get('bom_id')
        bom = None
        if bom_id:
            bom = get_object_or_404(BillOfMaterial, id=bom_id, product=media.product)
        
        media.process = process
        media.bom = bom
        media.media_type = request.POST.get('media_type')
        media.duration = int(request.POST.get('duration', 15))
        media.display_screen_1 = request.POST.get('display_screen_1') == 'on'
        media.display_screen_2 = request.POST.get('display_screen_2') == 'on'
        media.display_screen_3 = request.POST.get('display_screen_3') == 'on'
        
        # Update file if provided
        if request.FILES.get('file'):
            media.file = request.FILES.get('file')
            
        media.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Product media updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_product_media(request, media_id):
    """Delete a product media"""
    try:
        media = get_object_or_404(ProductMedia, id=media_id)
        media_name = str(media)
        media.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Product media "{media_name}" deleted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_product_media_data(request, media_id):
    """Get product media data for editing with file preview info"""
    try:
        media = get_object_or_404(ProductMedia, id=media_id)
        
        # Determine file type and preview info
        file_url = None
        file_type = None
        file_size = None
        
        if media.file:
            file_url = media.file.url
            file_name = media.file.name.lower()
            file_size = media.file.size if hasattr(media.file, 'size') else None
            
            # Determine file type
            if file_name.endswith('.pdf'):
                file_type = 'pdf'
            elif file_name.endswith(('.mp4', '.mov', '.avi', '.mkv', '.wmv', '.flv', '.webm')):
                file_type = 'video'
            elif file_name.endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp')):
                file_type = 'image'
            elif file_name.endswith(('.xlsx', '.xls')):
                file_type = 'excel'
            elif file_name.endswith(('.docx', '.doc')):
                file_type = 'document'
            else:
                file_type = 'other'
        
        media_data = {
            'id': media.id,
            'media_type': media.media_type,
            'duration': media.duration,
            'process_id': media.process.id if media.process else None,
            'bom_id': media.bom.id if media.bom else None,
            'display_screen_1': media.display_screen_1,
            'display_screen_2': media.display_screen_2,
            'display_screen_3': media.display_screen_3,
            'file_name': media.file.name if media.file else None,
            'file_url': file_url,
            'file_type': file_type,
            'file_size': file_size,
            'has_file': bool(media.file)
        }
        return JsonResponse({'success': True, 'media': media_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ==============================================
# AJAX HELPER VIEWS
# ==============================================

def get_stages_for_product(request, product_id):
    """Get assembly stages for a specific product (AJAX helper)"""
    try:
        product = get_object_or_404(Product, id=product_id)
        stages = AssemblyStage.objects.filter(product=product).order_by('order')
        
        stage_data = [{
            'id': stage.id,
            'name': stage.name,
            'display_name': stage.display_name,
            'order': stage.order,
            'process_count': stage.processes.count()
        } for stage in stages]
        
        return JsonResponse({
            'success': True,
            'stages': stage_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_processes_for_product(request, product_id):
    """Get all processes for a specific product"""
    try:
        product = get_object_or_404(Product, id=product_id)
        processes = AssemblyProcess.objects.filter(
            stage__product=product
        ).select_related('stage').order_by('stage__order', 'order')
        
        process_data = [{
            'id': process.id,
            'name': process.name,
            'display_name': process.display_name,
            'stage_name': process.stage.name,
            'stage_id': process.stage.id,
            'order': process.order,
            'location': process.location
        } for process in processes]
        
        return JsonResponse({
            'success': True,
            'processes': process_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_processes_for_stage(request, stage_id):
    """Get processes for a specific stage"""
    try:
        stage = get_object_or_404(AssemblyStage, id=stage_id)
        processes = stage.processes.all().order_by('order')
        
        process_data = [{
            'id': process.id,
            'name': process.name,
            'display_name': process.display_name,
            'order': process.order,
            'location': process.location,
            'is_looped': process.is_looped,
            'loop_group': process.loop_group
        } for process in processes]
        
        return JsonResponse({
            'success': True,
            'processes': process_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ==============================================
# BULK UPLOAD OPERATIONS
# ==============================================

@require_http_methods(["POST"])  
def upload_bom_items_excel(request, template_id):
    """Upload BOM items from Excel file to a specific BOM template"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        
        if 'excel_file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No Excel file provided'})
        
        excel_file = request.FILES['excel_file']
        unit_of_measure = request.POST.get('unit_of_measure', 'NO.')
        supplier = request.POST.get('supplier', '')
        overwrite_existing = request.POST.get('overwrite_existing') == 'true'
        
        result = process_excel_file_for_template(
            excel_file, template, unit_of_measure, supplier, overwrite_existing
        )
        
        if result['success']:
            return JsonResponse({
                'success': True,
                'message': f"Successfully processed {result['created']} new items and {result['updated']} updated items. Skipped {result['skipped']} items.",
                'errors': result['errors'] if result['errors'] else None
            })
        else:
            return JsonResponse({'success': False, 'error': result['error']})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def upload_product_media_zip(request):
    """Upload product media from ZIP file"""
    try:
        product_id = request.POST.get('product_id')
        product = get_object_or_404(Product, id=product_id)
        
        if 'zip_file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No ZIP file provided'})
        
        zip_file = request.FILES['zip_file']
        
        # Get form data and validate process belongs to product
        process_id = request.POST.get('process_id')
        process = None
        if process_id:
            process = get_object_or_404(AssemblyProcess, id=process_id, stage__product=product)
        
        bom_id = request.POST.get('bom_id')
        bom = None
        if bom_id:
            bom = get_object_or_404(BillOfMaterial, id=bom_id, product=product)
        
        default_duration = int(request.POST.get('default_duration', 30))
        display_screen_1 = request.POST.get('display_screen_1') == 'true'
        display_screen_2 = request.POST.get('display_screen_2') == 'true'
        display_screen_3 = request.POST.get('display_screen_3') == 'true'
        
        result = process_zip_file_for_media(
            product, zip_file, process, bom, default_duration,
            display_screen_1, display_screen_2, display_screen_3
        )
        
        if result['success']:
            return JsonResponse({
                'success': True,
                'message': f"Successfully processed {result['total_files']} files: {result['pdfs']} PDFs, {result['videos']} videos. Skipped {result['skipped']} files.",
                'errors': result['errors'] if result['errors'] else None
            })
        else:
            return JsonResponse({'success': False, 'error': result['error']})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ==============================================
# UTILITY FUNCTIONS
# ==============================================

def get_available_serial_numbers(request, template_id):
    """Get next available serial number for BOM template"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        used_serials = list(template.bom_items.values_list('serial_number', flat=True))
        
        # Find next available serial number
        next_serial = 1
        while next_serial in used_serials:
            next_serial += 1
            
        return JsonResponse({
            'success': True,
            'next_serial': next_serial,
            'used_serials': used_serials
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ==============================================
# BULK UPLOAD PROCESSING FUNCTIONS
# ==============================================

def process_excel_file_for_template(excel_file, bom_template, default_unit, default_supplier, overwrite_existing):
    """Process Excel file and create BOM items and template items"""
    result = {
        'success': False,
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': []
    }
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = workbook.active
        
        # Load images from the sheet
        image_loader = SheetImageLoader(sheet)
        
        # Find header row
        header_row = None
        for row_num in range(1, 6):
            row_values = [str(cell.value).strip().upper() if cell.value else '' for cell in sheet[row_num]]
            if any('ITEM DESCRIPTION' in val or 'DESCRIPTION' in val for val in row_values):
                header_row = row_num
                break
        
        if header_row is None:
            result['error'] = "Could not find header row with 'ITEM DESCRIPTION' column"
            return result
        
        # Map column indices
        headers = [str(cell.value).strip().upper() if cell.value else '' for cell in sheet[header_row]]
        
        col_mapping = {}
        for idx, header in enumerate(headers):
            if 'S.' in header and 'NO' in header:
                col_mapping['s_no'] = idx
            elif 'CODE' in header and 'ITEM' not in header:  # Just "Code" column, not "Item Code"
                col_mapping['code'] = idx
            elif 'ITEM DESCRIPTION' in header or 'DESCRIPTION' in header:
                col_mapping['description'] = idx
            elif 'PART NO' in header or 'PART_NO' in header:
                col_mapping['part_no'] = idx
            elif 'QTY' in header or 'QUANTITY' in header:
                col_mapping['qty'] = idx
            elif 'UOM' in header or ('UNIT' in header and 'MEASURE' in header):
                col_mapping['uom'] = idx
            elif 'PHOTO' in header or 'IMAGE' in header:
                col_mapping['photo'] = idx
        
        # Validate required columns
        required_cols = ['description', 'part_no']
        missing_cols = [col for col in required_cols if col not in col_mapping]
        if missing_cols:
            result['error'] = f"Missing required columns: {missing_cols}"
            return result
        
        # Process data rows
        for row_num in range(header_row + 1, sheet.max_row + 1):
            row = sheet[row_num]
            
            # Skip empty rows
            if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                continue
            
            try:
                # Extract data
                s_no = str(row[col_mapping.get('s_no', 0)].value).strip() if col_mapping.get('s_no') is not None else str(row_num - header_row)
                code = str(row[col_mapping.get('code', 0)].value).strip() if col_mapping.get('code') is not None and row[col_mapping.get('code', 0)].value else ''
                description = str(row[col_mapping['description']].value).strip() if row[col_mapping['description']].value else ''
                part_no = str(row[col_mapping['part_no']].value).strip() if row[col_mapping['part_no']].value else ''
                qty = str(row[col_mapping.get('qty', 0)].value).strip() if col_mapping.get('qty') is not None and row[col_mapping.get('qty', 0)].value else '1'
                
                # Extract UOM - use from Excel if available, otherwise use default
                uom = ''
                if 'uom' in col_mapping and col_mapping['uom'] is not None:
                    uom_cell_value = row[col_mapping['uom']].value
                    if uom_cell_value:
                        uom = str(uom_cell_value).strip()
                
                # Use extracted UOM if available, otherwise fall back to default_unit
                unit_of_measure = uom if uom else default_unit
                
                # Skip if essential data is missing
                if not description or not part_no:
                    result['skipped'] += 1
                    continue
                
                # Use the Code column if provided, otherwise generate from description
                if code:
                    item_code = code.upper().replace(' ', '_').replace('-', '_')
                    item_code = ''.join(c for c in item_code if c.isalnum() or c == '_')
                else:
                    item_code = description.upper().replace(' ', '_').replace('-', '_')
                    item_code = ''.join(c for c in item_code if c.isalnum() or c == '_')
                
                # Ensure item_code is not empty
                if not item_code:
                    item_code = f"ITEM_{row_num - header_row}"
                
                # Check if BOM item exists
                bom_item = None
                try:
                    bom_item = BOMItem.objects.get(item_code=item_code)
                    if overwrite_existing:
                        # Update existing item
                        bom_item.item_description = description
                        bom_item.part_number = part_no
                        bom_item.unit_of_measure = unit_of_measure  # Use extracted or default UOM
                        if default_supplier:
                            bom_item.supplier = default_supplier
                        
                        # Handle image extraction for existing item
                        if 'photo' in col_mapping:
                            photo_cell = f"{openpyxl.utils.get_column_letter(col_mapping['photo'] + 1)}{row_num}"
                            if image_loader.image_in(photo_cell):
                                try:
                                    image = image_loader.get(photo_cell)
                                    if image:
                                        # Delete old image if exists
                                        if bom_item.item_photo:
                                            bom_item.item_photo.delete(save=False)
                                        
                                        img_io = io.BytesIO()
                                        if image.mode == 'RGBA':
                                            image.save(img_io, format='PNG', optimize=True)
                                            filename = f"{item_code.lower()}.png"
                                        else:
                                            if image.mode not in ['RGB', 'L']:
                                                image = image.convert('RGB')
                                            image.save(img_io, format='JPEG', quality=95, optimize=True)
                                            filename = f"{item_code.lower()}.jpg"
                                        
                                        img_io.seek(0)
                                        image_file = ContentFile(img_io.getvalue(), name=filename)
                                        bom_item.item_photo = image_file
                                except Exception as e:
                                    result['errors'].append(f"Row {row_num}: Could not extract image - {str(e)}")
                        
                        bom_item.save()
                        result['updated'] += 1
                    # If not overwriting, just use the existing item
                except BOMItem.DoesNotExist:
                    # Create new BOM item
                    bom_item_data = {
                        'item_code': item_code,
                        'item_description': description,
                        'part_number': part_no,
                        'unit_of_measure': unit_of_measure,  # Use extracted or default UOM
                        'supplier': default_supplier or '',
                        'is_active': True
                    }
                    
                    # Handle image extraction for new item
                    if 'photo' in col_mapping:
                        photo_cell = f"{openpyxl.utils.get_column_letter(col_mapping['photo'] + 1)}{row_num}"
                        if image_loader.image_in(photo_cell):
                            try:
                                image = image_loader.get(photo_cell)
                                if image:
                                    img_io = io.BytesIO()
                                    if image.mode == 'RGBA':
                                        image.save(img_io, format='PNG', optimize=True)
                                        filename = f"{item_code.lower()}.png"
                                    else:
                                        if image.mode not in ['RGB', 'L']:
                                            image = image.convert('RGB')
                                        image.save(img_io, format='JPEG', quality=95, optimize=True)
                                        filename = f"{item_code.lower()}.jpg"
                                    
                                    img_io.seek(0)
                                    image_file = ContentFile(img_io.getvalue(), name=filename)
                                    bom_item_data['item_photo'] = image_file
                            except Exception as e:
                                result['errors'].append(f"Row {row_num}: Could not extract image - {str(e)}")
                    
                    bom_item = BOMItem.objects.create(**bom_item_data)
                    result['created'] += 1
                
                # Create or update BOM Template Item
                try:
                    serial_number = int(s_no) if s_no.isdigit() else row_num - header_row
                    
                    # Parse quantity - handle different formats
                    try:
                        base_quantity = float(qty) if qty.replace('.', '').replace(',', '').isdigit() else 1.0
                    except (ValueError, AttributeError):
                        base_quantity = 1.0
                    
                    # Check if template item already exists with this serial number
                    try:
                        template_item = BOMTemplateItem.objects.get(
                            bom_template=bom_template,
                            serial_number=serial_number
                        )
                        
                        if overwrite_existing:
                            template_item.item = bom_item
                            template_item.base_quantity = base_quantity
                            template_item.save()
                        
                    except BOMTemplateItem.DoesNotExist:
                        # Create new template item
                        BOMTemplateItem.objects.create(
                            bom_template=bom_template,
                            item=bom_item,
                            serial_number=serial_number,
                            base_quantity=base_quantity,
                            is_active=True
                        )
                        
                except Exception as e:
                    result['errors'].append(f"Row {row_num}: Error creating template item - {str(e)}")
                    
            except Exception as e:
                result['errors'].append(f"Row {row_num}: {str(e)}")
                continue
        
        result['success'] = True
        
    except Exception as e:
        result['error'] = str(e)
    
    return result

def process_zip_file_for_media(product, zip_file, process, bom, default_duration, 
                              display_screen_1, display_screen_2, display_screen_3):
    """Process uploaded ZIP file and create ProductMedia objects"""
    result = {
        'success': False,
        'total_files': 0,
        'pdfs': 0,
        'videos': 0,
        'skipped': 0,
        'errors': []
    }
    
    # Supported file extensions
    PDF_EXTENSIONS = {'.pdf'}
    VIDEO_EXTENSIONS = {'.mp4', '.mov', '.avi', '.mkv', '.wmv', '.flv', '.webm'}
    
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            # Extract zip file
            with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # Process all files in the extracted directory
            for root, dirs, files in os.walk(temp_dir):
                for filename in files:
                    file_path = os.path.join(root, filename)
                    file_ext = Path(filename).suffix.lower()
                    
                    # Skip hidden files and system files
                    if filename.startswith('.') or filename.startswith('__MACOSX'):
                        continue
                    
                    try:
                        # Determine media type based on extension
                        media_type = None
                        duration = None
                        
                        if file_ext in PDF_EXTENSIONS:
                            media_type = 'PROCESS_DOCUMENT'
                            result['pdfs'] += 1
                        elif file_ext in VIDEO_EXTENSIONS:
                            media_type = 'VIDEO'
                            duration = default_duration
                            result['videos'] += 1
                        else:
                            result['skipped'] += 1
                            continue
                        
                        # Read file content
                        with open(file_path, 'rb') as f:
                            file_content = f.read()
                        
                        # Create ContentFile
                        django_file = ContentFile(file_content, name=filename)
                        
                        # Create ProductMedia object
                        media_data = {
                            'product': product,
                            'media_type': media_type,
                            'file': django_file,
                            'display_screen_1': display_screen_1,
                            'display_screen_2': display_screen_2,
                            'display_screen_3': display_screen_3,
                        }
                        
                        # Add optional fields
                        if process:
                            media_data['process'] = process
                        if bom:
                            media_data['bom'] = bom
                        if duration:
                            media_data['duration'] = duration
                        
                        # Create the ProductMedia object
                        ProductMedia.objects.create(**media_data)
                        result['total_files'] += 1
                        
                    except Exception as e:
                        result['errors'].append(f"{filename}: {str(e)}")
                        continue
        
        result['success'] = True
        
    except zipfile.BadZipFile:
        result['error'] = "Invalid ZIP file"
    except Exception as e:
        result['error'] = str(e)
    
    return result










# Add this ONE simple endpoint to your views.py

import json
import time
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils.decorators import method_decorator
from django.views import View
from django.core.cache import cache
import logging

logger = logging.getLogger(__name__)

@method_decorator(csrf_exempt, name='dispatch')
class TriggerStationsReloadView(View):
    """
    Endpoint to trigger page reload on other stations
    """
    
    def post(self, request, station_id):
        try:
            data = json.loads(request.body)
            source_station = data.get('source_station', 1)
            target_stations = data.get('target_stations', [2, 3])
            
            logger.info(f"Station {source_station} triggering reload on stations: {target_stations}")
            
            # Create reload signals for target stations
            reload_signals = {}
            timestamp = int(time.time() * 1000)  # Current timestamp in milliseconds
            
            for target_station in target_stations:
                signal_key = f"station_{target_station}_reload_signal"
                signal_data = {
                    'triggered_by': source_station,
                    'timestamp': timestamp,
                    'action': 'reload_page',
                    'reason': 'triggered_by_station_1_left_arrow'
                }
                
                # Store in cache with 30 second expiration
                cache.set(signal_key, signal_data, 30)
                reload_signals[target_station] = signal_data
                
                logger.info(f"Created reload signal for station {target_station}: {signal_data}")
            
            return JsonResponse({
                'success': True,
                'message': f'Reload triggered for stations {target_stations}',
                'source_station': source_station,
                'target_stations': target_stations,
                'timestamp': timestamp,
                'signals_created': reload_signals
            })
            
        except json.JSONDecodeError:
            logger.error("Invalid JSON in request body")
            return JsonResponse({
                'success': False,
                'error': 'Invalid JSON in request body'
            }, status=400)
            
        except Exception as e:
            logger.error(f"Error triggering stations reload: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Server error: {str(e)}'
            }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class CheckReloadSignalView(View):
    """
    Endpoint for stations to check if they should reload
    """
    
    def get(self, request, station_id):
        try:
            # Get the station display number from your station model
            # Adjust this based on your actual model structure
            display_number = request.GET.get('display_number')
            if not display_number:
                return JsonResponse({
                    'success': False,
                    'error': 'Display number required'
                }, status=400)
            
            signal_key = f"station_{display_number}_reload_signal"
            signal_data = cache.get(signal_key)
            
            if signal_data:
                # Signal found - clear it so it only triggers once
                cache.delete(signal_key)
                
                logger.info(f"Station {display_number} found reload signal: {signal_data}")
                
                return JsonResponse({
                    'success': True,
                    'should_reload': True,
                    'signal_data': signal_data
                })
            else:
                return JsonResponse({
                    'success': True,
                    'should_reload': False,
                    'message': 'No reload signal found'
                })
                
        except Exception as e:
            logger.error(f"Error checking reload signal for station {station_id}: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Server error: {str(e)}'
            }, status=500)





# In your views.py

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.core.cache import cache
import json
import time

@require_http_methods(["POST"])
def sync_bom_pagination(request, station_id):
    """
    Broadcast BOM pagination changes from Display 1 to other displays
    """
    try:
        data = json.loads(request.body)
        source_display = data.get('source_display')
        target_page = data.get('target_page')
        total_pages = data.get('total_pages')
        mode = data.get('mode', 'split')
        items_per_screen = data.get('items_per_screen', 8)
        
        # Validate data
        if not all([source_display, target_page, total_pages]):
            return JsonResponse({
                'success': False,
                'error': 'Missing required parameters'
            }, status=400)
        
        # Store sync signal in cache for Display 2
        cache_key_display2 = f'bom_sync_station_{station_id}_display_2'
        cache.set(cache_key_display2, {
            'triggered_by': source_display,
            'target_page': target_page,
            'total_pages': total_pages,
            'mode': mode,
            'items_per_screen': items_per_screen,
            'timestamp': time.time()
        }, timeout=10)  # Signal expires after 10 seconds
        
        # Store sync signal in cache for Display 3
        cache_key_display3 = f'bom_sync_station_{station_id}_display_3'
        cache.set(cache_key_display3, {
            'triggered_by': source_display,
            'target_page': target_page,
            'total_pages': total_pages,
            'mode': mode,
            'items_per_screen': items_per_screen,
            'timestamp': time.time()
        }, timeout=10)
        
        return JsonResponse({
            'success': True,
            'message': f'BOM sync signal broadcasted to displays 2 & 3',
            'sync_data': {
                'source_display': source_display,
                'target_page': target_page,
                'total_pages': total_pages
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@require_http_methods(["GET"])
def check_bom_sync(request, station_id):
    """
    Check if there's a BOM sync signal for this display
    """
    try:
        display_number = request.GET.get('display_number')
        
        if not display_number:
            return JsonResponse({
                'success': False,
                'error': 'display_number parameter required'
            }, status=400)
        
        # Check cache for sync signal
        cache_key = f'bom_sync_station_{station_id}_display_{display_number}'
        sync_data = cache.get(cache_key)
        
        if sync_data:
            # Clear the signal after reading (one-time use)
            cache.delete(cache_key)
            
            return JsonResponse({
                'success': True,
                'should_sync': True,
                'sync_data': sync_data
            })
        else:
            return JsonResponse({
                'success': True,
                'should_sync': False
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)