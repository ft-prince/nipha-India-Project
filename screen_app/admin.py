from django.contrib import admin
from django.utils.html import format_html
from django import forms
from django.db import models
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from PIL import Image
import io
import os
import zipfile
import tempfile
from pathlib import Path
from .models import (
    Product, AssemblyStage, AssemblyProcess, BillOfMaterial, 
    ProductMedia, Station, AssemblySession,
    # New models
    BOMItem, BOMTemplate, BOMTemplateItem
)




# Zip Upload Form for ProductMedia
class ZipUploadForm(forms.Form):
    zip_file = forms.FileField(
        label='Zip File',
        help_text='Upload zip file containing PDFs and videos. PDFs will be marked as Process Documents, videos as Videos.',
        widget=forms.FileInput(attrs={'accept': '.zip'})
    )
    process = forms.ModelChoiceField(
        queryset=AssemblyProcess.objects.all(),
        required=False,
        help_text='Optional: Assign all extracted media to this process'
    )
    bom = forms.ModelChoiceField(
        queryset=BillOfMaterial.objects.all(),
        required=False,
        help_text='Optional: Assign all extracted media to this BOM'
    )
    default_duration = forms.IntegerField(
        initial=30,
        min_value=1,
        help_text='Default duration in seconds for video files'
    )
    # Display assignments
    display_screen_1 = forms.BooleanField(required=False, label='Display Screen 1')
    display_screen_2 = forms.BooleanField(required=False, label='Display Screen 2') 
    display_screen_3 = forms.BooleanField(required=False, label='Display Screen 3')
    
    def __init__(self, *args, **kwargs):
        product = kwargs.pop('product', None)
        super().__init__(*args, **kwargs)
        
        if product:
            # Filter processes and BOMs by product
            self.fields['process'].queryset = AssemblyProcess.objects.filter(
                stage__in=AssemblyStage.objects.all()
            ).order_by('stage__order', 'order')
            
            self.fields['bom'].queryset = BillOfMaterial.objects.filter(
                product=product
            )
  


# Excel Upload Form for BOM Items
class ExcelUploadForm(forms.Form):
    excel_file = forms.FileField(
        label='Excel File',
        help_text='Upload Excel file with columns: S. NO, ITEM DESCRIPTION, PART NO., QTY, ITEM PHOTO',
        widget=forms.FileInput(attrs={'accept': '.xlsx,.xls'})
    )
    unit_of_measure = forms.CharField(
        max_length=20,
        initial='NO.',
        help_text='Default unit of measure for all items (e.g., NO., KGS, GM, LTR)'
    )
    supplier = forms.CharField(
        max_length=100,
        required=False,
        help_text='Optional: Default supplier for all items'
    )
    overwrite_existing = forms.BooleanField(
        required=False,
        initial=False,
        help_text='Overwrite existing items with same item code'
    )

# BOM Item Management with Excel Upload
@admin.register(BOMItem)
class BOMItemAdmin(admin.ModelAdmin):
    list_display = ('item_code', 'item_description', 'part_number', 'unit_of_measure', 'cost_per_unit', 'item_photo_preview', 'is_active')
    list_filter = ('unit_of_measure', 'is_active', 'supplier')
    search_fields = ('item_code', 'item_description', 'part_number')
    list_editable = ('is_active', 'cost_per_unit')
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('item_code', 'item_description', 'part_number', 'unit_of_measure')
        }),
        ('Visual', {
            'fields': ('item_photo',),
            'description': 'Upload clear photos of items for BOM display'
        }),
        ('Additional Details', {
            'fields': ('supplier', 'cost_per_unit', 'weight_per_unit'),
            'classes': ('collapse',)
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )
    
    def item_photo_preview(self, obj):
        if obj.item_photo:
            return format_html(
                '<img src="{}" width="50" height="50" style="object-fit: cover; border-radius: 4px;"/>',
                obj.item_photo.url
            )
        return "No photo"
    item_photo_preview.short_description = 'Photo'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.admin_site.admin_view(self.upload_excel_view), name='bomitem_upload_excel'),
        ]
        return custom_urls + urls
    
    def upload_excel_view(self, request):
        """Handle Excel file upload for bulk BOM item creation"""
        if request.method == 'POST':
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    excel_file = form.cleaned_data['excel_file']
                    unit_of_measure = form.cleaned_data['unit_of_measure']
                    supplier = form.cleaned_data['supplier']
                    overwrite_existing = form.cleaned_data['overwrite_existing']
                    
                    # Process the Excel file
                    result = self.process_excel_file(
                        excel_file, 
                        unit_of_measure, 
                        supplier, 
                        overwrite_existing
                    )
                    
                    if result['success']:
                        messages.success(
                            request, 
                            f"Successfully processed {result['created']} new items and {result['updated']} updated items. "
                            f"Skipped {result['skipped']} items."
                        )
                        if result['errors']:
                            messages.warning(request, f"Errors encountered: {'; '.join(result['errors'])}")
                    else:
                        messages.error(request, f"Error processing file: {result['error']}")
                    
                    return redirect('admin:screen_app_bomitem_changelist')
                    
                except Exception as e:
                    messages.error(request, f"Error processing Excel file: {str(e)}")
        else:
            form = ExcelUploadForm()
        
        context = {
            'form': form,
            'title': 'Upload BOM Items from Excel',
            'opts': self.model._meta,
            'has_change_permission': True,
        }
        
        return render(request, 'admin/bomitem_excel_upload.html', context)
    
    def process_excel_file(self, excel_file, default_unit, default_supplier, overwrite_existing):
        """Process the uploaded Excel file and create BOM items"""
        result = {
            'success': False,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': []
        }
        
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = workbook.active
            
            # Load images from the sheet
            image_loader = SheetImageLoader(sheet)
            
            # Expected columns: S. NO, ITEM DESCRIPTION, PART NO., QTY, ITEM PHOTO
            # Find header row (usually row 1)
            header_row = None
            for row_num in range(1, 6):  # Check first 5 rows for headers
                row_values = [str(cell.value).strip().upper() if cell.value else '' for cell in sheet[row_num]]
                if any('ITEM DESCRIPTION' in val or 'DESCRIPTION' in val for val in row_values):
                    header_row = row_num
                    break
            
            if header_row is None:
                result['error'] = "Could not find header row with 'ITEM DESCRIPTION' column"
                return result
            
            # Map column indices
            headers = [str(cell.value).strip().upper() if cell.value else '' for cell in sheet[header_row]]
            
            col_mapping = {}
            for idx, header in enumerate(headers):
                if 'S.' in header and 'NO' in header:
                    col_mapping['s_no'] = idx
                elif 'ITEM DESCRIPTION' in header or 'DESCRIPTION' in header:
                    col_mapping['description'] = idx
                elif 'PART NO' in header or 'PART_NO' in header:
                    col_mapping['part_no'] = idx
                elif 'QTY' in header or 'QUANTITY' in header:
                    col_mapping['qty'] = idx
                elif 'PHOTO' in header or 'IMAGE' in header:
                    col_mapping['photo'] = idx
            
            # Validate required columns
            required_cols = ['description', 'part_no']
            missing_cols = [col for col in required_cols if col not in col_mapping]
            if missing_cols:
                result['error'] = f"Missing required columns: {missing_cols}"
                return result
            
            # Process data rows
            for row_num in range(header_row + 1, sheet.max_row + 1):
                row = sheet[row_num]
                
                # Skip empty rows
                if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                    continue
                
                try:
                    # Extract data
                    s_no = str(row[col_mapping.get('s_no', 0)].value).strip() if col_mapping.get('s_no') is not None else str(row_num - header_row)
                    description = str(row[col_mapping['description']].value).strip() if row[col_mapping['description']].value else ''
                    part_no = str(row[col_mapping['part_no']].value).strip() if row[col_mapping['part_no']].value else ''
                    qty = str(row[col_mapping.get('qty', 0)].value).strip() if col_mapping.get('qty') is not None and row[col_mapping.get('qty', 0)].value else '1'
                    
                    # Skip if essential data is missing
                    if not description or not part_no:
                        result['skipped'] += 1
                        continue
                    
                    # Generate item_code from description (clean and uppercase)
                    item_code = description.upper().replace(' ', '_').replace('-', '_')
                    # Remove special characters and keep only alphanumeric and underscore
                    item_code = ''.join(c for c in item_code if c.isalnum() or c == '_')
                    
                    # Check if item exists
                    existing_item = None
                    try:
                        existing_item = BOMItem.objects.get(item_code=item_code)
                        if not overwrite_existing:
                            result['skipped'] += 1
                            continue
                    except BOMItem.DoesNotExist:
                        pass
                    
                    # Handle image extraction - use original image without padding
                    image_file = None
                    if 'photo' in col_mapping:
                        photo_cell = f"{openpyxl.utils.get_column_letter(col_mapping['photo'] + 1)}{row_num}"
                        if image_loader.image_in(photo_cell):
                            try:
                                # Get the image using the image loader
                                image = image_loader.get(photo_cell)
                                if image:
                                    # Save the image without any processing - just as extracted
                                    img_io = io.BytesIO()
                                    
                                    # Use the original format and quality
                                    if image.mode == 'RGBA':
                                        image.save(img_io, format='PNG', optimize=True)
                                        filename = f"{item_code.lower()}.png"
                                    else:
                                        if image.mode not in ['RGB', 'L']:
                                            image = image.convert('RGB')
                                        image.save(img_io, format='JPEG', quality=95, optimize=True)
                                        filename = f"{item_code.lower()}.jpg"
                                    
                                    img_io.seek(0)
                                    image_file = ContentFile(img_io.getvalue(), name=filename)
                                    
                            except Exception as e:
                                result['errors'].append(f"Row {row_num}: Could not extract image - {str(e)}")
                    
                    # Create or update BOM item
                    item_data = {
                        'item_description': description,
                        'part_number': part_no,
                        'unit_of_measure': default_unit,
                        'supplier': default_supplier,
                        'is_active': True
                    }
                    
                    if existing_item:
                        # Update existing item
                        for key, value in item_data.items():
                            setattr(existing_item, key, value)
                        
                        if image_file:
                            # Delete old image if exists
                            if existing_item.item_photo:
                                existing_item.item_photo.delete(save=False)
                            existing_item.item_photo = image_file
                        
                        existing_item.save()
                        result['updated'] += 1
                    else:
                        # Create new item
                        item_data['item_code'] = item_code
                        if image_file:
                            item_data['item_photo'] = image_file
                        
                        BOMItem.objects.create(**item_data)
                        result['created'] += 1
                
                except Exception as e:
                    result['errors'].append(f"Row {row_num}: {str(e)}")
                    continue
            
            result['success'] = True
            
        except Exception as e:
            result['error'] = str(e)
        
        return result
    
    def changelist_view(self, request, extra_context=None):
        """Add upload button to changelist"""
        extra_context = extra_context or {}
        extra_context['show_excel_upload'] = True
        return super().changelist_view(request, extra_context=extra_context)
 
# BOM Template Item Inline
class BOMTemplateItemInline(admin.TabularInline):
    model = BOMTemplateItem
    extra = 1
    fields = ['serial_number', 'item', 'base_quantity', 'notes', 'is_active']
    ordering = ['serial_number']
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "item":
            kwargs["queryset"] = BOMItem.objects.filter(is_active=True).order_by('item_description')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

# BOM Template Management
@admin.register(BOMTemplate)
class BOMTemplateAdmin(admin.ModelAdmin):
    list_display = ('template_name', 'product', 'bom_type', 'stage', 'item_count', 'display_assignment', 'duration_info', 'is_active')
    list_filter = ('product', 'bom_type', 'stage', 'is_active', 'is_duration_active')
    search_fields = ('template_name', 'product__code', 'product__name')
    list_editable = ('is_active',)
    inlines = [BOMTemplateItemInline]
    
    fieldsets = (
        ('Template Information', {
            'fields': ('template_name', 'product', 'bom_type', 'stage', 'description')
        }),
        ('Display Assignment', {
            'fields': ('display_screen_1', 'display_screen_2', 'display_screen_3'),
            'description': 'Select which displays should show this BOM'
        }),
        ('Duration Settings', {
            'fields': ('duration', 'is_duration_active'),
            'description': 'Control how long this BOM is displayed'
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )
    
    def item_count(self, obj):
        return obj.bom_items.filter(is_active=True).count()
    item_count.short_description = 'Items'
    
    def display_assignment(self, obj):
        displays = []
        if obj.display_screen_1: displays.append('D1')
        if obj.display_screen_2: displays.append('D2')
        if obj.display_screen_3: displays.append('D3')
        
        if displays:
            display_badges = []
            colors = ['red', 'blue', 'green']
            for i, display in enumerate(displays):
                color = colors[i] if i < len(colors) else 'gray'
                display_badges.append(f'<span style="background-color: {color}; color: white; padding: 2px 6px; border-radius: 3px; margin: 1px;">{display}</span>')
            return format_html(' '.join(display_badges))
        return format_html('<span style="color: gray;">None</span>')
    display_assignment.short_description = 'Displays'
    
    def duration_info(self, obj):
        if obj.is_duration_active:
            return format_html(
                '<span style="color: green; font-weight: bold;">{}s</span>',
                obj.duration
            )
        else:
            return format_html(
                '<span style="color: gray;">{}s (inactive)</span>',
                obj.duration
            )
    duration_info.short_description = 'Duration'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:object_id>/preview-bom/', self.admin_site.admin_view(self.preview_bom_view), name='bomtemplate_preview'),
        ]
        return custom_urls + urls
    
    def preview_bom_view(self, request, object_id):
        """Preview BOM for different quantities"""
        template = self.get_object(request, object_id)
        quantity = int(request.GET.get('quantity', 1))
        
        bom_data = template.generate_bom_for_quantity(quantity)
        
        context = {
            'template': template,
            'quantity': quantity,
            'bom_data': bom_data,
            'title': f'BOM Preview - {template.template_name}',
        }
        
        return render(request, 'admin/bom_preview.html', context)
# Enhanced BOM Template Item Admin
@admin.register(BOMTemplateItem)
class BOMTemplateItemAdmin(admin.ModelAdmin):
    list_display = ('bom_template', 'serial_number', 'item', 'base_quantity', 'formatted_quantity_50', 'is_active')
    list_filter = ('bom_template__product', 'bom_template__bom_type', 'is_active')
    search_fields = ('item__item_description', 'bom_template__template_name')
    ordering = ['bom_template', 'serial_number']
    
    fieldsets = (
        ('BOM Line Information', {
            'fields': ('bom_template', 'serial_number', 'item', 'base_quantity')
        }),
        ('Additional Details', {
            'fields': ('notes', 'is_active')
        }),
    )
    
    def formatted_quantity_50(self, obj):
        """Show what quantity would be for 50 units"""
        calc_qty = obj.calculate_quantity_for_production(50)
        if obj.item.unit_of_measure in ['KGS', 'GM', 'LTR']:
            return f"{calc_qty:.3f} {obj.item.unit_of_measure}"
        else:
            return f"{int(calc_qty)} {obj.item.unit_of_measure}"
    formatted_quantity_50.short_description = 'Qty for 50 Units'
  
# Enhanced existing admins - Updated sections for AssemblyStage and AssemblyProcess

# Enhanced existing admins - Updated sections for AssemblyStage and AssemblyProcess

@admin.register(AssemblyStage)
class AssemblyStageAdmin(admin.ModelAdmin):
    list_display = ('product', 'order', 'name', 'display_name', 'process_count', 'bom_template_count')
    list_editable = ('order', 'display_name')
    list_filter = ('product', 'name')
    search_fields = ('name', 'display_name', 'product__code', 'product__name')
    ordering = ('product', 'order')
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('product', 'name', 'display_name')
        }),
        ('Ordering', {
            'fields': ('order',),
            'description': 'Order within the selected product (stages will be sorted by this number)'
        }),
    )
    
    def get_queryset(self, request):
        """Optimize queryset with select_related"""
        return super().get_queryset(request).select_related('product').prefetch_related('processes', 'bom_templates')
    
    def process_count(self, obj):
        count = obj.processes.count()
        if count > 0:
            return format_html(
                '<span style="background-color: #28a745; color: white; padding: 2px 8px; border-radius: 12px; font-size: 11px;">{}</span>',
                count
            )
        return format_html('<span style="color: #6c757d;">0</span>')
    process_count.short_description = 'Processes'
    
    def bom_template_count(self, obj):
        count = obj.bom_templates.count()
        if count > 0:
            return format_html(
                '<span style="background-color: #17a2b8; color: white; padding: 2px 8px; border-radius: 12px; font-size: 11px;">{}</span>',
                count
            )
        return format_html('<span style="color: #6c757d;">0</span>')
    bom_template_count.short_description = 'BOM Templates'
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "product":
            kwargs["queryset"] = Product.objects.all().order_by('code', 'name')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

class AssemblyProcessInline(admin.TabularInline):
    model = AssemblyProcess
    extra = 1
    fields = ['order', 'name', 'display_name', 'location', 'is_looped', 'loop_group']
    ordering = ['order']
    
    def get_queryset(self, request):
        return super().get_queryset(request).order_by('order')

@admin.register(AssemblyProcess)
class AssemblyProcessAdmin(admin.ModelAdmin):
    list_display = ('get_product', 'stage', 'order', 'name', 'display_name', 'location', 'loop_info', 'media_count')
    list_filter = (
        'stage__product',
        'stage',
        'location', 
        'is_looped'
    )
    list_editable = ('order', 'display_name', 'location')
    search_fields = ('name', 'display_name', 'stage__name', 'stage__product__code', 'stage__product__name')
    ordering = ('stage__product', 'stage__order', 'order')
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('stage', 'name', 'display_name')
        }),
        ('Process Details', {
            'fields': ('location', 'order'),
            'description': 'Location and order within the stage'
        }),
        ('Loop Configuration', {
            'fields': ('is_looped', 'loop_group'),
            'description': 'Configure if this process should loop and its loop group'
        }),
    )
    
    def get_queryset(self, request):
        """Optimize queryset with select_related"""
        return super().get_queryset(request).select_related('stage', 'stage__product').prefetch_related('media')
    
    def get_product(self, obj):
        if obj.stage and obj.stage.product:
            return format_html(
                '<strong>{}</strong><br><small style="color: #6c757d;">{}</small>',
                obj.stage.product.code,
                obj.stage.product.name[:30] + ('...' if len(obj.stage.product.name) > 30 else '')
            )
        return format_html('<span style="color: #dc3545;">No Product</span>')
    get_product.short_description = 'Product'
    get_product.admin_order_field = 'stage__product__code'
    
    def loop_info(self, obj):
        if obj.is_looped:
            loop_text = f"Group {obj.loop_group}" if obj.loop_group else "No Group"
            return format_html(
                '<span style="background-color: #ffc107; color: #000; padding: 2px 6px; border-radius: 3px; font-size: 11px;">🔄 {}</span>',
                loop_text
            )
        return format_html('<span style="color: #6c757d;">—</span>')
    loop_info.short_description = 'Loop Status'
    
    def media_count(self, obj):
        count = obj.media.count()
        if count > 0:
            return format_html(
                '<span style="background-color: #6f42c1; color: white; padding: 2px 8px; border-radius: 12px; font-size: 11px;">{}</span>',
                count
            )
        return format_html('<span style="color: #6c757d;">0</span>')
    media_count.short_description = 'Media'
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "stage":
            # Show stages ordered by product and then by order
            kwargs["queryset"] = AssemblyStage.objects.select_related('product').order_by('product__code', 'order')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    def save_model(self, request, obj, form, change):
        """Custom save to handle any business logic"""
        super().save_model(request, obj, form, change)
    
    def get_list_display_links(self, request, list_display):
        """Make name clickable for editing"""
        return ('name',)

# Custom admin actions for stages and processes
@admin.action(description='Duplicate selected stages with all processes')
def duplicate_stages_with_processes(modeladmin, request, queryset):
    """Duplicate stages and their associated processes"""
    duplicated_count = 0
    for stage in queryset:
        # Create new stage
        new_stage = AssemblyStage.objects.create(
            product=stage.product,
            name=f"{stage.name} (Copy)",
            display_name=f"{stage.display_name} (Copy)" if stage.display_name else None,
            order=stage.order + 1000  # Add high number to avoid conflicts
        )
        
        # Copy all processes
        for process in stage.processes.all():
            AssemblyProcess.objects.create(
                stage=new_stage,
                name=process.name,
                display_name=process.display_name,
                location=process.location,
                order=process.order,
                is_looped=process.is_looped,
                loop_group=process.loop_group
            )
        
        duplicated_count += 1
    
    messages.success(request, f'Successfully duplicated {duplicated_count} stage(s) with their processes')

@admin.action(description='Reset order numbers (1, 2, 3, ...)')
def reset_stage_order(modeladmin, request, queryset):
    """Reset order numbers for selected stages within each product"""
    products_updated = set()
    
    for product_id in queryset.values_list('product_id', flat=True).distinct():
        stages = queryset.filter(product_id=product_id).order_by('order')
        for index, stage in enumerate(stages, start=1):
            stage.order = index
            stage.save()
        products_updated.add(product_id)
    
    product_names = Product.objects.filter(id__in=products_updated).values_list('code', flat=True)
    messages.success(
        request, 
        f'Reset order numbers for stages in products: {", ".join(product_names)}'
    )

@admin.action(description='Reset process order numbers within stages')
def reset_process_order(modeladmin, request, queryset):
    """Reset order numbers for processes within their stages"""
    stages_updated = set()
    
    for stage_id in queryset.values_list('stage_id', flat=True).distinct():
        processes = queryset.filter(stage_id=stage_id).order_by('order')
        for index, process in enumerate(processes, start=1):
            process.order = index
            process.save()
        stages_updated.add(stage_id)
    
    stage_names = AssemblyStage.objects.filter(id__in=stages_updated).values_list('name', flat=True)
    messages.success(
        request, 
        f'Reset process order numbers in stages: {", ".join(stage_names)}'
    )

# Add the actions to the admin classes
AssemblyStageAdmin.actions = [duplicate_stages_with_processes, reset_stage_order]
AssemblyProcessAdmin.actions = [reset_process_order]

# Updated Product Admin to include Assembly Stages
class AssemblyStageInline(admin.TabularInline):
    model = AssemblyStage
    extra = 0
    fields = ['order', 'name', 'display_name', 'process_count_inline', 'bom_count_inline']
    readonly_fields = ['process_count_inline', 'bom_count_inline']
    ordering = ['order']
    
    def process_count_inline(self, obj):
        if obj.pk:
            count = obj.processes.count()
            return format_html('<span style="color: #28a745; font-weight: bold;">{}</span>', count)
        return "—"
    process_count_inline.short_description = 'Processes'
    
    def bom_count_inline(self, obj):
        if obj.pk:
            count = obj.bom_templates.count()
            return format_html('<span style="color: #17a2b8; font-weight: bold;">{}</span>', count)
        return "—"
    bom_count_inline.short_description = 'BOMs'


@admin.register(BillOfMaterial)
class BillOfMaterialAdmin(admin.ModelAdmin):
    list_display = ('product', 'bom_type', 'stage', 'source_type', 'file_or_template', 'preview_link')
    list_filter = ('bom_type', 'stage', 'product', 'source_type')
    search_fields = ('product__code', 'product__name')
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('product', 'bom_type', 'stage', 'source_type')
        }),
        ('PDF Source (Legacy)', {
            'fields': ('file',),
            'classes': ('collapse',),
            'description': 'Use for PDF-based BOMs'
        }),
        ('Database Source (Recommended)', {
            'fields': ('bom_template',),
            'description': 'Link to database BOM template for dynamic quantity calculation'
        }),
    )
    
    def file_or_template(self, obj):
        if obj.source_type == 'PDF' and obj.file:
            return format_html('<a href="{}" target="_blank">📄 PDF File</a>', obj.file.url)
        elif obj.source_type == 'DATABASE' and obj.bom_template:
            return format_html('📊 {}', obj.bom_template.template_name)
        return "Not configured"
    file_or_template.short_description = 'Source'
    
    def preview_link(self, obj):
        if obj.source_type == 'DATABASE' and obj.bom_template:
            return format_html(
                '<a href="{}?quantity=1" target="_blank">Preview 1x</a> | '
                '<a href="{}?quantity=50" target="_blank">Preview 50x</a>',
                f'/admin/screen_app/bomtemplate/{obj.bom_template.id}/preview-bom/',
                f'/admin/screen_app/bomtemplate/{obj.bom_template.id}/preview-bom/'
            )
        return "N/A"
    preview_link.short_description = 'Preview'

class ProductMediaInline(admin.TabularInline):
    model = ProductMedia
    extra = 1
    fields = ['media_type', 'file', 'file_preview', 'duration', 'process', 'bom', 
             'display_screen_1', 'display_screen_2', 'display_screen_3']
    readonly_fields = ['file_preview']

    def file_preview(self, obj):
        if obj.file:
            file_url = obj.file.url
            file_name = obj.file.name.lower()
            if file_name.endswith('.pdf'):
                return format_html('<a href="{}" target="_blank">📄 View PDF</a>', file_url)
            elif file_name.endswith(('.mp4', '.mov', '.avi', '.mkv')):
                return format_html('<video width="100" height="60" controls><source src="{}" type="video/mp4"></video>', file_url)
            elif file_name.endswith(('.xlsx', '.docx')):
                return format_html('<a href="{}" target="_blank">📊 Download</a>', file_url)
            else:
                return format_html('<a href="{}">📁 View File</a>', file_url)
        return "No file"
    file_preview.short_description = 'Preview'

 

@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('code', 'name', 'media_count', 'bom_count', 'bom_template_count')
    search_fields = ('code', 'name')
    inlines = [ProductMediaInline]

    def media_count(self, obj):
        return obj.media.count()
    media_count.short_description = 'Media Files'
    
    def bom_count(self, obj):
        return obj.boms.count()
    bom_count.short_description = 'BOMs'
    
    def bom_template_count(self, obj):
        return obj.bom_templates.count()
    bom_template_count.short_description = 'BOM Templates'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:object_id>/upload-zip/', self.admin_site.admin_view(self.upload_zip_view), name='product_upload_zip'),
        ]
        return custom_urls + urls
    
    def upload_zip_view(self, request, object_id):
        """Handle zip file upload for bulk ProductMedia creation"""
        product = self.get_object(request, object_id)
        
        if request.method == 'POST':
            form = ZipUploadForm(request.POST, request.FILES, product=product)
            if form.is_valid():
                try:
                    result = self.process_zip_file(product, form.cleaned_data)
                    
                    if result['success']:
                        messages.success(
                            request, 
                            f"Successfully processed {result['total_files']} files: "
                            f"{result['pdfs']} PDFs, {result['videos']} videos. "
                            f"Skipped {result['skipped']} files."
                        )
                        if result['errors']:
                            messages.warning(request, f"Some errors: {'; '.join(result['errors'])}")
                    else:
                        messages.error(request, f"Error processing zip: {result['error']}")
                    
                    return redirect('admin:screen_app_product_change', object_id=object_id)
                    
                except Exception as e:
                    messages.error(request, f"Error processing zip file: {str(e)}")
        else:
            form = ZipUploadForm(product=product)
        
        context = {
            'form': form,
            'product': product,
            'title': f'Upload Media Zip for {product.name}',
            'opts': self.model._meta,
            'has_change_permission': True,
        }
        
        return render(request, 'admin/product_zip_upload.html', context)
    
    def process_zip_file(self, product, form_data):
        """Process uploaded zip file and create ProductMedia objects"""
        result = {
            'success': False,
            'total_files': 0,
            'pdfs': 0,
            'videos': 0,
            'skipped': 0,
            'errors': []
        }
        
        zip_file = form_data['zip_file']
        process = form_data.get('process')
        bom = form_data.get('bom')
        default_duration = form_data['default_duration']
        display_screen_1 = form_data['display_screen_1']
        display_screen_2 = form_data['display_screen_2']
        display_screen_3 = form_data['display_screen_3']
        
        # Supported file extensions
        PDF_EXTENSIONS = {'.pdf'}
        VIDEO_EXTENSIONS = {'.mp4', '.mov', '.avi', '.mkv', '.wmv', '.flv', '.webm'}
        
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                # Extract zip file
                with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir)
                
                # Process all files in the extracted directory
                for root, dirs, files in os.walk(temp_dir):
                    for filename in files:
                        file_path = os.path.join(root, filename)
                        file_ext = Path(filename).suffix.lower()
                        
                        # Skip hidden files and system files
                        if filename.startswith('.') or filename.startswith('__MACOSX'):
                            continue
                        
                        try:
                            # Determine media type based on extension
                            media_type = None
                            duration = None
                            
                            if file_ext in PDF_EXTENSIONS:
                                media_type = 'PROCESS_DOCUMENT'
                                result['pdfs'] += 1
                            elif file_ext in VIDEO_EXTENSIONS:
                                media_type = 'VIDEO'
                                duration = default_duration
                                result['videos'] += 1
                            else:
                                result['skipped'] += 1
                                continue
                            
                            # Read file content
                            with open(file_path, 'rb') as f:
                                file_content = f.read()
                            
                            # Create ContentFile
                            django_file = ContentFile(file_content, name=filename)
                            
                            # Create ProductMedia object
                            media_data = {
                                'product': product,
                                'media_type': media_type,
                                'file': django_file,
                                'display_screen_1': display_screen_1,
                                'display_screen_2': display_screen_2,
                                'display_screen_3': display_screen_3,
                            }
                            
                            # Add optional fields
                            if process:
                                media_data['process'] = process
                            if bom:
                                media_data['bom'] = bom
                            if duration:
                                media_data['duration'] = duration
                            
                            # Create the ProductMedia object
                            ProductMedia.objects.create(**media_data)
                            result['total_files'] += 1
                            
                        except Exception as e:
                            result['errors'].append(f"{filename}: {str(e)}")
                            continue
            
            result['success'] = True
            
        except zipfile.BadZipFile:
            result['error'] = "Invalid zip file"
        except Exception as e:
            result['error'] = str(e)
        
        return result
    
    def change_view(self, request, object_id, form_url='', extra_context=None):
        """Add zip upload button to product change view"""
        extra_context = extra_context or {}
        extra_context['show_zip_upload'] = True
        extra_context['zip_upload_url'] = f'upload-zip/'
        return super().change_view(request, object_id, form_url, extra_context=extra_context)
   


@admin.register(ProductMedia)
class ProductMediaAdmin(admin.ModelAdmin):
    list_display = ['product', 'media_type', 'process', 'bom', 'file_preview', 'display_assignment', 'duration']
    list_filter = ['media_type', 'product', 'process__stage', 'display_screen_1', 'display_screen_2', 'display_screen_3']
    search_fields = ['product__code', 'product__name', 'file']
    readonly_fields = ['file_preview']

    def file_preview(self, obj):
        if obj.file:
            file_url = obj.file.url
            file_name = obj.file.name.lower()
            if file_name.endswith('.pdf'):
                return format_html('<a href="{}" target="_blank">📄 PDF</a>', file_url)
            elif file_name.endswith(('.mp4', '.mov', '.avi', '.mkv')):
                return format_html('<video width="100" height="60" controls><source src="{}" type="video/mp4"></video>', file_url)
            else:
                return format_html('<a href="{}">📁 File</a>', file_url)
        return "No file"
    file_preview.short_description = 'Preview'
    
    def display_assignment(self, obj):
        displays = obj.get_assigned_displays()
        if displays:
            display_badges = []
            for display in displays:
                color = ['red', 'blue', 'green'][display-1]
                display_badges.append(f'<span style="background-color: {color}; color: white; padding: 2px 6px; border-radius: 3px; margin: 1px;">D{display}</span>')
            return format_html(' '.join(display_badges))
        return format_html('<span style="color: gray;">None</span>')
    display_assignment.short_description = 'Displays'
    
class StationAdminForm(forms.ModelForm):
    class Meta:
        model = Station
        fields = '__all__'
        
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk and self.instance.current_product:
            # Filter current_stage and current_process based on selected product
            self.fields['current_stage'].queryset = AssemblyStage.objects.all()
            if self.instance.current_stage:
                self.fields['current_process'].queryset = self.instance.current_stage.processes.all()
            else:
                self.fields['current_process'].queryset = AssemblyProcess.objects.none()

@admin.register(Station)
class StationAdmin(admin.ModelAdmin):
    form = StationAdminForm
    list_display = ('name', 'display_number', 'current_product', 'current_stage', 'current_process', 
                   'product_quantity', 'display_status', 'clicker_status', 'bom_data_available')
    list_filter = ('display_number', 'clicker_enabled', 'loop_mode', 'current_stage')
    search_fields = ('name',)
    
    fieldsets = (
        ('Display Configuration', {
            'fields': ('name', 'display_number', 'manager', 'products')
        }),
        ('Current Assembly State', {
            'fields': ('current_product', 'current_stage', 'current_process', 'product_quantity'),
            'description': 'Set the current assembly status'
        }),
        ('BOM Display Settings', {
            'fields': ('show_single_unit_bom', 'show_batch_bom'),
            'description': 'Choose which BOMs to display'
        }),
        ('Control Settings', {
            'fields': ('clicker_enabled', 'auto_advance', 'loop_mode'),
            'description': 'Control and automation settings'
        }),
    )
    
    def display_status(self, obj):
        colors = ['#FF6B6B', '#4ECDC4', '#45B7D1']
        color = colors[obj.display_number - 1] if obj.display_number <= 3 else '#666'
        return format_html(
            '<span style="background-color: {}; color: white; padding: 4px 8px; border-radius: 4px; font-weight: bold;">Display {}</span>',
            color, obj.display_number
        )
    display_status.short_description = 'Display'
    
    def clicker_status(self, obj):
        if obj.clicker_enabled:
            loop_text = " 🔄" if obj.loop_mode else ""
            return format_html('<span style="color: green;">✓ Enabled{}</span>', loop_text)
        return format_html('<span style="color: red;">✗ Disabled</span>')
    clicker_status.short_description = 'Clicker'
    
    def bom_data_available(self, obj):
        bom_data = obj.get_current_bom_data()
        if bom_data:
            return format_html('<span style="color: green;">✓ {} items</span>', len(bom_data))
        return format_html('<span style="color: orange;">⚠ No BOM data</span>')
    bom_data_available.short_description = 'BOM Data'

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        station = self.get_object(request, object_id)
        if station:
            # Show current media for this display
            current_media = station.get_current_media()
            extra_context['current_media'] = current_media
            extra_context['next_process'] = station.get_next_process()
            extra_context['previous_process'] = station.get_previous_process()
            
            # Show BOM data
            bom_data = station.get_current_bom_data()
            extra_context['current_bom_data'] = bom_data
            
            # Show workflow information
            if station.current_product:
                workflow_info = {
                    'stages': AssemblyStage.objects.all().order_by('order'),
                    'current_stage_processes': station.current_stage.processes.all() if station.current_stage else None,
                }
                extra_context['workflow_info'] = workflow_info
                
        return super().change_view(request, object_id, form_url, extra_context=extra_context)

@admin.register(AssemblySession)
class AssemblySessionAdmin(admin.ModelAdmin):
    list_display = ('product', 'quantity', 'current_stage', 'current_process', 'start_time', 
                   'completed', 'display_stations')
    list_filter = ('completed', 'product', 'current_stage', 'start_time')
    readonly_fields = ('start_time',)
    date_hierarchy = 'start_time'
    
    fieldsets = (
        ('Session Info', {
            'fields': ('product', 'quantity', 'start_time', 'end_time', 'completed')
        }),
        ('Current Status', {
            'fields': ('current_stage', 'current_process')
        }),
        ('Display Stations', {
            'fields': ('display_1_station', 'display_2_station', 'display_3_station'),
            'description': 'Stations assigned to each display for this session'
        }),
    )
    
    def display_stations(self, obj):
        stations = []
        if obj.display_1_station:
            stations.append(f'D1: {obj.display_1_station.name}')
        if obj.display_2_station:
            stations.append(f'D2: {obj.display_2_station.name}')
        if obj.display_3_station:
            stations.append(f'D3: {obj.display_3_station.name}')
        return ' | '.join(stations) if stations else 'No stations assigned'
    display_stations.short_description = 'Assigned Stations'

# Custom admin actions
@admin.action(description='Enable loop mode for selected stations')
def enable_loop_mode(modeladmin, request, queryset):
    queryset.update(loop_mode=True)

@admin.action(description='Disable loop mode for selected stations')
def disable_loop_mode(modeladmin, request, queryset):
    queryset.update(loop_mode=False)

@admin.action(description='Clone BOM template with items')
def clone_bom_template(modeladmin, request, queryset):
    for template in queryset:
        # Create new template
        new_template = BOMTemplate.objects.create(
            product=template.product,
            bom_type=template.bom_type,
            stage=template.stage,
            template_name=f"{template.template_name} (Copy)",
            description=f"Copy of {template.template_name}",
            display_screen_1=template.display_screen_1,
            display_screen_2=template.display_screen_2,
            display_screen_3=template.display_screen_3,
            is_active=False  # Set as inactive by default
        )
        
        # Copy all items
        for item in template.bom_items.all():
            BOMTemplateItem.objects.create(
                bom_template=new_template,
                item=item.item,
                base_quantity=item.base_quantity,
                serial_number=item.serial_number,
                notes=item.notes,
                is_active=item.is_active
            )
    
    messages.success(request, f'Successfully cloned {queryset.count()} BOM template(s)')

StationAdmin.actions = [enable_loop_mode, disable_loop_mode]
BOMTemplateAdmin.actions = [clone_bom_template]

# Custom admin site title
admin.site.site_header = "Nipha Export Private Limited Admin"
admin.site.site_title = "Nipha Export Private Limited"
admin.site.index_title = "Welcome to Nipha Export Private Limited"
