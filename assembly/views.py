from django.shortcuts import render

# Create your views here.





from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.db.models import Prefetch, Count
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.core.paginator import Paginator
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
import json
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from PIL import Image
import io
import os
import zipfile
import tempfile
from pathlib import Path
from .models import (
    Product, ProductAssemblyProcess, ProductStage, Station, ProductMedia, BOMTemplate, BOMTemplateItem, 
    AssemblyProcess, AssemblyStage, BillOfMaterial, BOMItem
)


def get_product_stages_for_dropdown(request, product_id):
    """Get product-specific stages for dropdown (keeps your existing AJAX working)"""
    try:
        product = get_object_or_404(Product, id=product_id)
        stages = ProductStage.objects.filter(product=product, is_active=True).order_by('order')
        
        stage_data = [{
            'id': stage.id,
            'stage_code': stage.stage_code,
            'display_name': stage.display_name,
            'order': stage.order
        } for stage in stages]
        
        return JsonResponse({
            'success': True,
            'stages': stage_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def get_product_processes_for_dropdown(request, product_id, stage_id):
    """Get product-specific processes for dropdown"""
    try:
        product = get_object_or_404(Product, id=product_id)
        stage = get_object_or_404(ProductStage, id=stage_id, product=product)
        processes = ProductAssemblyProcess.objects.filter(
            product=product, 
            stage=stage, 
            is_active=True
        ).order_by('order')
        
        process_data = [{
            'id': process.id,
            'process_name': process.process_name,
            'display_name': process.display_name,
            'order': process.order
        } for process in processes]
        
        return JsonResponse({
            'success': True,
            'processes': process_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def product_information_view(request):
    """Main view to display product information page with CRUD operations"""
    products = Product.objects.all().order_by('code')
    selected_product = None
    product_data = {}
    
    # Get selected product from request
    product_id = request.GET.get('product_id')
    if product_id:
        try:
            selected_product = get_object_or_404(Product, id=product_id)
            product_data = get_product_data(selected_product)
        except Exception as e:
            messages.error(request, f"Error loading product data: {e}")
            product_data = {}
    
    # Get additional data for dropdowns
    assembly_stages = AssemblyStage.objects.all().order_by('order')
    assembly_processes = AssemblyProcess.objects.all().order_by('stage__order', 'order')
    bom_items = BOMItem.objects.filter(is_active=True).order_by('item_description')
    
    context = {
        'products': products,
        'selected_product': selected_product,
        'product_data': product_data,
        'assembly_stages': assembly_stages,
        'assembly_processes': assembly_processes,
        'bom_items': bom_items,
        'bom_type_choices': BOMTemplate.BOM_TYPE_CHOICES,
        'media_type_choices': ProductMedia.MEDIA_TYPE_CHOICES,
    }
    
    return render(request, 'assembly/product_information.html', context)

# Excel Upload for BOM Items
@require_http_methods(["POST"])  
def upload_bom_items_excel(request, template_id):
    """Upload BOM items from Excel file to a specific BOM template"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        
        if 'excel_file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No Excel file provided'})
        
        excel_file = request.FILES['excel_file']
        unit_of_measure = request.POST.get('unit_of_measure', 'NO.')
        supplier = request.POST.get('supplier', '')
        overwrite_existing = request.POST.get('overwrite_existing') == 'true'
        
        result = process_excel_file_for_template(
            excel_file, template, unit_of_measure, supplier, overwrite_existing
        )
        
        if result['success']:
            return JsonResponse({
                'success': True,
                'message': f"Successfully processed {result['created']} new items and {result['updated']} updated items. Skipped {result['skipped']} items.",
                'errors': result['errors'] if result['errors'] else None
            })
        else:
            return JsonResponse({'success': False, 'error': result['error']})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def process_excel_file_for_template(excel_file, bom_template, default_unit, default_supplier, overwrite_existing):
    """Process Excel file and create BOM items and template items"""
    result = {
        'success': False,
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': []
    }
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = workbook.active
        
        # Load images from the sheet
        image_loader = SheetImageLoader(sheet)
        
        # Find header row
        header_row = None
        for row_num in range(1, 6):
            row_values = [str(cell.value).strip().upper() if cell.value else '' for cell in sheet[row_num]]
            if any('ITEM DESCRIPTION' in val or 'DESCRIPTION' in val for val in row_values):
                header_row = row_num
                break
        
        if header_row is None:
            result['error'] = "Could not find header row with 'ITEM DESCRIPTION' column"
            return result
        
        # Map column indices
        headers = [str(cell.value).strip().upper() if cell.value else '' for cell in sheet[header_row]]
        
        col_mapping = {}
        for idx, header in enumerate(headers):
            if 'S.' in header and 'NO' in header:
                col_mapping['s_no'] = idx
            elif 'CODE' in header and 'ITEM' not in header:  # Just "Code" column, not "Item Code"
                col_mapping['code'] = idx
            elif 'ITEM DESCRIPTION' in header or 'DESCRIPTION' in header:
                col_mapping['description'] = idx
            elif 'PART NO' in header or 'PART_NO' in header:
                col_mapping['part_no'] = idx
            elif 'QTY' in header or 'QUANTITY' in header:
                col_mapping['qty'] = idx
            elif 'PHOTO' in header or 'IMAGE' in header:
                col_mapping['photo'] = idx
        
        # Validate required columns
        required_cols = ['description', 'part_no']
        missing_cols = [col for col in required_cols if col not in col_mapping]
        if missing_cols:
            result['error'] = f"Missing required columns: {missing_cols}"
            return result
        
        # Process data rows
        for row_num in range(header_row + 1, sheet.max_row + 1):
            row = sheet[row_num]
            
            # Skip empty rows
            if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                continue
            
            try:
                # Extract data
                s_no = str(row[col_mapping.get('s_no', 0)].value).strip() if col_mapping.get('s_no') is not None else str(row_num - header_row)
                code = str(row[col_mapping.get('code', 0)].value).strip() if col_mapping.get('code') is not None and row[col_mapping.get('code', 0)].value else ''
                description = str(row[col_mapping['description']].value).strip() if row[col_mapping['description']].value else ''
                part_no = str(row[col_mapping['part_no']].value).strip() if row[col_mapping['part_no']].value else ''
                qty = str(row[col_mapping.get('qty', 0)].value).strip() if col_mapping.get('qty') is not None and row[col_mapping.get('qty', 0)].value else '1'
                
                # Skip if essential data is missing
                if not description or not part_no:
                    result['skipped'] += 1
                    continue
                
                # Use the Code column if provided, otherwise generate from description
                if code:
                    # Use the provided code, clean it up
                    item_code = code.upper().replace(' ', '_').replace('-', '_')
                    item_code = ''.join(c for c in item_code if c.isalnum() or c == '_')
                else:
                    # Generate item_code from description as fallback
                    item_code = description.upper().replace(' ', '_').replace('-', '_')
                    item_code = ''.join(c for c in item_code if c.isalnum() or c == '_')
                
                # Ensure item_code is not empty
                if not item_code:
                    item_code = f"ITEM_{row_num - header_row}"
                
                # Check if BOM item exists
                bom_item = None
                item_created = False
                try:
                    bom_item = BOMItem.objects.get(item_code=item_code)
                    if overwrite_existing:
                        # Update existing item
                        bom_item.item_description = description
                        bom_item.part_number = part_no
                        bom_item.unit_of_measure = default_unit
                        if default_supplier:
                            bom_item.supplier = default_supplier
                        
                        # Handle image extraction for existing item
                        if 'photo' in col_mapping:
                            photo_cell = f"{openpyxl.utils.get_column_letter(col_mapping['photo'] + 1)}{row_num}"
                            if image_loader.image_in(photo_cell):
                                try:
                                    image = image_loader.get(photo_cell)
                                    if image:
                                        # Delete old image if exists
                                        if bom_item.item_photo:
                                            bom_item.item_photo.delete(save=False)
                                        
                                        img_io = io.BytesIO()
                                        if image.mode == 'RGBA':
                                            image.save(img_io, format='PNG', optimize=True)
                                            filename = f"{item_code.lower()}.png"
                                        else:
                                            if image.mode not in ['RGB', 'L']:
                                                image = image.convert('RGB')
                                            image.save(img_io, format='JPEG', quality=95, optimize=True)
                                            filename = f"{item_code.lower()}.jpg"
                                        
                                        img_io.seek(0)
                                        image_file = ContentFile(img_io.getvalue(), name=filename)
                                        bom_item.item_photo = image_file
                                except Exception as e:
                                    result['errors'].append(f"Row {row_num}: Could not extract image - {str(e)}")
                        
                        bom_item.save()
                        result['updated'] += 1
                    # If not overwriting, just use the existing item
                except BOMItem.DoesNotExist:
                    # Create new BOM item
                    bom_item_data = {
                        'item_code': item_code,
                        'item_description': description,
                        'part_number': part_no,
                        'unit_of_measure': default_unit,
                        'supplier': default_supplier or '',
                        'is_active': True
                    }
                    
                    # Handle image extraction for new item
                    if 'photo' in col_mapping:
                        photo_cell = f"{openpyxl.utils.get_column_letter(col_mapping['photo'] + 1)}{row_num}"
                        if image_loader.image_in(photo_cell):
                            try:
                                image = image_loader.get(photo_cell)
                                if image:
                                    img_io = io.BytesIO()
                                    if image.mode == 'RGBA':
                                        image.save(img_io, format='PNG', optimize=True)
                                        filename = f"{item_code.lower()}.png"
                                    else:
                                        if image.mode not in ['RGB', 'L']:
                                            image = image.convert('RGB')
                                        image.save(img_io, format='JPEG', quality=95, optimize=True)
                                        filename = f"{item_code.lower()}.jpg"
                                    
                                    img_io.seek(0)
                                    image_file = ContentFile(img_io.getvalue(), name=filename)
                                    bom_item_data['item_photo'] = image_file
                            except Exception as e:
                                result['errors'].append(f"Row {row_num}: Could not extract image - {str(e)}")
                    
                    bom_item = BOMItem.objects.create(**bom_item_data)
                    result['created'] += 1
                    item_created = True
                
                # Create or update BOM Template Item
                try:
                    serial_number = int(s_no) if s_no.isdigit() else row_num - header_row
                    
                    # Parse quantity - handle different formats
                    try:
                        base_quantity = float(qty) if qty.replace('.', '').replace(',', '').isdigit() else 1.0
                    except (ValueError, AttributeError):
                        base_quantity = 1.0
                    
                    # Check if template item already exists with this serial number
                    try:
                        template_item = BOMTemplateItem.objects.get(
                            bom_template=bom_template,
                            serial_number=serial_number
                        )
                        
                        if overwrite_existing:
                            template_item.item = bom_item
                            template_item.base_quantity = base_quantity
                            template_item.save()
                        # If not overwriting and item exists, skip
                        
                    except BOMTemplateItem.DoesNotExist:
                        # Create new template item
                        BOMTemplateItem.objects.create(
                            bom_template=bom_template,
                            item=bom_item,
                            serial_number=serial_number,
                            base_quantity=base_quantity,
                            is_active=True
                        )
                        
                except Exception as e:
                    result['errors'].append(f"Row {row_num}: Error creating template item - {str(e)}")
                    # If we created a BOM item but failed to create template item, we might want to keep the BOM item
                    # for potential future use, so we don't delete it here
                    
            except Exception as e:
                result['errors'].append(f"Row {row_num}: {str(e)}")
                continue
        
        result['success'] = True
        
    except Exception as e:
        result['error'] = str(e)
    
    return result 

 
# ZIP Upload for Product Media
@require_http_methods(["POST"])
def upload_product_media_zip(request):
    """Upload product media from ZIP file"""
    try:
        product_id = request.POST.get('product_id')
        product = get_object_or_404(Product, id=product_id)
        
        if 'zip_file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No ZIP file provided'})
        
        zip_file = request.FILES['zip_file']
        
        # Get form data
        process_id = request.POST.get('process_id')
        process = get_object_or_404(AssemblyProcess, id=process_id) if process_id else None
        
        bom_id = request.POST.get('bom_id')
        bom = get_object_or_404(BillOfMaterial, id=bom_id) if bom_id else None
        
        default_duration = int(request.POST.get('default_duration', 30))
        display_screen_1 = request.POST.get('display_screen_1') == 'true'
        display_screen_2 = request.POST.get('display_screen_2') == 'true'
        display_screen_3 = request.POST.get('display_screen_3') == 'true'
        
        result = process_zip_file_for_media(
            product, zip_file, process, bom, default_duration,
            display_screen_1, display_screen_2, display_screen_3
        )
        
        if result['success']:
            return JsonResponse({
                'success': True,
                'message': f"Successfully processed {result['total_files']} files: {result['pdfs']} PDFs, {result['videos']} videos. Skipped {result['skipped']} files.",
                'errors': result['errors'] if result['errors'] else None
            })
        else:
            return JsonResponse({'success': False, 'error': result['error']})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def process_zip_file_for_media(product, zip_file, process, bom, default_duration, 
                              display_screen_1, display_screen_2, display_screen_3):
    """Process uploaded ZIP file and create ProductMedia objects"""
    result = {
        'success': False,
        'total_files': 0,
        'pdfs': 0,
        'videos': 0,
        'skipped': 0,
        'errors': []
    }
    
    # Supported file extensions
    PDF_EXTENSIONS = {'.pdf'}
    VIDEO_EXTENSIONS = {'.mp4', '.mov', '.avi', '.mkv', '.wmv', '.flv', '.webm'}
    
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            # Extract zip file
            with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # Process all files in the extracted directory
            for root, dirs, files in os.walk(temp_dir):
                for filename in files:
                    file_path = os.path.join(root, filename)
                    file_ext = Path(filename).suffix.lower()
                    
                    # Skip hidden files and system files
                    if filename.startswith('.') or filename.startswith('__MACOSX'):
                        continue
                    
                    try:
                        # Determine media type based on extension
                        media_type = None
                        duration = None
                        
                        if file_ext in PDF_EXTENSIONS:
                            media_type = 'PROCESS_DOC'
                            result['pdfs'] += 1
                        elif file_ext in VIDEO_EXTENSIONS:
                            media_type = 'VIDEO'
                            duration = default_duration
                            result['videos'] += 1
                        else:
                            result['skipped'] += 1
                            continue
                        
                        # Read file content
                        with open(file_path, 'rb') as f:
                            file_content = f.read()
                        
                        # Create ContentFile
                        django_file = ContentFile(file_content, name=filename)
                        
                        # Create ProductMedia object
                        media_data = {
                            'product': product,
                            'media_type': media_type,
                            'file': django_file,
                            'display_screen_1': display_screen_1,
                            'display_screen_2': display_screen_2,
                            'display_screen_3': display_screen_3,
                        }
                        
                        # Add optional fields
                        if process:
                            media_data['process'] = process
                        if bom:
                            media_data['bom'] = bom
                        if duration:
                            media_data['duration'] = duration
                        
                        # Create the ProductMedia object
                        ProductMedia.objects.create(**media_data)
                        result['total_files'] += 1
                        
                    except Exception as e:
                        result['errors'].append(f"{filename}: {str(e)}")
                        continue
        
        result['success'] = True
        
    except zipfile.BadZipFile:
        result['error'] = "Invalid ZIP file"
    except Exception as e:
        result['error'] = str(e)
    
    return result

# BOM Template CRUD Operations
@require_http_methods(["POST"])
def create_bom_template(request):
    """Create a new BOM template"""
    try:
        product_id = request.POST.get('product_id')
        product = get_object_or_404(Product, id=product_id)
        
        # Get stage if provided
        stage_id = request.POST.get('stage_id')
        stage = get_object_or_404(AssemblyStage, id=stage_id) if stage_id else None
        
        bom_template = BOMTemplate.objects.create(
            product=product,
            bom_type=request.POST.get('bom_type'),
            stage=stage,
            template_name=request.POST.get('template_name'),
            description=request.POST.get('description', ''),
            duration=int(request.POST.get('duration', 20)),
            is_duration_active=request.POST.get('is_duration_active') == 'on',
            display_screen_1=request.POST.get('display_screen_1') == 'on',
            display_screen_2=request.POST.get('display_screen_2') == 'on',
            display_screen_3=request.POST.get('display_screen_3') == 'on',
        )
        
        messages.success(request, f'BOM Template "{bom_template.template_name}" created successfully!')
        return JsonResponse({
            'success': True, 
            'message': 'BOM Template created successfully!',
            'template_id': bom_template.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def update_bom_template(request, template_id):
    """Update an existing BOM template"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        
        # Get stage if provided
        stage_id = request.POST.get('stage_id')
        stage = get_object_or_404(AssemblyStage, id=stage_id) if stage_id else None
        
        template.bom_type = request.POST.get('bom_type')
        template.stage = stage
        template.template_name = request.POST.get('template_name')
        template.description = request.POST.get('description', '')
        template.duration = int(request.POST.get('duration', 20))
        template.is_duration_active = request.POST.get('is_duration_active') == 'on'
        template.display_screen_1 = request.POST.get('display_screen_1') == 'on'
        template.display_screen_2 = request.POST.get('display_screen_2') == 'on'
        template.display_screen_3 = request.POST.get('display_screen_3') == 'on'
        template.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'BOM Template updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_bom_template(request, template_id):
    """Delete a BOM template"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        template_name = template.template_name
        template.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'BOM Template "{template_name}" deleted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# BOM Template Item CRUD Operations
@require_http_methods(["POST"])
def create_bom_template_item(request):
    """Create a new BOM template item"""
    try:
        template_id = request.POST.get('template_id')
        template = get_object_or_404(BOMTemplate, id=template_id)
        
        item_id = request.POST.get('item_id')
        item = get_object_or_404(BOMItem, id=item_id)
        
        bom_item = BOMTemplateItem.objects.create(
            bom_template=template,
            item=item,
            base_quantity=float(request.POST.get('base_quantity', 1)),
            serial_number=int(request.POST.get('serial_number')),
            notes=request.POST.get('notes', ''),
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'BOM item added successfully!',
            'item_id': bom_item.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def update_bom_template_item(request, item_id):
    """Update an existing BOM template item"""
    try:
        bom_item = get_object_or_404(BOMTemplateItem, id=item_id)
        
        item_id = request.POST.get('item_id')
        item = get_object_or_404(BOMItem, id=item_id)
        
        bom_item.item = item
        bom_item.base_quantity = float(request.POST.get('base_quantity', 1))
        bom_item.serial_number = int(request.POST.get('serial_number'))
        bom_item.notes = request.POST.get('notes', '')
        bom_item.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'BOM item updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_bom_template_item(request, item_id):
    """Delete a BOM template item"""
    try:
        bom_item = get_object_or_404(BOMTemplateItem, id=item_id)
        item_description = bom_item.item.item_description
        bom_item.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'BOM item "{item_description}" removed successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Product Media CRUD Operations
@require_http_methods(["POST"])
def create_product_media(request):
    """Create a new product media"""
    try:
        product_id = request.POST.get('product_id')
        product = get_object_or_404(Product, id=product_id)
        
        # Get process if provided
        process_id = request.POST.get('process_id')
        process = get_object_or_404(AssemblyProcess, id=process_id) if process_id else None
        
        # Get BOM if provided
        bom_id = request.POST.get('bom_id')
        bom = get_object_or_404(BillOfMaterial, id=bom_id) if bom_id else None
        
        media = ProductMedia.objects.create(
            product=product,
            process=process,
            bom=bom,
            media_type=request.POST.get('media_type'),
            file=request.FILES.get('file'),
            duration=int(request.POST.get('duration', 15)),
            display_screen_1=request.POST.get('display_screen_1') == 'on',
            display_screen_2=request.POST.get('display_screen_2') == 'on',
            display_screen_3=request.POST.get('display_screen_3') == 'on',
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Product media created successfully!',
            'media_id': media.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def update_product_media(request, media_id):
    """Update an existing product media"""
    try:
        media = get_object_or_404(ProductMedia, id=media_id)
        
        # Get process if provided
        process_id = request.POST.get('process_id')
        process = get_object_or_404(AssemblyProcess, id=process_id) if process_id else None
        
        # Get BOM if provided
        bom_id = request.POST.get('bom_id')
        bom = get_object_or_404(BillOfMaterial, id=bom_id) if bom_id else None
        
        media.process = process
        media.bom = bom
        media.media_type = request.POST.get('media_type')
        media.duration = int(request.POST.get('duration', 15))
        media.display_screen_1 = request.POST.get('display_screen_1') == 'on'
        media.display_screen_2 = request.POST.get('display_screen_2') == 'on'
        media.display_screen_3 = request.POST.get('display_screen_3') == 'on'
        
        # Update file if provided
        if request.FILES.get('file'):
            media.file = request.FILES.get('file')
            
        media.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Product media updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_product_media(request, media_id):
    """Delete a product media"""
    try:
        media = get_object_or_404(ProductMedia, id=media_id)
        media_name = str(media)
        media.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Product media "{media_name}" deleted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Assembly Process CRUD Operations
@require_http_methods(["POST"])
def create_assembly_process(request):
    """Create a new assembly process"""
    try:
        stage_id = request.POST.get('stage_id')
        stage = get_object_or_404(AssemblyStage, id=stage_id)
        
        process = AssemblyProcess.objects.create(
            stage=stage,
            name=request.POST.get('name'),
            display_name=request.POST.get('display_name', ''),
            location=request.POST.get('location', ''),
            order=int(request.POST.get('order', 1)),
            is_looped=request.POST.get('is_looped') == 'on',
            loop_group=request.POST.get('loop_group', ''),
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Assembly process created successfully!',
            'process_id': process.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def update_assembly_process(request, process_id):
    """Update an existing assembly process"""
    try:
        process = get_object_or_404(AssemblyProcess, id=process_id)
        
        stage_id = request.POST.get('stage_id')
        stage = get_object_or_404(AssemblyStage, id=stage_id)
        
        process.stage = stage
        process.name = request.POST.get('name')
        process.display_name = request.POST.get('display_name', '')
        process.location = request.POST.get('location', '')
        process.order = int(request.POST.get('order', 1))
        process.is_looped = request.POST.get('is_looped') == 'on'
        process.loop_group = request.POST.get('loop_group', '')
        process.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Assembly process updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_assembly_process(request, process_id):
    """Delete an assembly process"""
    try:
        process = get_object_or_404(AssemblyProcess, id=process_id)
        process_name = process.name
        process.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Assembly process "{process_name}" deleted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Utility function to get available serial numbers for BOM items
def get_available_serial_numbers(request, template_id):
    """Get next available serial number for BOM template"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        used_serials = list(template.bom_items.values_list('serial_number', flat=True))
        
        # Find next available serial number
        next_serial = 1
        while next_serial in used_serials:
            next_serial += 1
            
        return JsonResponse({
            'success': True,
            'next_serial': next_serial,
            'used_serials': used_serials
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Get processes for a specific stage (AJAX helper)
def get_processes_for_stage(request, stage_id):
    """Get processes for a specific stage"""
    try:
        stage = get_object_or_404(AssemblyStage, id=stage_id)
        processes = stage.processes.all().order_by('order')
        
        process_data = [{
            'id': process.id,
            'name': process.name,
            'display_name': process.display_name,
            'order': process.order
        } for process in processes]
        
        return JsonResponse({
            'success': True,
            'processes': process_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# BOM Template CRUD Operations
@require_http_methods(["POST"])
def create_bom_template(request):
    """Create a new BOM template"""
    try:
        product_id = request.POST.get('product_id')
        product = get_object_or_404(Product, id=product_id)
        
        # Get stage if provided
        stage_id = request.POST.get('stage_id')
        stage = get_object_or_404(AssemblyStage, id=stage_id) if stage_id else None
        
        bom_template = BOMTemplate.objects.create(
            product=product,
            bom_type=request.POST.get('bom_type'),
            stage=stage,
            template_name=request.POST.get('template_name'),
            description=request.POST.get('description', ''),
            duration=int(request.POST.get('duration', 20)),
            is_duration_active=request.POST.get('is_duration_active') == 'on',
            display_screen_1=request.POST.get('display_screen_1') == 'on',
            display_screen_2=request.POST.get('display_screen_2') == 'on',
            display_screen_3=request.POST.get('display_screen_3') == 'on',
        )
        
        messages.success(request, f'BOM Template "{bom_template.template_name}" created successfully!')
        return JsonResponse({
            'success': True, 
            'message': 'BOM Template created successfully!',
            'template_id': bom_template.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def update_bom_template(request, template_id):
    """Update an existing BOM template"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        
        # Get stage if provided
        stage_id = request.POST.get('stage_id')
        stage = get_object_or_404(AssemblyStage, id=stage_id) if stage_id else None
        
        template.bom_type = request.POST.get('bom_type')
        template.stage = stage
        template.template_name = request.POST.get('template_name')
        template.description = request.POST.get('description', '')
        template.duration = int(request.POST.get('duration', 20))
        template.is_duration_active = request.POST.get('is_duration_active') == 'on'
        template.display_screen_1 = request.POST.get('display_screen_1') == 'on'
        template.display_screen_2 = request.POST.get('display_screen_2') == 'on'
        template.display_screen_3 = request.POST.get('display_screen_3') == 'on'
        template.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'BOM Template updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_bom_template(request, template_id):
    """Delete a BOM template"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        template_name = template.template_name
        template.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'BOM Template "{template_name}" deleted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# BOM Template Item CRUD Operations
@require_http_methods(["POST"])
def create_bom_template_item(request):
    """Create a new BOM template item"""
    try:
        template_id = request.POST.get('template_id')
        template = get_object_or_404(BOMTemplate, id=template_id)
        
        item_id = request.POST.get('item_id')
        item = get_object_or_404(BOMItem, id=item_id)
        
        bom_item = BOMTemplateItem.objects.create(
            bom_template=template,
            item=item,
            base_quantity=float(request.POST.get('base_quantity', 1)),
            serial_number=int(request.POST.get('serial_number')),
            notes=request.POST.get('notes', ''),
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'BOM item added successfully!',
            'item_id': bom_item.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def update_bom_template_item(request, item_id):
    """Update an existing BOM template item"""
    try:
        bom_item = get_object_or_404(BOMTemplateItem, id=item_id)
        
        item_id = request.POST.get('item_id')
        item = get_object_or_404(BOMItem, id=item_id)
        
        bom_item.item = item
        bom_item.base_quantity = float(request.POST.get('base_quantity', 1))
        bom_item.serial_number = int(request.POST.get('serial_number'))
        bom_item.notes = request.POST.get('notes', '')
        bom_item.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'BOM item updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_bom_template_item(request, item_id):
    """Delete a BOM template item"""
    try:
        bom_item = get_object_or_404(BOMTemplateItem, id=item_id)
        item_description = bom_item.item.item_description
        bom_item.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'BOM item "{item_description}" removed successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Product Media CRUD Operations
@require_http_methods(["POST"])
def create_product_media(request):
    """Create a new product media"""
    try:
        product_id = request.POST.get('product_id')
        product = get_object_or_404(Product, id=product_id)
        
        # Get process if provided
        process_id = request.POST.get('process_id')
        process = get_object_or_404(AssemblyProcess, id=process_id) if process_id else None
        
        # Get BOM if provided
        bom_id = request.POST.get('bom_id')
        bom = get_object_or_404(BillOfMaterial, id=bom_id) if bom_id else None
        
        media = ProductMedia.objects.create(
            product=product,
            process=process,
            bom=bom,
            media_type=request.POST.get('media_type'),
            file=request.FILES.get('file'),
            duration=int(request.POST.get('duration', 15)),
            display_screen_1=request.POST.get('display_screen_1') == 'on',
            display_screen_2=request.POST.get('display_screen_2') == 'on',
            display_screen_3=request.POST.get('display_screen_3') == 'on',
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Product media created successfully!',
            'media_id': media.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def update_product_media(request, media_id):
    """Update an existing product media"""
    try:
        media = get_object_or_404(ProductMedia, id=media_id)
        
        # Get process if provided
        process_id = request.POST.get('process_id')
        process = get_object_or_404(AssemblyProcess, id=process_id) if process_id else None
        
        # Get BOM if provided
        bom_id = request.POST.get('bom_id')
        bom = get_object_or_404(BillOfMaterial, id=bom_id) if bom_id else None
        
        media.process = process
        media.bom = bom
        media.media_type = request.POST.get('media_type')
        media.duration = int(request.POST.get('duration', 15))
        media.display_screen_1 = request.POST.get('display_screen_1') == 'on'
        media.display_screen_2 = request.POST.get('display_screen_2') == 'on'
        media.display_screen_3 = request.POST.get('display_screen_3') == 'on'
        
        # Update file if provided
        if request.FILES.get('file'):
            media.file = request.FILES.get('file')
            
        media.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Product media updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_product_media(request, media_id):
    """Delete a product media"""
    try:
        media = get_object_or_404(ProductMedia, id=media_id)
        media_name = str(media)
        media.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Product media "{media_name}" deleted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Assembly Process CRUD Operations
@require_http_methods(["POST"])
def create_assembly_process(request):
    """Create a new assembly process"""
    try:
        stage_id = request.POST.get('stage_id')
        stage = get_object_or_404(AssemblyStage, id=stage_id)
        
        process = AssemblyProcess.objects.create(
            stage=stage,
            name=request.POST.get('name'),
            display_name=request.POST.get('display_name', ''),
            location=request.POST.get('location', ''),
            order=int(request.POST.get('order', 1)),
            is_looped=request.POST.get('is_looped') == 'on',
            loop_group=request.POST.get('loop_group', ''),
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Assembly process created successfully!',
            'process_id': process.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def update_assembly_process(request, process_id):
    """Update an existing assembly process"""
    try:
        process = get_object_or_404(AssemblyProcess, id=process_id)
        
        stage_id = request.POST.get('stage_id')
        stage = get_object_or_404(AssemblyStage, id=stage_id)
        
        process.stage = stage
        process.name = request.POST.get('name')
        process.display_name = request.POST.get('display_name', '')
        process.location = request.POST.get('location', '')
        process.order = int(request.POST.get('order', 1))
        process.is_looped = request.POST.get('is_looped') == 'on'
        process.loop_group = request.POST.get('loop_group', '')
        process.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Assembly process updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_assembly_process(request, process_id):
    """Delete an assembly process"""
    try:
        process = get_object_or_404(AssemblyProcess, id=process_id)
        process_name = process.name
        process.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Assembly process "{process_name}" deleted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Utility function to get available serial numbers for BOM items
def get_available_serial_numbers(request, template_id):
    """Get next available serial number for BOM template"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        used_serials = list(template.bom_items.values_list('serial_number', flat=True))
        
        # Find next available serial number
        next_serial = 1
        while next_serial in used_serials:
            next_serial += 1
            
        return JsonResponse({
            'success': True,
            'next_serial': next_serial,
            'used_serials': used_serials
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Get processes for a specific stage (AJAX helper)
def get_processes_for_stage(request, stage_id):
    """Get processes for a specific stage"""
    try:
        stage = get_object_or_404(AssemblyStage, id=stage_id)
        processes = stage.processes.all().order_by('order')
        
        process_data = [{
            'id': process.id,
            'name': process.name,
            'display_name': process.display_name,
            'order': process.order
        } for process in processes]
        
        return JsonResponse({
            'success': True,
            'processes': process_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Helper views for getting individual record data
def get_bom_template_data(request, template_id):
    """Get BOM template data for editing"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        template_data = {
            'id': template.id,
            'template_name': template.template_name,
            'bom_type': template.bom_type,
            'stage_id': template.stage.id if template.stage else None,
            'description': template.description,
            'duration': template.duration,
            'is_duration_active': template.is_duration_active,
            'display_screen_1': template.display_screen_1,
            'display_screen_2': template.display_screen_2,
            'display_screen_3': template.display_screen_3,
        }
        return JsonResponse({'success': True, 'template': template_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_bom_template_items(request, template_id):
    """Get BOM template items with images"""
    try:
        template = get_object_or_404(BOMTemplate, id=template_id)
        items = template.bom_items.filter(is_active=True).order_by('serial_number')
        
        items_data = [{
            'id': item.id,
            'serial_number': item.serial_number,
            'item_code': item.item.item_code,
            'item_description': item.item.item_description,
            'part_number': item.item.part_number,
            'base_quantity': float(item.base_quantity),
            'unit_of_measure': item.item.unit_of_measure,
            'notes': item.notes,
            'item_photo_url': item.item.item_photo.url if item.item.item_photo else None,
            'has_photo': bool(item.item.item_photo)
        } for item in items]
        
        return JsonResponse({'success': True, 'items': items_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def get_bom_template_item_data(request, item_id):
    """Get BOM template item data for editing"""
    try:
        item = get_object_or_404(BOMTemplateItem, id=item_id)
        item_data = {
            'id': item.id,
            'item_id': item.item.id,
            'serial_number': item.serial_number,
            'base_quantity': float(item.base_quantity),
            'notes': item.notes
        }
        return JsonResponse({'success': True, 'item': item_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_product_media_data(request, media_id):
    """Get product media data for editing with file preview info"""
    try:
        media = get_object_or_404(ProductMedia, id=media_id)
        
        # Determine file type and preview info
        file_url = None
        file_type = None
        file_size = None
        
        if media.file:
            file_url = media.file.url
            file_name = media.file.name.lower()
            file_size = media.file.size if hasattr(media.file, 'size') else None
            
            # Determine file type
            if file_name.endswith('.pdf'):
                file_type = 'pdf'
            elif file_name.endswith(('.mp4', '.mov', '.avi', '.mkv', '.wmv', '.flv', '.webm')):
                file_type = 'video'
            elif file_name.endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp')):
                file_type = 'image'
            elif file_name.endswith(('.xlsx', '.xls')):
                file_type = 'excel'
            elif file_name.endswith(('.docx', '.doc')):
                file_type = 'document'
            else:
                file_type = 'other'
        
        media_data = {
            'id': media.id,
            'media_type': media.media_type,
            'duration': media.duration,
            'process_id': media.process.id if media.process else None,
            'bom_id': media.bom.id if media.bom else None,
            'display_screen_1': media.display_screen_1,
            'display_screen_2': media.display_screen_2,
            'display_screen_3': media.display_screen_3,
            'file_name': media.file.name if media.file else None,
            'file_url': file_url,
            'file_type': file_type,
            'file_size': file_size,
            'has_file': bool(media.file)
        }
        return JsonResponse({'success': True, 'media': media_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_assembly_process_data(request, process_id):
    """Get assembly process data for editing"""
    try:
        process = get_object_or_404(AssemblyProcess, id=process_id)
        process_data = {
            'id': process.id,
            'stage_id': process.stage.id if process.stage else None,
            'name': process.name,
            'display_name': process.display_name,
            'location': process.location,
            'order': process.order,
            'is_looped': process.is_looped,
            'loop_group': process.loop_group
        }
        return JsonResponse({'success': True, 'process': process_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_product_data(product):
    """Get all data related to a product (unchanged from previous version)"""
    # ... (same as before)
    
    # 1. Connected Stations
    active_stations = Station.objects.filter(
        current_product=product
    ).select_related(
        'current_stage', 'current_process', 'manager'
    ).order_by('display_number')
    
    # Also get stations that have this product in their products many-to-many
    available_stations = Station.objects.filter(
        products=product
    ).select_related('manager').order_by('display_number')
    
    # 2. Product Media
    media_files = ProductMedia.objects.filter(
        product=product
    ).select_related('process', 'bom').order_by('media_type', 'id')
    
    # 3. BOM Templates with Items
    bom_templates = BOMTemplate.objects.filter(
        product=product,
        is_active=True
    ).prefetch_related(
        Prefetch(
            'bom_items', 
            queryset=BOMTemplateItem.objects.filter(is_active=True)
                .select_related('item')
                .order_by('serial_number')
        )
    ).select_related('stage').order_by('bom_type')
    
    # 4. Assembly Stages and Processes
    assembly_stages = AssemblyStage.objects.prefetch_related(
        Prefetch(
            'processes', 
            queryset=AssemblyProcess.objects.order_by('order')
        )
    ).order_by('order')
    
    # 5. Bill of Materials (PDF/Legacy)
    legacy_boms = BillOfMaterial.objects.filter(
        product=product
    ).select_related('stage', 'bom_template').order_by('bom_type')
    
    # Group media by type for easier display
    media_by_type = {}
    for media_item in media_files:
        media_type = media_item.get_media_type_display()
        if media_type not in media_by_type:
            media_by_type[media_type] = []
        media_by_type[media_type].append(media_item)
    
    # Group BOM templates by type
    templates_by_type = {}
    for template in bom_templates:
        template_type = template.get_bom_type_display()
        if template_type not in templates_by_type:
            templates_by_type[template_type] = []
        templates_by_type[template_type].append(template)
    
    # Prepare context data
    context_data = {
        'product': product,
        'active_stations': active_stations,
        'available_stations': available_stations,
        'media_files': media_files,
        'bom_templates': bom_templates,
        'assembly_stages': assembly_stages,
        'legacy_boms': legacy_boms,
        'media_by_type': media_by_type,
        'templates_by_type': templates_by_type,
        'station_summary': {
            'total_active': active_stations.count(),
            'total_available': available_stations.count(),
            'displays_in_use': list(active_stations.values_list('display_number', flat=True)),
        }
    }
    
    return context_data


def get_product_data_updated(product):
    """Updated version of your get_product_data function with product-specific stages"""
    
    # Your existing code stays the same, just add this at the end:
    
    # Get your existing data
    context_data = get_product_data(product)  # Your existing function
    
    # Add product-specific stages and processes
    product_stages = ProductStage.objects.filter(
        product=product,
        is_active=True
    ).prefetch_related(
        'product_processes'
    ).order_by('order')
    
    # Add to context
    context_data['product_stages'] = product_stages
    context_data['has_product_stages'] = product_stages.exists()
    
    return context_data
